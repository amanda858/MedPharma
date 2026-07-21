"""Batch coverage + medical-necessity + prior-auth review.

Upload a spreadsheet of patients (Excel or CSV) and get one reviewed row per
ordered test back: disposition, coverage status, prior-auth, patient $, expected
value, and the exact action a biller must take. This is the "review all at once"
workflow a lab runs every morning against the day's accessions.
"""
from __future__ import annotations

import csv
from datetime import date as _date, datetime
from typing import Iterable, Optional

from .gate import AccessionGate
from .models import PatientRequest

# Input column aliases (case/space-insensitive) -> canonical field.
_COLUMN_ALIASES = {
    "first_name": ["first", "firstname", "first name", "patient first", "fname"],
    "last_name": ["last", "lastname", "last name", "patient last", "lname"],
    "dob": ["dob", "date of birth", "birthdate", "birth date"],
    "gender": ["gender", "sex"],
    "member_id": ["member id", "memberid", "member", "policy id", "policy",
                  "subscriber id", "insurance id", "insurance #", "policy number"],
    "payer_name": ["payer", "payer name", "insurance", "carrier", "plan", "insurance company"],
    "payer_id": ["payer id", "payerid", "payer code"],
    "ssn_last4": ["ssn", "ssn last4", "ssn last 4", "ssn4"],
    "zip_code": ["zip", "zip code", "zipcode", "postal", "postal code"],
    "date_of_service": ["dos", "date of service", "service date", "collection date"],
    "cpt_codes": ["cpt", "cpts", "cpt codes", "tests", "test codes", "procedure", "procedures"],
    "icd10_codes": ["icd", "icd10", "icd-10", "dx", "diagnosis", "diagnosis codes", "dx codes"],
    "provider_npi": ["npi", "provider npi", "rendering npi", "ordering npi"],
    "provider_name": ["provider", "provider name", "ordering provider", "physician"],
}

REVIEW_COLUMNS = [
    "Patient", "DOB", "Payer", "Member ID", "CPT", "Test",
    "Disposition", "Coverage", "Medically Necessary", "Prior Auth", "Auth #",
    "Patient $", "Plan $", "Expected $", "Action", "Reasons",
]

_DISPOSITION_FILL = {
    "CLEAR TO RUN": "C6EFCE",
    "HOLD — PRIOR AUTH": "FFEB9C",
    "HOLD — MEDICAL NECESSITY": "FFC7CE",
    "GET ABN": "FFD966",
    "SELF-PAY": "D9D9D9",
    "DENY RISK": "FF9999",
}


def _canon_header(h: str) -> Optional[str]:
    s = (h or "").strip().lower()
    for field, aliases in _COLUMN_ALIASES.items():
        if s == field or s in aliases:
            return field
    return None


def _clean(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _split_codes(v) -> list[str]:
    if v is None:
        return []
    s = str(v)
    for sep in (";", "|", "/", ",", " "):
        s = s.replace(sep, ",")
    return [c.strip().upper() for c in s.split(",") if c.strip()]


def _norm_date(v) -> str:
    if v in (None, ""):
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, _date):
        return v.isoformat()
    s = str(v).strip()
    if not s:
        return ""
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%Y/%m/%d", "%d-%b-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s


def row_to_request(row: dict) -> PatientRequest:
    g: dict = {}
    for k, v in row.items():
        field = _canon_header(k)
        if field and (field not in g or _clean(v)):
            g[field] = v
    return PatientRequest(
        first_name=str(g.get("first_name", "") or "").strip(),
        last_name=str(g.get("last_name", "") or "").strip(),
        dob=_norm_date(g.get("dob")),
        gender=(str(g.get("gender", "U") or "U").strip()[:1].upper() or "U"),
        member_id=_clean(g.get("member_id")),
        payer_name=_clean(g.get("payer_name")),
        payer_id=_clean(g.get("payer_id")),
        ssn_last4=_clean(g.get("ssn_last4")),
        zip_code=_clean(g.get("zip_code")),
        date_of_service=_norm_date(g.get("date_of_service")) or None,
        cpt_codes=_split_codes(g.get("cpt_codes")),
        icd10_codes=_split_codes(g.get("icd10_codes")),
        provider_npi=str(g.get("provider_npi", "") or "").strip(),
        provider_name=str(g.get("provider_name", "") or "").strip(),
    )


def build_review_gate(allow_live: bool | None = None) -> AccessionGate:
    from .config import build_default_engine
    from .prior_auth import PriorAuthEngine
    engine = build_default_engine(allow_live=allow_live)
    return AccessionGate(engine, PriorAuthEngine(sandbox=engine.pverify.sandbox),
                         auto_submit_pa=True)


def rows_from_accession(req: PatientRequest, res) -> list[dict]:
    """Turn an already-evaluated AccessionResult into reviewed spreadsheet rows.

    Split out from review_request so a caller that already holds the full
    AccessionResult (e.g. a single-patient API view) can reuse the exact same
    row shape without evaluating the gate twice.
    """
    payer = res.coverage.payer_name + (" (discovered)" if res.coverage.discovered else "")
    rows: list[dict] = []
    for ln in res.lines:
        rows.append({
            "Patient": req.full_name, "DOB": req.dob, "Payer": payer,
            "Member ID": res.coverage.member_id, "CPT": ln.cpt, "Test": ln.description,
            "Disposition": ln.disposition.value, "Coverage": ln.coverage_status,
            "Medically Necessary": "Yes" if ln.medically_necessary else "No",
            "Prior Auth": ln.prior_auth.value, "Auth #": ln.auth_number,
            "Patient $": ln.patient_responsibility, "Plan $": ln.plan_pays,
            "Expected $": ln.expected_value, "Action": "; ".join(ln.actions),
            "Reasons": " | ".join(ln.reasons),
        })
    if not rows:  # no CPTs ordered — still report coverage
        rows.append({
            "Patient": req.full_name, "DOB": req.dob, "Payer": payer,
            "Member ID": res.coverage.member_id, "CPT": "", "Test": "(no tests listed)",
            "Disposition": res.overall.value, "Coverage": res.coverage.status.value,
            "Medically Necessary": "", "Prior Auth": "", "Auth #": "",
            "Patient $": "", "Plan $": "", "Expected $": "",
            "Action": "Add ordered CPT(s) to evaluate.", "Reasons": "",
        })
    return rows


def review_request(req: PatientRequest, gate: AccessionGate) -> list[dict]:
    return rows_from_accession(req, gate.evaluate(req))


def review_batch(requests: Iterable[PatientRequest],
                 gate: Optional[AccessionGate] = None) -> list[dict]:
    gate = gate or build_review_gate()
    out: list[dict] = []
    for req in requests:
        out.extend(review_request(req, gate))
    return out


def review_rows(rows: Iterable[dict], gate: Optional[AccessionGate] = None) -> list[dict]:
    return review_batch((row_to_request(r) for r in rows), gate)


# ── file I/O (Excel + CSV) ──────────────────────────────────────────────────
def read_rows(path: str) -> list[dict]:
    if path.lower().endswith((".xlsx", ".xlsm", ".xls")):
        return _read_xlsx(path)
    return _read_csv(path)


def _read_csv(path: str) -> list[dict]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def _read_xlsx(path: str) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    try:
        headers = [str(h or "").strip() for h in next(it)]
    except StopIteration:
        return []
    out: list[dict] = []
    for r in it:
        if r is None or all(c is None or str(c).strip() == "" for c in r):
            continue
        out.append({headers[i]: r[i] for i in range(min(len(headers), len(r)))})
    return out


def write_review_csv(path: str, rows: list[dict]) -> str:
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=REVIEW_COLUMNS)
        w.writeheader()
        for row in rows:
            w.writerow({c: row.get(c, "") for c in REVIEW_COLUMNS})
    return path


def write_review_xlsx(path: str, rows: list[dict]) -> str:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Coverage Review"
    ws.append(REVIEW_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="374151")
    disp_col = REVIEW_COLUMNS.index("Disposition") + 1
    for row in rows:
        ws.append([row.get(c, "") for c in REVIEW_COLUMNS])
        hexc = _DISPOSITION_FILL.get(str(row.get("Disposition", "")))
        if hexc:
            ws.cell(row=ws.max_row, column=disp_col).fill = PatternFill("solid", fgColor=hexc)
    widths = [18, 12, 22, 14, 8, 34, 24, 10, 10, 12, 12, 10, 9, 10, 40, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    wb.save(path)
    return path


def summarize(rows: list[dict]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in rows:
        d = str(row.get("Disposition", ""))
        counts[d] = counts.get(d, 0) + 1
    return counts
