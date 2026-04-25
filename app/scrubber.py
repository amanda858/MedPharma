"""Generic CSV/XLSX scrubber.

Accepts ANY uploaded sheet of companies/orgs and:
  1. Auto-detects columns for name, phone, city, state, zip, website, email.
  2. For each row, runs rule_intercept to classify + score fit.
  3. Attempts to verify a real website using domain-guessing + content match
     (lab phone or city/state or org token must appear on the candidate site).
  4. Scrapes contact/about/team pages for on-domain emails (filters junk).
  5. Returns a normalized list of rows ready for Excel/CSV download.

Zero external paid APIs. Free, deterministic, rule-based.
"""

from __future__ import annotations

import asyncio
import csv
import io
import re
from typing import Any, Iterable, Optional

import httpx

from rule_intercept import intercept_excel_upload, intercept_request, score_lab_lead
try:
    from app.config import HUNTER_API_KEY as _HUNTER_API_KEY
except ImportError:
    _HUNTER_API_KEY = ""


# ─── Column auto-detection ──────────────────────────────────────────────

_COL_HINTS: dict[str, list[str]] = {
    "name":    ["organization", "org", "company", "lab name", "name", "business", "account", "entity"],
    "phone":   ["phone", "telephone", "tel", "mobile", "contact phone", "main"],
    "city":    ["city", "town", "locality"],
    "state":   ["state", "region", "province", "st"],
    "zip":     ["zip", "postal", "postcode"],
    "website": ["website", "url", "domain", "site", "homepage", "web"],
    "email":   ["email", "e-mail", "mail", "contact email"],
    "address": ["address", "street", "addr"],
    "npi":     ["npi"],
    "taxonomy":["taxonomy", "specialty", "type", "category"],
}


def detect_columns(headers: list[str]) -> dict[str, Optional[str]]:
    norm = {h: re.sub(r"[^a-z0-9]+", " ", (h or "").lower()).strip() for h in headers}
    mapped: dict[str, Optional[str]] = {k: None for k in _COL_HINTS}

    for field, hints in _COL_HINTS.items():
        for h, n in norm.items():
            if not n:
                continue
            for hint in hints:
                if hint == n or hint in n.split() or (len(hint) >= 4 and hint in n):
                    mapped[field] = h
                    break
            if mapped[field]:
                break
    return mapped


# ─── File parsing ───────────────────────────────────────────────────────

def parse_uploaded(content: bytes, filename: str) -> tuple[list[str], list[dict]]:
    """Return (headers, rows) from CSV/XLSX/XLS bytes."""
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xls")):
        try:
            from openpyxl import load_workbook
        except Exception as e:  # pragma: no cover
            raise RuntimeError(f"openpyxl required for Excel uploads: {e}")
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return [], []
        headers = [str(c).strip() if c is not None else "" for c in header_row]
        rows: list[dict] = []
        for r in rows_iter:
            if r is None:
                continue
            row = {headers[i]: ("" if v is None else str(v).strip()) for i, v in enumerate(r) if i < len(headers)}
            if any(v for v in row.values()):
                rows.append(row)
        return headers, rows

    # Default: CSV (or unknown — try CSV)
    text = content.decode("utf-8-sig", errors="replace")
    sniff = csv.Sniffer()
    try:
        dialect = sniff.sniff(text[:4096], delimiters=",;\t|")
    except Exception:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    headers = list(reader.fieldnames or [])
    rows = []
    for r in reader:
        clean = {k: (v.strip() if isinstance(v, str) else "") for k, v in r.items() if k}
        if any(clean.values()):
            rows.append(clean)
    return headers, rows


# ─── Email + domain heuristics ─────────────────────────────────────────

EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")
ASSET_RE = re.compile(r"(@2x|\.png|\.jpg|\.jpeg|\.svg|\.webp|\.gif)$", re.I)
PHONE_DIGITS = re.compile(r"\D+")
# Matches US phone numbers in common formats: (555) 555-5555 / 555-555-5555 / 5555555555
PHONE_RE = re.compile(r"\b(\(?\d{3}\)?[\s.\-]?\d{3}[\s.\-]?\d{4})\b")

_PUBLIC_MAIL = {
    "gmail.com", "yahoo.com", "hotmail.com", "outlook.com", "aol.com",
    "icloud.com", "protonmail.com", "live.com", "msn.com",
}
_BLOCKED_DOMAINS = {
    "sentry.io", "wixpress.com", "wordpress.com", "godaddy.com", "domainmarket.com",
    "wix.com", "squarespace.com", "cloudflare.com", "2x.com", "tektronix.com",
    "questdiagnostics.com", "google.com", "facebook.com", "instagram.com",
    "twitter.com", "linkedin.com", "youtube.com",
}
_GENERIC_LOCAL = {
    "info", "contact", "hello", "hi", "admin", "support", "sales", "office",
    "reception", "webmaster", "mail", "general", "inquiries", "enquiries",
    "marketing", "help", "noreply", "no-reply", "donotreply", "postmaster",
    "notifications", "alerts", "press", "media", "careers", "jobs", "hr",
    "customerservice", "service", "interested",
}
_STOP_WORDS = {
    # Corporate suffixes only — these are never part of a brand domain.
    # Keep healthcare descriptors (diagnostics, labs, pathology, health, etc.)
    # because they're often part of the actual domain (e.g. "Genova Diagnostics"
    # → genovadiagnostics.com).
    "llc", "inc", "pllc", "pa", "pc", "llp", "lp", "ltd", "corp", "corporation",
    "company", "co", "the", "and", "of", "for",
}
_PLACEHOLDER_LOCALS = {"user", "example", "yourname", "firstname", "lastname", "test", "demo"}

# Decision-maker title taxonomy — score 0-100
_DM_TITLES: dict[str, int] = {
    "chief executive officer": 100, "ceo": 100,
    "chief operating officer": 92, "coo": 92,
    "chief financial officer": 88, "cfo": 88,
    "chief medical officer": 92, "cmo": 92,
    "president": 88, "owner": 88, "founder": 86, "co-founder": 84,
    "vice president": 82, " vp,": 82, " vp ": 82,
    "medical director": 86, "lab director": 86, "laboratory director": 86,
    "director of": 78, "director,": 78, "director ": 76,
    "general manager": 74, "practice manager": 72, "office manager": 66,
    "billing manager": 74, "rcm manager": 76, "revenue cycle manager": 74,
    "credentialing manager": 72, "enrollment manager": 70, "compliance officer": 74,
    "operations manager": 70, "quality manager": 68,
    "manager": 60, "supervisor": 52, "administrator": 50,
    "coordinator": 44, "specialist": 36,
}

# Contact roles that indicate non-decision-makers
_SKIP_CONTACT_ROLES = {
    "customer service", "customer care", "receptionist", "front desk",
    "call center", "help desk", "technical support",
}


def _digits(s: str) -> str:
    return PHONE_DIGITS.sub("", s or "")


def _tokens(name: str) -> list[str]:
    toks = re.findall(r"[a-z][a-z0-9]{2,}", (name or "").lower())
    return [t for t in toks if t not in _STOP_WORDS and len(t) >= 3]


def _is_personal_local(local: str) -> bool:
    """Return True if the email local part looks like a real person\'s name (not a generic role)."""
    local = local.lower()
    if local in _GENERIC_LOCAL:
        return False
    # firstname.lastname, f.lastname, first_last patterns
    if "." in local or "_" in local:
        parts = re.split(r"[._]", local)
        return len(parts) >= 2 and all(p.isalpha() and 1 <= len(p) <= 20 for p in parts)
    # Likely concatenated name like jsmith, johnsmith — alpha only, 4-18 chars
    return local.isalpha() and 4 <= len(local) <= 18


def _dm_title_score(title: str) -> int:
    """Score a job title for decision-maker value (0-100)."""
    if not title:
        return 0
    tl = title.lower()
    best = 0
    for key, sc in _DM_TITLES.items():
        if key in tl:
            best = max(best, sc)
    for role in _SKIP_CONTACT_ROLES:
        if role in tl:
            return max(0, best - 35)
    return best


def _candidate_domains(name: str, hint: str = "") -> list[str]:
    out: list[str] = []
    seen: set[str] = set()

    def add(d: str):
        d = (d or "").lower().strip().lstrip(".")
        d = re.sub(r"^https?://", "", d).rstrip("/")
        d = d.split("/")[0]
        if d and d not in seen and "." in d and 4 <= len(d) <= 60:
            seen.add(d)
            out.append(d)

    if hint:
        add(hint)
        # Also try root domain of hint
        parts = hint.split(".")
        if len(parts) > 2:
            add(".".join(parts[-2:]))

    toks = _tokens(name)
    if toks:
        # Deterministic priority: multi-token org-like bases FIRST so we don't
        # collide with unrelated short-name companies (e.g. "Genova Diagnostics"
        # → genovadiagnostics.com before genova.com which is a different firm).
        preferred_tlds = (".com", ".org", ".net", ".co", ".us", ".io", ".health", ".bio")

        ordered_bases: list[str] = []

        def push_base(base: str) -> None:
            b = (base or "").strip().lower()
            if b and b not in ordered_bases and 3 <= len(b) <= 40:
                ordered_bases.append(b)

        # Multi-token first — these are far more likely to be the real domain
        if len(toks) >= 3:
            push_base("".join(toks[:3]))
            push_base("-".join(toks[:3]))
        if len(toks) >= 2:
            push_base(toks[0] + toks[1])
            push_base("-".join(toks[:2]))
        # Healthcare-suffix variants on the multi-token base
        base_root = "".join(toks[:2]) if len(toks) >= 2 else toks[0]
        for suf in ("lab", "labs", "diagnostics", "health", "medical", "rx", "group"):
            push_base(base_root + suf)
        # Single-token base last (most likely to be wrong for common words)
        push_base(toks[0])

        for b in ordered_bases:
            for tld in preferred_tlds:
                add(b + tld)

    return out[:30]


def _email_quality(email: str, name: str = "", title: str = "", org: str = "", verified_domain: str = "") -> int:
    """Score 0-100. <30 = junk, drop. Personal decision-maker emails score highest."""
    if not email or "@" not in email:
        return 0
    email = email.lower().strip()
    local, _, domain = email.partition("@")
    if ASSET_RE.search(email):
        return 0
    if domain in _BLOCKED_DOMAINS:
        return 0
    if local in _PLACEHOLDER_LOCALS:
        return 0
    if not re.match(r"^[a-z0-9._%+\-]+@[a-z0-9.\-]+\.[a-z]{2,}$", email):
        return 0

    # Domain must match org — unless it's the pre-verified domain, or public mail, or no org
    _vd_root = re.sub(r"^www\.", "", verified_domain) if verified_domain else ""
    if org and domain not in _PUBLIC_MAIL and domain != _vd_root:
        droot = re.sub(r"^www\.", "", domain).split(".")[0]
        toks = _tokens(org)
        match = False
        for t in toks:
            if t == droot or (len(t) >= 4 and t in droot) or (len(droot) >= 4 and droot in t):
                match = True
                break
        if not match:
            return 0

    score = 0
    # Domain quality
    if domain not in _PUBLIC_MAIL:
        score += 25
    else:
        score += 5

    # Email local part: personal names score much higher than generic roles
    if _is_personal_local(local):
        score += 48  # firstname.lastname — almost certainly a real decision maker
    elif local in _GENERIC_LOCAL:
        score += 5   # generic role address — low actionability
    else:
        score += 22  # ambiguous but not generic

    # Title-based DM boost
    ts = _dm_title_score(title)
    if ts >= 75:
        score += 22
    elif ts >= 50:
        score += 14
    elif ts >= 30:
        score += 7
    elif title and any(r in title.lower() for r in _SKIP_CONTACT_ROLES):
        score -= 15  # Customer-service penalty

    # Known name provides a small trust boost
    if name and len(name.strip()) >= 4:
        score += 5

    return min(100, max(0, score))


def _input_email_score(email: str) -> int:
    """Score 0-100 for user-supplied emails from spreadsheet.
    No org-domain matching — user already sourced these. Only junk-filter."""
    if not email or "@" not in email:
        return 0
    email = email.lower().strip()
    local, _, domain = email.partition("@")
    if ASSET_RE.search(email):
        return 0
    if domain in _BLOCKED_DOMAINS:
        return 0
    if local in _PLACEHOLDER_LOCALS:
        return 0
    if not re.match(r"^[a-z0-9._%+\-]+@[a-z0-9.\-]+\.[a-z]{2,}$", email):
        return 0
    score = 30  # base — user-sourced, trust it
    if domain not in _PUBLIC_MAIL:
        score += 20  # corporate domain bonus
    if _is_personal_local(local):
        score += 25  # firstname.lastname pattern
    elif local in _GENERIC_LOCAL:
        score += 0   # generic role address, still keep it
    else:
        score += 12  # ambiguous but not generic
    return min(100, score)


def _extract_named_contacts(html: str, domain: str) -> list[dict]:
    """
    Parse HTML to extract (name, title, email) contacts.
    Returns contacts sorted by decision-maker relevance — personal named emails first.
    """
    clean = re.sub(r"<(?:script|style)[^>]*>.*?</(?:script|style)>", " ", html, flags=re.S | re.I)
    text = re.sub(r"<[^>]+>", " ", clean)
    text = re.sub(r"\s+", " ", text)

    EMAIL_FIND = re.compile(r"([A-Za-z0-9._%+\-]+)@([A-Za-z0-9.\-]+\.[A-Za-z]{2,})")
    NAME_RE = re.compile(r"\b([A-Z][a-z]{2,20}(?:\s+[A-Z][a-z]{2,25}){1,2})\b")
    _HEADING_SKIP = {
        "contact us", "about us", "our team", "meet the", "learn more", "read more",
        "click here", "copyright", "all rights", "privacy policy", "terms of",
        "united states", "new york", "los angeles",
    }
    # Any word in a candidate "name" matching one of these is a geographic
    # token, not a human name. Rejects matches like "Clairsville Ohio",
    # "Fort Lauderdale", "Panama City".
    _GEO_NAME_TOKENS = {
        "alabama", "alaska", "arizona", "arkansas", "california", "colorado",
        "connecticut", "delaware", "florida", "georgia", "hawaii", "idaho",
        "illinois", "indiana", "iowa", "kansas", "kentucky", "louisiana",
        "maine", "maryland", "massachusetts", "michigan", "minnesota",
        "mississippi", "missouri", "montana", "nebraska", "nevada",
        "ohio", "oklahoma", "oregon", "pennsylvania", "tennessee", "texas",
        "utah", "vermont", "virginia", "washington", "wisconsin", "wyoming",
        "city", "county", "avenue", "street", "road", "boulevard", "drive",
        "lane", "parkway", "highway", "suite", "building", "floor", "plaza",
        "north", "south", "east", "west", "fort", "saint", "port", "mount",
        "lake", "river", "valley", "hills", "heights", "park", "beach",
        "clairsville", "lauderdale", "angeles", "diego", "francisco",
        "vegas", "orleans", "york", "jersey", "hampshire", "carolina",
        "dakota", "rica",
        # Common non-name nouns caught by the capitalized regex
        "main", "campus", "office", "department", "center", "centre",
        "team", "service", "services", "group", "staff", "clinic",
        "hospital", "laboratory", "laboratories", "medical", "health",
        "healthcare", "company", "corp", "corporation", "inc", "llc",
        "phone", "fax", "email", "contact", "customer", "support",
        "location", "locations", "patient", "patients", "provider",
        "hours", "monday", "tuesday", "wednesday", "thursday", "friday",
        "saturday", "sunday", "weekdays", "billing", "appointment",
        "schedule", "welcome", "home", "about", "mission", "vision",
        "history", "leadership", "management", "careers", "privacy",
        "terms", "policy", "rights", "reserved",
    }

    contacts: dict[str, dict] = {}
    for m in EMAIL_FIND.finditer(text):
        local, dom = m.group(1), m.group(2)
        if dom.lower() != domain:
            continue
        email = f"{local.lower()}@{dom.lower()}"
        if email in contacts or local.lower() in _PLACEHOLDER_LOCALS:
            continue

        # Context window: 600 chars before, 100 after
        ctx = text[max(0, m.start() - 600): m.end() + 100]

        # Find the closest capitalized name phrase before the email
        name = ""
        for nm in NAME_RE.finditer(ctx):
            candidate = nm.group(0).strip()
            if candidate.lower() in _HEADING_SKIP:
                continue
            parts = candidate.split()
            if len(parts) < 2:
                continue
            # Reject if ANY token in the candidate is a geographic word —
            # avoids matches like "Clairsville Ohio", "Fort Lauderdale".
            if any(p.lower() in _GEO_NAME_TOKENS for p in parts):
                continue
            name = candidate  # keep updating — last one wins (closest to email)

        # Best DM title in the same context window
        title, best_ts = "", 0
        ctx_lower = ctx.lower()
        for key, sc in _DM_TITLES.items():
            if key in ctx_lower and sc > best_ts:
                title = key.title()
                best_ts = sc

        is_personal = _is_personal_local(local)
        is_skip_role = any(r in ctx_lower for r in _SKIP_CONTACT_ROLES)

        contacts[email] = {
            "email": email,
            "name": name,
            "title": title,
            "dm_score": best_ts,
            "is_personal": is_personal,
            "is_skip_role": is_skip_role,
        }

    # Sort: personal named contacts first, then by DM score, generics last
    return sorted(
        contacts.values(),
        key=lambda c: (
            not c["is_personal"],
            -c["dm_score"],
            c["email"].split("@")[0] in _GENERIC_LOCAL,
        ),
    )


# ─── Async fetching / scraping ──────────────────────────────────────────

_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/120.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml",
}
_PAGES = [
    "", "/contact", "/contact-us", "/about", "/about-us", "/team",
    "/leadership", "/staff", "/providers", "/people", "/our-team",
    "/administration", "/management", "/billing", "/credentialing",
    "/compliance", "/directory", "/physicians", "/medical-staff",
    "/our-staff", "/who-we-are", "/meet-the-team", "/physician-directory",
    "/locations", "/resources",
]


async def _fetch(client: httpx.AsyncClient, url: str) -> Optional[str]:
    try:
        r = await client.get(url)
        if r.status_code == 200 and r.text:
            return r.text
    except Exception:
        return None
    return None


async def _verify_and_scrape(
    client: httpx.AsyncClient,
    org: str,
    phone: str = "",
    city: str = "",
    state: str = "",
    website_hint: str = "",
) -> dict:
    """Find a verified domain, scrape on-domain emails, extract a direct line, and find named decision-maker contacts."""
    phone_d = _digits(phone)[-10:]
    city_l = (city or "").lower().strip()
    state_l = (state or "").lower().strip()

    _raw_candidates = _candidate_domains(org, website_hint)
    # Interleave www. variants — many labs only resolve on www.
    candidates: list[str] = []
    _seen_c: set[str] = set()
    for _d in _raw_candidates:
        if _d not in _seen_c:
            _seen_c.add(_d)
            candidates.append(_d)
        _wd = ("www." + _d) if not _d.startswith("www.") else _d
        if _wd not in _seen_c:
            _seen_c.add(_wd)
            candidates.append(_wd)
    chosen: Optional[str] = None

    if phone_d:
        min_score = 30
    elif city_l:
        min_score = 20
    else:
        min_score = 16

    for dom in candidates:
        for sch in ("https://", "http://"):
            html = await _fetch(client, sch + dom)
            if not html:
                continue
            txt = html.lower()
            score = 0
            if phone_d and phone_d in _digits(html):
                score += 50
            if city_l and len(city_l) > 2 and city_l in txt:
                score += 15
            if state_l and len(state_l) == 2 and state_l in txt:
                score += 6
            for t in _tokens(org):
                if t in txt:
                    score += 8
            if score >= min_score:
                chosen = dom
                break
        if chosen:
            break

    if not chosen:
        return {"domain": None, "emails": [], "pattern_emails": [], "direct_line": "", "named_contacts": []}

    # Normalize: strip www. so email @domain matching is consistent
    root_chosen = re.sub(r"^www\.", "", chosen)

    # Scrape all contact/team/staff pages for emails, phones, named contacts
    emails: set[str] = set()
    raw_phones: list[str] = []
    html_pages: list[str] = []

    async def _pull(url: str) -> None:
        html = await _fetch(client, url)
        if html:
            for e in EMAIL_RE.findall(html):
                emails.add(e.lower())
            for p in PHONE_RE.findall(html):
                raw_phones.append(p)
            html_pages.append(html)

    # Try both www. and root forms so no page is missed
    pull_urls = [f"https://{chosen}{p}" for p in _PAGES]
    if chosen != root_chosen:
        pull_urls += [f"https://{root_chosen}{p}" for p in _PAGES[:6]]
    await asyncio.gather(*[_pull(u) for u in pull_urls])

    # Accept emails whose domain matches root (strips www. from both sides)
    good: list[str] = []
    for e in emails:
        local, _, dom = e.partition("@")
        if re.sub(r"^www\.", "", dom) != root_chosen:
            continue
        if ASSET_RE.search(e):
            continue
        if local in _PLACEHOLDER_LOCALS:
            continue
        good.append(e)
    good.sort(key=lambda e: (e.split("@")[0] in _GENERIC_LOCAL, e))

    # Extract best direct line
    direct_line = ""
    seen_digits: set[str] = set()
    _TOLL_FREE = {"800", "888", "877", "866", "855", "844", "833"}
    # Placeholder/fake phones commonly used as website filler
    _PLACEHOLDER_PHONES = {
        "9009009009", "0000000000", "1111111111", "1234567890",
        "9999999999", "5555555555", "1231231234", "8008008000",
        "0123456789",
    }
    def _is_placeholder(d: str) -> bool:
        if d in _PLACEHOLDER_PHONES:
            return True
        # repeat-digit or sequential-digit patterns
        if len(set(d)) <= 2:
            return True
        return False
    for raw in raw_phones:
        d = _digits(raw)[-10:]
        if len(d) != 10 or d in seen_digits:
            continue
        seen_digits.add(d)
        if d[:3] in _TOLL_FREE:
            continue
        if _is_placeholder(d):
            continue
        if phone_d and d == phone_d:
            continue
        direct_line = f"({d[0:3]}) {d[3:6]}-{d[6:10]}"
        break
    if not direct_line:
        for d in seen_digits:
            if len(d) == 10 and (not phone_d or d != phone_d) and not _is_placeholder(d):
                direct_line = f"({d[0:3]}) {d[3:6]}-{d[6:10]}"
                break

    # Extract named decision-maker contacts from all scraped pages
    combined_html = "\n".join(html_pages)
    named_contacts = _extract_named_contacts(combined_html, root_chosen) if combined_html else []

    # --- Pattern email generation ---
    # Build first.last@domain patterns from every named contact found on site
    pattern_addrs: list[str] = []
    _seen_p: set[str] = set(good)
    for nc in named_contacts[:6]:
        parts = (nc.get("name") or "").split()
        if len(parts) >= 2:
            fn = parts[0].lower().replace("-", "")
            ln = parts[-1].lower().replace("-", "")
            for pat in [
                f"{fn}.{ln}@{root_chosen}",
                f"{fn[0]}{ln}@{root_chosen}",
                f"{fn}{ln[0]}@{root_chosen}",
                f"{fn}.{ln[0]}@{root_chosen}",
            ]:
                if pat not in _seen_p and pat.split("@")[0] not in _PLACEHOLDER_LOCALS:
                    _seen_p.add(pat)
                    pattern_addrs.append(pat)
    # Standard high-value role prefixes common across all lab/medical orgs
    for _pfx in ("billing", "credentialing", "rcm", "lab", "director", "manager", "admin", "office"):
        _addr = f"{_pfx}@{root_chosen}"
        if _addr not in _seen_p:
            _seen_p.add(_addr)
            pattern_addrs.append(_addr)

    return {
        "domain": root_chosen,
        "emails": good,                    # all on-domain scraped emails (no cap)
        "pattern_emails": pattern_addrs,   # generated patterns (unverified)
        "direct_line": direct_line,
        "named_contacts": named_contacts,
    }


# ─── Public scrub API ───────────────────────────────────────────────────

async def scrub_rows(
    headers: list[str],
    rows: list[dict],
    *,
    concurrency: int = 8,
    per_row_timeout: float = 60.0,
    max_rows: int = 1000,
) -> dict[str, Any]:
    """Run the full scrub pass. Returns dict with mapped columns and result rows."""
    cols = detect_columns(headers)
    rows = rows[:max_rows]

    sem = asyncio.Semaphore(concurrency)
    limits = httpx.Limits(max_connections=concurrency * 4, max_keepalive_connections=concurrency * 2)
    client = httpx.AsyncClient(
        headers=_HEADERS, timeout=8.0, follow_redirects=True, limits=limits,
    )

    async def _do(row: dict) -> dict:
        org = (row.get(cols["name"]) or "").strip() if cols["name"] else ""
        phone = (row.get(cols["phone"]) or "").strip() if cols["phone"] else ""
        city = (row.get(cols["city"]) or "").strip() if cols["city"] else ""
        state = (row.get(cols["state"]) or "").strip() if cols["state"] else ""
        zipc = (row.get(cols["zip"]) or "").strip() if cols["zip"] else ""
        web = (row.get(cols["website"]) or "").strip() if cols["website"] else ""
        existing_email = (row.get(cols["email"]) or "").strip() if cols["email"] else ""
        addr = (row.get(cols["address"]) or "").strip() if cols["address"] else ""
        npi = (row.get(cols["npi"]) or "").strip() if cols["npi"] else ""
        tax = (row.get(cols["taxonomy"]) or "").strip() if cols["taxonomy"] else ""

        verified_domain: Optional[str] = None
        scraped_emails: list[str] = []
        pattern_email_list: list[str] = []
        direct_line: str = ""
        scrub_status = "ok"
        scrub_error = ""
        named_contacts: list[dict] = []

        if org:
            try:
                async with sem:
                    _scrape = await asyncio.wait_for(
                        _verify_and_scrape(client, org, phone=phone, city=city, state=state, website_hint=web),
                        timeout=per_row_timeout,
                    )
                    verified_domain    = _scrape.get("domain")
                    scraped_emails     = _scrape.get("emails", [])
                    pattern_email_list = _scrape.get("pattern_emails", [])
                    direct_line        = _scrape.get("direct_line", "")
                    named_contacts     = _scrape.get("named_contacts", [])
            except asyncio.TimeoutError:
                scrub_status = "timeout"
                scrub_error = f"row scrape exceeded {per_row_timeout}s"
            except Exception as e:
                scrub_status = "error"
                scrub_error = str(e)[:200]
        else:
            scrub_status = "skipped"
            scrub_error = "no organization name detected"

        # Build candidate emails: scraped → pattern-generated → Hunter → existing
        candidates: list[tuple[int, str]] = []
        vd = verified_domain or ""
        for e in scraped_emails:
            s = _email_quality(e, org=org, verified_domain=vd)
            if s > 0:
                candidates.append((s, e))
        # Pattern-generated (slight confidence penalty — not scrape-confirmed)
        for e in pattern_email_list:
            s = _email_quality(e, org=org, verified_domain=vd)
            s = max(0, s - 10)
            if s > 0:
                candidates.append((s, e))
        # Hunter.io — highest-trust source when API key is configured
        if vd and _HUNTER_API_KEY:
            try:
                from app.email_finder import hunter_domain_search
                _h_results, _ = await asyncio.wait_for(
                    hunter_domain_search(vd, _HUNTER_API_KEY), timeout=12.0
                )
                for _he in _h_results[:15]:
                    _he_email = _he.get("email", "")
                    if not _he_email:
                        continue
                    _hs = _email_quality(_he_email, org=org, title=_he.get("position", ""), verified_domain=vd)
                    _hs = min(100, _hs + 20)  # Hunter verified — trust boost
                    if _hs > 0:
                        candidates.append((_hs, _he_email))
            except Exception:
                pass
        # Blind role patterns — always generate against the best known domain
        # (verified domain if we got one, otherwise org-derived candidates) so
        # every row has usable outreach addresses even when scrape finds nothing.
        # CRITICAL: only use fallback domains that actually have MX records,
        # otherwise we generate emails for unrelated companies.
        if org:
            if verified_domain:
                _blind_doms = [verified_domain]
                _blind_score = 35  # verified domain → higher trust
            else:
                # MX-filter the candidate list. Domains without mail servers
                # are useless and frequently belong to unrelated companies.
                from app.email_verifier import lookup_mx
                _raw_doms = _candidate_domains(org, web)[:8]
                _blind_doms = []
                for _d in _raw_doms:
                    try:
                        _mx = await asyncio.wait_for(lookup_mx(_d), timeout=4.0)
                    except Exception:
                        _mx = []
                    if _mx:
                        _blind_doms.append(_d)
                    if len(_blind_doms) >= 2:
                        break
                _blind_score = 22  # MX-confirmed but org-match unproven
            _seen_blind: set[str] = {e for _, e in candidates}
            for _bd in _blind_doms:
                for _pfx in ("billing", "credentialing", "rcm", "lab", "director", "admin", "office", "info"):
                    _ba = f"{_pfx}@{_bd}"
                    if _ba not in _seen_blind:
                        _seen_blind.add(_ba)
                        candidates.append((_blind_score, _ba))

        # ── Real-human enrichment via NPI registry ─────────────────────
        # If the org is a US healthcare provider, NPPES exposes the
        # authorized official's name + title for free. Use that to
        # generate person-specific email patterns (firstname.lastname@,
        # flastname@, …) which beat role-account guessing.
        npi_official = None
        if org:
            try:
                from app.npi_client import find_org_official, person_email_patterns
                npi_official = await asyncio.wait_for(
                    find_org_official(org, state=state, city=city), timeout=10.0
                )
            except Exception:
                npi_official = None

            if npi_official and (verified_domain or _blind_doms):
                _person_doms = [verified_domain] if verified_domain else _blind_doms[:1]
                _person_score = 60 if verified_domain else 38
                from app.npi_client import person_email_patterns as _ppat
                _seen_p: set[str] = {e for _, e in candidates}
                for _pd in _person_doms:
                    for _pa in _ppat(npi_official["first"], npi_official["last"], _pd):
                        if _pa not in _seen_p:
                            _seen_p.add(_pa)
                            candidates.append((_person_score, _pa))

                # Promote the NPI official into the named-contact list.
                # The authorized official is a legally-registered signer —
                # higher trust than scraped page text. If the scraper's
                # current top_dm doesn't have a personal email tied to it,
                # put the NPI official in front.
                full_name = f"{npi_official['first']} {npi_official['last']}".strip()
                npi_contact = {
                    "name": full_name,
                    "title": npi_official.get("title", ""),
                    "email": "",
                    "dm_score": 80,  # NPI authorized official = real DM
                    "source": "NPI registry",
                }
                if not named_contacts:
                    named_contacts = [npi_contact]
                else:
                    _top = named_contacts[0]
                    _top_has_personal_email = bool(_top.get("email")) and _top.get("is_personal")
                    if not _top_has_personal_email:
                        # Prefer the verified NPI official over a scraped phrase
                        named_contacts = [npi_contact] + named_contacts
                if not direct_line and npi_official.get("phone"):
                    direct_line = npi_official["phone"]

        # ── Multi-platform social DM links (no scraping; user clicks in browser) ─
        # Spam filters kill cold email. DMs land. We build clickable
        # search URLs for LinkedIn / Facebook / Instagram / X plus the
        # company's own pages — user clicks a URL, lands on the platform
        # already logged in, picks the right account, pastes our DM.
        # Fully ToS-compliant. Zero account-ban risk.
        social_li_url = ""
        social_li_sales = ""
        social_fb_url = ""
        social_ig_url = ""
        social_x_url = ""
        social_google_li = ""
        social_google_all = ""
        social_li_company = ""
        social_fb_page = ""
        social_ig_company = ""
        msg_li_note = ""
        msg_li_first = ""
        msg_li_followup = ""
        msg_fb = ""
        msg_ig = ""
        msg_x = ""
        msg_sms = ""
        try:
            from app.social_finder import find_social_profiles, social_outreach_templates
            if npi_official:
                _sp = await find_social_profiles(
                    npi_official["first"],
                    npi_official["last"],
                    org=org,
                    title=npi_official.get("title", ""),
                )
                _tmpl = social_outreach_templates(npi_official["first"], org)
            elif org:
                # No named DM — still give them company-page DM URLs.
                _sp = await find_social_profiles("", "", org=org)
                _tmpl = social_outreach_templates("", org)
            else:
                _sp = None
                _tmpl = None
            if _sp:
                social_li_url      = _sp.get("linkedin_url", "")
                social_li_sales    = _sp.get("linkedin_sales_nav", "")
                social_fb_url      = _sp.get("facebook_url", "")
                social_ig_url      = _sp.get("instagram_url", "")
                social_x_url       = _sp.get("x_url", "")
                social_google_li   = _sp.get("google_linkedin", "")
                social_google_all  = _sp.get("google_social", "")
                social_li_company  = _sp.get("linkedin_company_url", "")
                social_fb_page     = _sp.get("facebook_page_url", "")
                social_ig_company  = _sp.get("instagram_company_url", "")
            if _tmpl:
                msg_li_note     = _tmpl["linkedin_connection_note"]
                msg_li_first    = _tmpl["linkedin_first_message"]
                msg_li_followup = _tmpl["linkedin_follow_up"]
                msg_fb          = _tmpl["facebook_dm"]
                msg_ig          = _tmpl["instagram_dm"]
                msg_x           = _tmpl["x_dm"]
                msg_sms         = _tmpl["sms"]
        except Exception:
            pass

        # ── Playbook: personalized hook + objection handlers + heat score ─
        # The personalized hook makes every DM feel custom-written.
        # Objection handlers are paste-ready replies for the 5 most common
        # pushbacks (already-have-biller, send-info, what-cost, etc.).
        # Heat score combines lab-fit + contact richness + NPI recency
        # to drive the "Daily Top 10" prioritization.
        hook_line = ""
        objection_lib: dict[str, str] = {}
        heat = 0
        heat_reasons: list[str] = []
        recency_label = ""
        npi_last_updated = ""
        try:
            from app.playbook import (
                personalized_hook, objection_handlers, heat_score,
                enrich_templates_with_hook, _recency_signal,
            )
            _first_for_hook = (npi_official or {}).get("first", "")
            _tax_for_hook = (npi_official or {}).get("taxonomy_desc", "") or tax
            _state_for_hook = (npi_official or {}).get("state", "") or state
            npi_last_updated = (npi_official or {}).get("last_updated", "")
            recency_label, _ = _recency_signal(npi_last_updated)

            hook_line = personalized_hook(
                first=_first_for_hook,
                org=org,
                taxonomy_desc=_tax_for_hook,
                lab_type_detected=lab_intel.get("lab_type_detected", "") if False else "",
                state=_state_for_hook,
                last_updated=npi_last_updated,
            )

            # Objection handlers
            objection_lib = objection_handlers(
                first=_first_for_hook,
                org=org,
            )

            # Inject hook + Calendly into the social templates
            if hook_line:
                _enriched = enrich_templates_with_hook({
                    "linkedin_first_message": msg_li_first,
                    "facebook_dm": msg_fb,
                    "instagram_dm": msg_ig,
                    "x_dm": msg_x,
                    "sms": msg_sms,
                }, hook_line)
                msg_li_first = _enriched.get("linkedin_first_message", msg_li_first)
                msg_fb = _enriched.get("facebook_dm", msg_fb)
                msg_ig = _enriched.get("instagram_dm", msg_ig)
                msg_x = _enriched.get("x_dm", msg_x)
                msg_sms = _enriched.get("sms", msg_sms)
        except Exception:
            pass

        # Existing input emails — user already sourced these; skip org-domain matching
        if existing_email:
            for e in re.split(r"[;,\s]+", existing_email):
                e = e.strip()
                if not e:
                    continue
                s = _input_email_score(e)
                if s > 0:
                    candidates.append((s, e))
        # Dedup, sort
        seen: set[str] = set()
        ranked: list[tuple[int, str]] = []
        for s, e in sorted(candidates, key=lambda x: -x[0]):
            if e in seen:
                continue
            seen.add(e)
            ranked.append((s, e))

        # ── Real-deliverability verification (in-house verifier) ─────────
        # Run MX + SMTP RCPT probe on top candidates. Drop confirmed-bad,
        # boost confirmed-good, leave catch-all/inconclusive in middle.
        verify_results: dict[str, dict] = {}
        if ranked:
            try:
                from app.email_verifier import verify_batch
                top_emails = [e for _, e in ranked[:8]]
                _vr = await asyncio.wait_for(
                    verify_batch(top_emails, do_smtp=True, concurrency=4),
                    timeout=25.0,
                )
                for r in _vr:
                    verify_results[r.get("email", "")] = r
            except Exception:
                pass

        if verify_results:
            adjusted: list[tuple[int, str]] = []
            for s, e in ranked:
                vr = verify_results.get(e)
                if not vr:
                    adjusted.append((s, e))
                    continue
                v = vr.get("verdict")
                if v == "undeliverable":
                    continue  # drop
                if v == "deliverable":
                    s = min(100, max(s, 85))
                elif v == "catch-all":
                    s = min(100, max(s, 50))
                elif v == "risky":
                    s = min(100, max(s, 40))
                adjusted.append((s, e))
            adjusted.sort(key=lambda x: -x[0])
            ranked = adjusted

        # ── Lab intelligence scoring (rule_intercept) ──────────────────
        lab_intel = score_lab_lead(org, lab_type=tax, state=state)

        # If we have an NPI authorized official AND a verified-deliverable
        # person-pattern email made it through, attach that as the DM email.
        if npi_official and not named_contacts:
            named_contacts = []
        if npi_official and named_contacts and not named_contacts[0].get("email"):
            _of_first = npi_official["first"].lower()
            _of_last = npi_official["last"].lower()
            for s, e in ranked:
                local = e.split("@")[0].lower()
                if (_of_first in local or _of_last in local) and s >= 50:
                    named_contacts[0]["email"] = e
                    break

        # Decision-maker contacts — top 3
        top_dm = named_contacts[0] if len(named_contacts) > 0 else {}
        dm2    = named_contacts[1] if len(named_contacts) > 1 else {}
        dm3    = named_contacts[2] if len(named_contacts) > 2 else {}
        dm_email = top_dm.get("email", "")

        # Fit score = lab quality score + contact-info richness bonuses
        fit = lab_intel["score"]
        if verified_domain:
            fit = min(100, fit + 5)
        if ranked:
            fit = min(100, fit + 8)
        if direct_line:
            fit = min(100, fit + 3)
        if top_dm.get("dm_score", 0) >= 60:
            fit = min(100, fit + 5)  # bonus for finding a real decision-maker
        if dm2:
            fit = min(100, fit + 3)  # bonus for multiple DM contacts

        # Re-personalize hook now that lab_intel is available (lab_type_detected
        # gives us a better signal than the raw taxonomy string).
        try:
            from app.playbook import (
                personalized_hook as _ph, heat_score as _hs,
                enrich_templates_with_hook as _eth,
            )
            _better_hook = _ph(
                first=(npi_official or {}).get("first", ""),
                org=org,
                taxonomy_desc=(npi_official or {}).get("taxonomy_desc", "") or tax,
                lab_type_detected=lab_intel.get("lab_type_detected", ""),
                state=(npi_official or {}).get("state", "") or state,
                last_updated=npi_last_updated,
            )
            if _better_hook:
                hook_line = _better_hook
                _enriched = _eth({
                    "linkedin_first_message": msg_li_first,
                    "facebook_dm": msg_fb,
                    "instagram_dm": msg_ig,
                    "x_dm": msg_x,
                    "sms": msg_sms,
                }, hook_line)
                msg_li_first = _enriched.get("linkedin_first_message", msg_li_first)
                msg_fb = _enriched.get("facebook_dm", msg_fb)
                msg_ig = _enriched.get("instagram_dm", msg_ig)
                msg_x = _enriched.get("x_dm", msg_x)
                msg_sms = _enriched.get("sms", msg_sms)

            heat, heat_reasons = _hs(
                lead_score=lab_intel["score"],
                fit_score=fit,
                has_dm=bool(top_dm.get("name")),
                has_direct_line=bool(direct_line),
                has_verified_domain=bool(verified_domain),
                has_social=bool(social_li_url or social_fb_url or social_ig_url or social_x_url),
                last_updated=npi_last_updated,
                state=state,
            )
        except Exception:
            pass

        return {
            "Lead Score": lab_intel["score"],
            "Heat Score": heat,
            "Heat Reasons": "; ".join(heat_reasons),
            "NPI Last Updated": npi_last_updated,
            "Recency Signal": recency_label,
            "Personalized Hook": hook_line,
            "Tier": lab_intel["tier"],
            "Priority": lab_intel["priority"],
            "Org Name": org,
            "Taxonomy / Type": tax,
            "Type Detected": lab_intel.get("lab_type_detected", ""),
            "NPI": npi,
            "Address": addr,
            "City": city,
            "State": state,
            "ZIP": zipc,
            "Phone": phone,
            "Direct Line": direct_line,
            # Decision-maker contacts (up to 3 named people found on site)
            "Decision Maker": top_dm.get("name", ""),
            "DM Title":       top_dm.get("title", ""),
            "DM Email":       dm_email,
            "DM 2":           dm2.get("name", ""),
            "DM 2 Title":     dm2.get("title", ""),
            "DM 2 Email":     dm2.get("email", ""),
            "DM 3":           dm3.get("name", ""),
            "DM 3 Title":     dm3.get("title", ""),
            "DM 3 Email":     dm3.get("email", ""),
            "Email 1": ranked[0][1] if len(ranked) >= 1 else "",
            "Email 1 Score": ranked[0][0] if len(ranked) >= 1 else "",
            "Email 2": ranked[1][1] if len(ranked) >= 2 else "",
            "Email 2 Score": ranked[1][0] if len(ranked) >= 2 else "",
            "Email 3": ranked[2][1] if len(ranked) >= 3 else "",
            "Email 3 Score": ranked[2][0] if len(ranked) >= 3 else "",
            "Email 4": ranked[3][1] if len(ranked) >= 4 else "",
            "Email 4 Score": ranked[3][0] if len(ranked) >= 4 else "",
            "Email 5": ranked[4][1] if len(ranked) >= 5 else "",
            "Email 5 Score": ranked[4][0] if len(ranked) >= 5 else "",
            "Existing Email (input)": existing_email,
            "Original Website": web,
            "Verified Domain": verified_domain or "",
            # ── Social DM channels (preferred outreach — DMs beat email) ──
            "LinkedIn URL":            social_li_url,
            "LinkedIn Sales Nav URL":  social_li_sales,
            "Facebook URL":            social_fb_url,
            "Instagram URL":           social_ig_url,
            "X / Twitter URL":         social_x_url,
            "Google LinkedIn Search":  social_google_li,
            "Google Social Search":    social_google_all,
            "LinkedIn Company Page":   social_li_company,
            "Facebook Company Page":   social_fb_page,
            "Instagram Company":       social_ig_company,
            # ── Pre-written DM messages (paste-ready) ──
            "LinkedIn Connection Note": msg_li_note,
            "LinkedIn First Message":   msg_li_first,
            "LinkedIn Follow-up":       msg_li_followup,
            "Facebook DM":              msg_fb,
            "Instagram DM":             msg_ig,
            "X / Twitter DM":           msg_x,
            "SMS Template":             msg_sms,
            # ── Objection handlers (paste-ready replies) ──
            "Reply: Already Have Biller":  objection_lib.get("objection_already_have_biller", ""),
            "Reply: Send Info First":      objection_lib.get("objection_send_info_first", ""),
            "Reply: What Does It Cost":    objection_lib.get("objection_what_does_it_cost", ""),
            "Reply: Not Interested":       objection_lib.get("objection_not_interested", ""),
            "Reply: Busy Now":             objection_lib.get("objection_busy_now", ""),
            "Reply: Who Are You":          objection_lib.get("objection_who_are_you", ""),
            "Lead Signals": "; ".join(lab_intel.get("signals", [])),
            "Fit Score": fit,
            "Intercept Category": lab_intel["category"],
            "Intercept Confidence": 1.0,
            "Scrub Status": scrub_status,
            "Scrub Error": scrub_error,
        }

    try:
        results = await asyncio.gather(*[_do(r) for r in rows], return_exceptions=True)
    finally:
        await client.aclose()

    out: list[dict] = []
    error_count = 0
    for r in results:
        if isinstance(r, Exception):
            error_count += 1
            out.append({"Scrub Status": "error", "Scrub Error": str(r)[:200]})
        else:
            out.append(r)

    # Sort by Heat Score (true buyability ranking) → Lead Score → Fit
    out.sort(key=lambda r: (
        -int(r.get("Heat Score", 0) or 0),
        -int(r.get("Lead Score", 0) or 0),
        -int(r.get("Fit Score", 0) or 0),
    ))

    # Daily Top 10 — the leads to hit RIGHT NOW
    daily_top_10 = [r for r in out if r.get("Heat Score", 0)][:10]

    summary = {
        "input_rows": len(rows),
        "output_rows": len(out),
        "verified_domains": sum(1 for r in out if r.get("Verified Domain")),
        "rows_with_email": sum(1 for r in out if r.get("Email 1")),
        "rows_with_dm": sum(1 for r in out if r.get("Decision Maker") or r.get("DM Email")),
        "rows_with_linkedin": sum(1 for r in out if r.get("LinkedIn URL")),
        "rows_with_social_dm": sum(
            1 for r in out
            if any(r.get(k) for k in ("LinkedIn URL", "Facebook URL", "Instagram URL", "X / Twitter URL"))
        ),
        "rows_hot_npi_recent": sum(1 for r in out if (r.get("Recency Signal") or "").startswith("HOT")),
        "rows_top_heat": sum(1 for r in out if int(r.get("Heat Score", 0) or 0) >= 70),
        "total_emails_found": sum(
            sum(1 for f in ("Email 1", "Email 2", "Email 3", "Email 4", "Email 5") if r.get(f))
            for r in out
        ),
        "errors": error_count,
        "detected_columns": cols,
    }
    return {"summary": summary, "rows": out, "daily_top_10": daily_top_10}


# ─── Output writers ─────────────────────────────────────────────────────

OUTPUT_FIELDS = [
    # ── Prioritization (heat score drives Daily Top 10) ──────────────
    "Heat Score", "Heat Reasons", "Recency Signal", "NPI Last Updated",
    "Lead Score", "Tier", "Priority",
    # ── Personalized opener ──────────────────────────────────────────
    "Personalized Hook",
    # ── Organization ─────────────────────────────────────────────────
    "Org Name", "Taxonomy / Type", "Type Detected",
    "NPI", "Address", "City", "State", "ZIP",
    # ── Contact ──────────────────────────────────────────────────────
    "Phone", "Direct Line",
    "Decision Maker", "DM Title", "DM Email",
    "DM 2", "DM 2 Title", "DM 2 Email",
    "DM 3", "DM 3 Title", "DM 3 Email",
    # ── PRIMARY OUTREACH: Social DM channels (paste-ready) ────────────
    # Spam blocks email; DMs land. These are the preferred channels.
    "LinkedIn URL",
    "LinkedIn Sales Nav URL",
    "Facebook URL",
    "Instagram URL",
    "X / Twitter URL",
    "Google Social Search",
    "Google LinkedIn Search",
    "LinkedIn Company Page",
    "Facebook Company Page",
    "Instagram Company",
    "LinkedIn Connection Note",
    "LinkedIn First Message",
    "LinkedIn Follow-up",
    "Facebook DM",
    "Instagram DM",
    "X / Twitter DM",
    "SMS Template",
    # ── Objection handlers (paste-ready replies) ─────────────────────
    "Reply: Already Have Biller",
    "Reply: Send Info First",
    "Reply: What Does It Cost",
    "Reply: Not Interested",
    "Reply: Busy Now",
    "Reply: Who Are You",
    # ── SECONDARY: Email (kept but de-prioritized) ───────────────────
    "Email 1", "Email 1 Score",
    "Email 2", "Email 2 Score",
    "Email 3", "Email 3 Score",
    "Email 4", "Email 4 Score",
    "Email 5", "Email 5 Score",
    "Existing Email (input)",
    # ── Web ───────────────────────────────────────────────────────────
    "Original Website", "Verified Domain",
    # ── Intelligence ──────────────────────────────────────────────────
    "Lead Signals",
    # ── Metadata ─────────────────────────────────────────────────────
    "Fit Score", "Intercept Category", "Intercept Confidence",
    "Scrub Status", "Scrub Error",
]


def to_csv_bytes(rows: Iterable[dict]) -> bytes:
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=OUTPUT_FIELDS, extrasaction="ignore")
    w.writeheader()
    for r in rows:
        w.writerow({k: r.get(k, "") for k in OUTPUT_FIELDS})
    return buf.getvalue().encode("utf-8")


def to_xlsx_bytes(rows: Iterable[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Scrubbed Leads"
    ws.append(OUTPUT_FIELDS)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for r in rows:
        ws.append([r.get(k, "") for k in OUTPUT_FIELDS])
    # Reasonable column widths
    widths = {
        "Lead Score": 12, "Tier": 10, "Priority": 12,
        "Org Name": 38, "Taxonomy / Type": 28, "Type Detected": 24,
        "NPI": 12, "Address": 28, "City": 16, "State": 8, "ZIP": 10,
        "Phone": 18, "Direct Line": 18,
        "Email 1": 34, "Email 1 Score": 12,
        "Email 2": 34, "Email 2 Score": 12,
        "Email 3": 34, "Email 3 Score": 12,
        "Existing Email (input)": 32,
        "Original Website": 26, "Verified Domain": 26,
        "Lead Signals": 48,
        "Fit Score": 10, "Intercept Category": 14, "Intercept Confidence": 14,
        "Scrub Status": 12, "Scrub Error": 28,
    }
    for i, h in enumerate(OUTPUT_FIELDS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = widths.get(h, 18)
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
