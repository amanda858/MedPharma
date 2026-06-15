"""
Notification system — Individual Progress Report.

Buffers all activity during a user's session and sends ONE consolidated
"Individual Progress" report when the user logs out, including:
  • Activity breakdown by section
  • Industry-standard RCM benchmarks comparison
  • AI-powered productivity analysis (via OpenAI, with rule-based fallback)

Configuration via environment variables:
  SENDGRID_API_KEY — SendGrid API key for sending email
  NOTIFY_EMAIL     — Comma-separated destination emails for notifications
  SENDGRID_FROM    — Sender email address (must be verified in SendGrid)
  TWILIO_SID     — Twilio Account SID
  TWILIO_TOKEN   — Twilio Auth Token
  TWILIO_FROM    — Twilio phone number (E.164 format, e.g. +18001234567)
  NOTIFY_PHONE   — Destination phone for SMS (E.164 format)
  OPENAI_API_KEY — For AI productivity narrative (optional; falls back to rule-based)
"""

import os
import logging
import threading
import json
import smtplib
import urllib.request
import urllib.parse
import urllib.error
import base64
from collections import defaultdict
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

log = logging.getLogger("notifications")


def _normalize_phone(value: str) -> str:
    """Normalize phone input to E.164 when possible (US-focused fallback)."""
    raw = (value or "").strip()
    if not raw:
        return ""
    if raw.startswith("+"):
        return raw
    digits = "".join(ch for ch in raw if ch.isdigit())
    if len(digits) == 10:
        return f"+1{digits}"
    if len(digits) == 11 and digits.startswith("1"):
        return f"+{digits}"
    return raw

# ── Configuration ──
SENDGRID_API_KEY = os.getenv("SENDGRID_API_KEY", "")
SENDGRID_FROM = os.getenv("SENDGRID_FROM", "notifications@medprosc.com")
NOTIFY_EMAILS = [e.strip() for e in os.getenv("NOTIFY_EMAIL", "eric@medprosc.com").split(",") if e.strip()]

TWILIO_SID = os.getenv("TWILIO_SID", "")
TWILIO_TOKEN = os.getenv("TWILIO_TOKEN", "")
TWILIO_FROM = os.getenv("TWILIO_FROM", "")
NOTIFY_PHONE = _normalize_phone(os.getenv("NOTIFY_PHONE", "+18036263500"))

# Free carrier email-to-SMS gateways — no Twilio required.
# Set NOTIFY_PHONE_CARRIER to your carrier name (e.g. "att", "verizon", "tmobile").
# The 10-digit phone from NOTIFY_PHONE is sent as an email to <number>@<gateway>.
_CARRIER_GATEWAYS: dict[str, str] = {
    "att": "txt.att.net",
    "at&t": "txt.att.net",
    "verizon": "vtext.com",
    "vzw": "vtext.com",
    "tmobile": "tmomail.net",
    "t-mobile": "tmomail.net",
    "sprint": "messaging.sprintpcs.com",
    "metro": "mymetropcs.com",
    "metropcs": "mymetropcs.com",
    "boost": "sms.myboostmobile.com",
    "cricket": "sms.cricketwireless.net",
    "uscellular": "email.uscc.net",
    "us cellular": "email.uscc.net",
    "straighttalk": "vtext.com",
    "tracfone": "txt.att.net",
    "googlefi": "msg.fi.google.com",
    "fi": "msg.fi.google.com",
}
NOTIFY_PHONE_CARRIER = os.getenv("NOTIFY_PHONE_CARRIER", "").strip().lower()

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
IN_APP_ONLY_MODE = os.getenv("NOTIFY_IN_APP_ONLY", "0").strip().lower() in {"1", "true", "yes", "on"}


def _carrier_sms_email(phone: str, carrier: str) -> str | None:
    """Return email-to-SMS gateway address for a phone+carrier combo, or None."""
    gateway = _CARRIER_GATEWAYS.get(carrier.strip().lower(), "")
    if not gateway:
        return None
    digits = "".join(c for c in (phone or "") if c.isdigit())
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    if len(digits) != 10:
        return None
    return f"{digits}@{gateway}"


def _live_config() -> dict:
    """Read notification credentials FRESH every call.

    Resolution order: in-DB app_settings (set via admin UI) → env vars →
    module-level defaults. This lets the admin paste credentials in the
    hub without ever touching Render env vars.
    """
    # Pull DB-stored secrets first; missing keys come back empty so the
    # env / default fallbacks still apply.
    try:
        from app.client_db import get_app_setting as _gs
        db_sg_key  = (_gs("SENDGRID_API_KEY") or "").strip()
        db_sg_from = (_gs("SENDGRID_FROM") or "").strip()
        db_smtp_h  = (_gs("SMTP_HOST") or "").strip()
        db_smtp_p  = (_gs("SMTP_PORT") or "").strip()
        db_smtp_u  = (_gs("SMTP_USER") or "").strip()
        db_smtp_pw = (_gs("SMTP_PASS") or "").strip()
        db_notify  = (_gs("NOTIFY_EMAILS") or "").strip()
    except Exception:
        db_sg_key = db_sg_from = db_smtp_h = db_smtp_p = ""
        db_smtp_u = db_smtp_pw = db_notify = ""

    sg_key = db_sg_key or os.getenv("SENDGRID_API_KEY", "") or SENDGRID_API_KEY
    sg_from = (db_sg_from
               or os.getenv("SENDGRID_FROM", "")
               or "notifications@medprosc.com"
               or SENDGRID_FROM)
    notify_raw = (db_notify
                  or os.getenv("NOTIFY_EMAIL", "eric@medprosc.com"))
    emails = [e.strip() for e in notify_raw.split(",") if e.strip()] or NOTIFY_EMAILS
    t_sid = os.getenv("TWILIO_SID", "") or TWILIO_SID
    t_tok = os.getenv("TWILIO_TOKEN", "") or TWILIO_TOKEN
    t_from = _normalize_phone(os.getenv("TWILIO_FROM", "") or TWILIO_FROM)
    phone = _normalize_phone(os.getenv("NOTIFY_PHONE", "+18036263500")) or NOTIFY_PHONE
    carrier = os.getenv("NOTIFY_PHONE_CARRIER", "").strip().lower() or NOTIFY_PHONE_CARRIER
    smtp_h = db_smtp_h or os.getenv("SMTP_HOST", "smtp.gmail.com") or SMTP_HOST
    smtp_p_raw = db_smtp_p or os.getenv("SMTP_PORT", "587") or str(SMTP_PORT)
    try:
        smtp_p = int(smtp_p_raw or 587)
    except (TypeError, ValueError):
        smtp_p = 587
    smtp_u = db_smtp_u or os.getenv("SMTP_USER", "") or SMTP_USER
    smtp_pw = db_smtp_pw or os.getenv("SMTP_PASS", "") or SMTP_PASS
    in_app = os.getenv("NOTIFY_IN_APP_ONLY", "0").strip().lower() in {"1", "true", "yes", "on"}
    return {
        "SENDGRID_API_KEY": sg_key, "SENDGRID_FROM": sg_from, "NOTIFY_EMAILS": emails,
        "TWILIO_SID": t_sid, "TWILIO_TOKEN": t_tok, "TWILIO_FROM": t_from,
        "NOTIFY_PHONE": phone, "NOTIFY_PHONE_CARRIER": carrier,
        "SMTP_HOST": smtp_h, "SMTP_PORT": smtp_p,
        "SMTP_USER": smtp_u, "SMTP_PASS": smtp_pw, "IN_APP_ONLY_MODE": in_app,
    }

# Users whose activity triggers notifications (team members only).
# Owner (Eric) should receive reports, not be tracked as a worker by default.
# Supports comma-separated env var, e.g. NOTIFY_ON_USERS="jessica,rcm"
# Use NOTIFY_ON_USERS="*" to enable for all users.
# Default is "*" so every staff/client action funnels to the admin inbox; set
# NOTIFY_ON_USERS in the environment to restrict it to a specific allowlist.
_notify_on_users_env = os.getenv("NOTIFY_ON_USERS", "*").strip()
NOTIFY_ON_USERS = {
    u.strip().lower() for u in _notify_on_users_env.split(",") if u.strip()
} if _notify_on_users_env and _notify_on_users_env != "*" else {"*"}

# SMTP (primary send path — no SendGrid account required)
SMTP_HOST = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASS = os.getenv("SMTP_PASS", "")


# ── Industry-standard RCM benchmarks (actions per 8-hour day) ──
# Sources: MGMA, HBMA, AAPC industry reports for medical billing / credentialing
INDUSTRY_BENCHMARKS = {
    "Claims":          {"daily_target": 180, "unit": "claims touched", "per_hour": 25,
                        "note": "Industry avg: 25-35 claims processed/hr (AR follow-up, posting, submissions)"},
    "Credentialing":   {"daily_target": 12,  "unit": "credentialing actions", "per_hour": 1.5,
                        "note": "Industry avg: 3-5 new apps + 8-15 follow-ups/day"},
    "EDI":             {"daily_target": 40,  "unit": "EDI transactions", "per_hour": 5,
                        "note": "Industry avg: 40-60 clearinghouse transactions/day"},
    "Production":      {"daily_target": 10,  "unit": "production log entries", "per_hour": 1.25,
                        "note": "Standard: logging tasks, time tracking, QA notes"},
    "SLA Tracking":    {"daily_target": 15,  "unit": "SLA updates", "per_hour": 2,
                        "note": "Industry avg: 15-25 SLA/TAT status updates/day"},
    "Notes":           {"daily_target": 20,  "unit": "notes", "per_hour": 2.5,
                        "note": "Standard documentation pace for RCM workflows"},
}
# Catch-all for sections not explicitly listed
_DEFAULT_BENCHMARK = {"daily_target": 15, "unit": "actions", "per_hour": 2,
                      "note": "General administrative RCM benchmark"}

# ── In-memory activity buffer (keyed by username) ──
# Each entry: list of {"action", "section", "detail", "timestamp", "raw_ts"}
_activity_buffer: dict[str, list[dict]] = defaultdict(list)
_session_start: dict[str, datetime] = {}          # first activity time per user
_buffer_lock = threading.Lock()

# Auto-flush tuning (no manual logout required)
AUTO_FLUSH_ACTION_THRESHOLD = int(os.getenv("NOTIFY_AUTO_FLUSH_THRESHOLD", "20"))
AUTO_FLUSH_MAX_AGE_MINUTES = int(os.getenv("NOTIFY_AUTO_FLUSH_MAX_AGE_MINUTES", "60"))


def _should_notify(username: str) -> bool:
    """Return True if this user's activity should trigger notifications."""
    u = (username or "").lower()
    if "*" in NOTIFY_ON_USERS:
        return True
    alias_map = {
        "jess": {"jess", "jessica"},
        "jessica": {"jess", "jessica"},
        "eric": {"eric", "admin"},
        "admin": {"eric", "admin"},
        "rcm": {"rcm"},
    }
    aliases = alias_map.get(u, {u})
    return any(a in NOTIFY_ON_USERS for a in aliases)


def _get_benchmark(section: str) -> dict:
    """Return the industry benchmark dict for a section (fuzzy match)."""
    s = section.strip().lower()
    for key, bench in INDUSTRY_BENCHMARKS.items():
        if key.lower() in s or s in key.lower():
            return {**bench, "section_key": key}
    return {**_DEFAULT_BENCHMARK, "section_key": section}


# ───────────────────────────────────────────────────────────────────────
#  AI Productivity Analysis
# ───────────────────────────────────────────────────────────────────────

def _generate_ai_summary(username: str, date_str: str, session_hrs: float,
                         by_section: dict, benchmarks_data: list, overall_pct: float) -> str:
    """
    Call OpenAI to produce a 3-5 sentence narrative evaluating the employee's
    productivity against RCM industry standards.  Falls back to rule-based if
    API key is missing or the call fails.
    """
    section_summary = "\n".join(
        f"  - {b['section']}: {b['actual']} actions done, benchmark {b['target']}/day "
        f"({b['pct']}% of target). {b['note']}"
        for b in benchmarks_data
    )

    prompt = f"""You are a medical billing team lead reviewing an employee's daily production.

Employee: {username}
Date: {date_str}
Active session: {session_hrs:.1f} hours
Overall productivity: {overall_pct:.0f}% of industry standard

Section-by-section breakdown:
{section_summary}

Industry context: Standard RCM workday is 7.5-8 hrs. Medical billing specialists
should process 25-35 claims/hr, credentialing staff handle 3-5 new apps + 15 follow-ups/day.

Write a concise 3-5 sentence "Individual Progress Assessment" that:
1. States whether this employee met, exceeded, or fell short of daily expectations
2. Highlights their strongest area and any area needing improvement
3. Assesses whether the employee worked a productive and sufficient day
4. Gives one specific, actionable recommendation

Keep the tone professional but direct — this is an internal team lead report.
Do NOT use bullet points; write in paragraph form. Do not include any greeting.
"""

    if not OPENAI_API_KEY:
        return _rule_based_summary(username, session_hrs, benchmarks_data, overall_pct)

    try:
        from openai import OpenAI
        client = OpenAI(api_key=OPENAI_API_KEY)
        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "system", "content": "You are a healthcare RCM operations team lead."},
                      {"role": "user", "content": prompt}],
            max_tokens=350,
            temperature=0.7,
        )
        return resp.choices[0].message.content.strip()
    except Exception as e:
        log.error(f"OpenAI productivity analysis failed: {e}")
        return _rule_based_summary(username, session_hrs, benchmarks_data, overall_pct)


def _rule_based_summary(username: str, session_hrs: float,
                        benchmarks_data: list, overall_pct: float) -> str:
    """Fallback narrative when OpenAI is unavailable."""
    # Find best and worst sections
    if not benchmarks_data:
        return f"{username} had no measurable activity to evaluate."

    best = max(benchmarks_data, key=lambda b: b["pct"])
    worst = min(benchmarks_data, key=lambda b: b["pct"])

    if overall_pct >= 100:
        rating = "exceeded daily production expectations"
    elif overall_pct >= 75:
        rating = "met most daily production expectations"
    elif overall_pct >= 50:
        rating = "fell below expected daily production volume"
    else:
        rating = "significantly underperformed against industry benchmarks"

    summary = (
        f"{username} {rating} with an overall productivity score of {overall_pct:.0f}% "
        f"across a {session_hrs:.1f}-hour session. "
    )
    if best["pct"] > 0:
        summary += (
            f"Their strongest area was {best['section']} at {best['pct']:.0f}% of target "
            f"({best['actual']} of {best['target']} expected). "
        )
    if worst["section"] != best["section"] and worst["pct"] < 80:
        summary += (
            f"{worst['section']} needs attention — only {worst['pct']:.0f}% of the daily benchmark was reached. "
        )

    if session_hrs < 6:
        summary += (
            f"The session duration of {session_hrs:.1f} hours is below the standard 7.5-8 hour workday, "
            f"which likely contributed to lower output. Consider reviewing time management."
        )
    elif overall_pct < 75:
        summary += "Recommend reviewing workflow efficiency and reducing non-productive time."
    else:
        summary += "Keep up the consistent work."

    return summary


# ── Public API — called from route handlers ──

def notify_activity(username: str, action: str, section: str, detail: str = "", sms_copy: bool = False):
    """Buffer a single activity event.
    If sms_copy=True, also send an immediate concise SMS copy."""
    if not _should_notify(username):
        return
    should_flush = False
    sms_text = ""
    with _buffer_lock:
        now = datetime.now()
        key = username.lower()
        if key not in _session_start:
            _session_start[key] = now
        _activity_buffer[key].append({
            "action": action,
            "section": section,
            "detail": detail,
            "timestamp": now.strftime("%I:%M %p"),
            "raw_ts": now,
        })
        should_flush = len(_activity_buffer[key]) >= AUTO_FLUSH_ACTION_THRESHOLD
        if sms_copy:
            sms_text = f"{username}: {action} {section}"
            if detail:
                sms_text += f" | {detail}"
            if len(sms_text) > 155:
                sms_text = sms_text[:152] + "…"

    if should_flush:
        threading.Thread(target=flush_and_notify, args=(username,), daemon=True).start()
    if sms_text:
        threading.Thread(target=_send_sms, args=(sms_text,), daemon=True).start()


def notify_bulk_activity(username: str, action: str, section: str, count: int, detail: str = ""):
    """Buffer a bulk activity event (does NOT send immediately)."""
    if not _should_notify(username):
        return
    should_flush = False
    with _buffer_lock:
        now = datetime.now()
        key = username.lower()
        if key not in _session_start:
            _session_start[key] = now
        _activity_buffer[key].append({
            "action": f"{action} {count} records in",
            "section": section,
            "detail": detail,
            "timestamp": now.strftime("%I:%M %p"),
            "raw_ts": now,
        })
        should_flush = len(_activity_buffer[key]) >= AUTO_FLUSH_ACTION_THRESHOLD

    if should_flush:
        threading.Thread(target=flush_and_notify, args=(username,), daemon=True).start()


def flush_all_pending_notifications():
    """
    Flush buffered notifications for users with enough activity or stale sessions.
    This prevents dependence on manual logout.
    """
    with _buffer_lock:
        snapshot = {
            user: list(items)
            for user, items in _activity_buffer.items()
            if items
        }

    now = datetime.now()
    for user, items in snapshot.items():
        if not items:
            continue
        last_ts = items[-1].get("raw_ts")
        age_min = ((now - last_ts).total_seconds() / 60.0) if isinstance(last_ts, datetime) else 0
        if len(items) >= AUTO_FLUSH_ACTION_THRESHOLD or age_min >= AUTO_FLUSH_MAX_AGE_MINUTES:
            try:
                flush_and_notify(user)
            except Exception as e:
                log.error(f"Auto flush failed for {user}: {e}")


def flush_and_notify(username: str):
    """
    Called at logout — builds a full **Individual Progress** report with:
      • Activity breakdown by section
      • Industry benchmark comparison
      • AI-powered productivity narrative
    Sends one email + SMS, then clears the buffer.
    """
    key = username.lower()
    with _buffer_lock:
        activities = list(_activity_buffer.pop(key, []))
        session_start = _session_start.pop(key, None)

    if not activities or not _should_notify(username):
        return

    # ── Timing ──
    now = datetime.now()
    date_str = now.strftime("%B %d, %Y")
    time_str = now.strftime("%I:%M %p")
    if session_start:
        session_hrs = (now - session_start).total_seconds() / 3600
        start_str = session_start.strftime("%I:%M %p")
    else:
        session_hrs = 0
        start_str = "N/A"

    # ── Group by section ──
    by_section: dict[str, list[dict]] = defaultdict(list)
    for a in activities:
        by_section[a["section"]].append(a)

    # ── Benchmark comparison ──
    benchmarks_data = []
    total_pct_sum, total_sections = 0, 0
    for section, items in by_section.items():
        bench = _get_benchmark(section)
        actual = len(items)
        # Pro-rate target if session < 8 hrs
        effective_hrs = min(session_hrs, 8) if session_hrs > 0 else 8
        prorated_target = max(1, round(bench["daily_target"] * (effective_hrs / 8)))
        pct = round((actual / prorated_target) * 100, 1) if prorated_target else 0
        benchmarks_data.append({
            "section": section,
            "actual": actual,
            "target": prorated_target,
            "full_day_target": bench["daily_target"],
            "pct": pct,
            "per_hour": bench["per_hour"],
            "unit": bench["unit"],
            "note": bench["note"],
        })
        total_pct_sum += pct
        total_sections += 1

    overall_pct = round(total_pct_sum / total_sections, 1) if total_sections else 0

    # ── Productivity rating badge ──
    if overall_pct >= 110:
        rating_label, rating_color, rating_bg = "EXCEEDS STANDARDS", "#15803d", "#dcfce7"
    elif overall_pct >= 85:
        rating_label, rating_color, rating_bg = "MEETS STANDARDS", "#2563eb", "#dbeafe"
    elif overall_pct >= 60:
        rating_label, rating_color, rating_bg = "BELOW EXPECTATIONS", "#d97706", "#fef3c7"
    else:
        rating_label, rating_color, rating_bg = "NEEDS IMPROVEMENT", "#dc2626", "#fee2e2"

    # ── AI Summary ──
    ai_summary = _generate_ai_summary(username, date_str, session_hrs,
                                       by_section, benchmarks_data, overall_pct)

    # ── Plain text body ──
    lines = [
        "═══════════════════════════════════════════",
        "       INDIVIDUAL PROGRESS REPORT",
        "═══════════════════════════════════════════",
        "",
        f"  Employee:  {username}",
        f"  Date:      {date_str}",
        f"  Session:   {start_str} — {time_str} ({session_hrs:.1f} hrs)",
        f"  Rating:    {rating_label} ({overall_pct:.0f}%)",
        "",
        "───────────────────────────────────────────",
        "  AI PRODUCTIVITY ASSESSMENT",
        "───────────────────────────────────────────",
        "",
        f"  {ai_summary}",
        "",
        "───────────────────────────────────────────",
        "  PRODUCTION vs INDUSTRY BENCHMARKS",
        "───────────────────────────────────────────",
        "",
    ]
    for b in benchmarks_data:
        bar_filled = min(20, round(b["pct"] / 5))
        bar = "█" * bar_filled + "░" * (20 - bar_filled)
        lines.append(f"  {b['section']}")
        lines.append(f"    Done: {b['actual']}  |  Target: {b['target']}  |  {b['pct']:.0f}%")
        lines.append(f"    [{bar}]")
        lines.append(f"    {b['note']}")
        lines.append("")

    lines += [
        "───────────────────────────────────────────",
        "  ACTIVITY DETAIL",
        "───────────────────────────────────────────",
        "",
    ]
    for section, items in by_section.items():
        lines.append(f"  ── {section} ({len(items)} actions) ──")
        for item in items:
            line = f"    • {item['timestamp']} — {item['action']} {item['section']}"
            if item.get("detail"):
                line += f" ({item['detail']})"
            lines.append(line)
        lines.append("")

    body = "\n".join(lines)

    # ── HTML body for email ──
    # Benchmark rows
    bench_rows_html = ""
    for b in benchmarks_data:
        pct_clamped = min(b["pct"], 100)
        if b["pct"] >= 100:
            bar_color = "#22c55e"
        elif b["pct"] >= 70:
            bar_color = "#3b82f6"
        elif b["pct"] >= 40:
            bar_color = "#f59e0b"
        else:
            bar_color = "#ef4444"
        bench_rows_html += f"""
        <tr>
            <td style="padding:10px 12px;border-bottom:1px solid #f1f5f9;font-weight:600;font-size:13px;white-space:nowrap">{b['section']}</td>
            <td style="padding:10px 12px;border-bottom:1px solid #f1f5f9;text-align:center;font-size:14px;font-weight:700">{b['actual']}</td>
            <td style="padding:10px 12px;border-bottom:1px solid #f1f5f9;text-align:center;font-size:13px;color:#64748b">{b['target']}</td>
            <td style="padding:10px 8px;border-bottom:1px solid #f1f5f9;width:140px">
                <div style="background:#f1f5f9;border-radius:6px;height:12px;overflow:hidden">
                    <div style="background:{bar_color};height:100%;width:{pct_clamped}%;border-radius:6px"></div>
                </div>
            </td>
            <td style="padding:10px 12px;border-bottom:1px solid #f1f5f9;text-align:center;font-weight:700;font-size:13px;color:{bar_color}">{b['pct']:.0f}%</td>
        </tr>"""

    # Activity detail rows
    section_html = ""
    for section, items in by_section.items():
        rows = ""
        for item in items:
            detail_txt = f"<br><span style='color:#64748b;font-size:11px'>{item['detail']}</span>" if item.get("detail") else ""
            rows += f"""<tr>
                <td style="padding:6px 12px;border-bottom:1px solid #f8fafc;font-size:12px;color:#94a3b8;white-space:nowrap">{item['timestamp']}</td>
                <td style="padding:6px 12px;border-bottom:1px solid #f8fafc;font-size:12px">{item['action']}{detail_txt}</td>
            </tr>"""
        section_html += f"""
        <div style="margin-bottom:12px">
            <div style="font-weight:600;font-size:13px;color:#475569;padding:6px 0;border-bottom:1px solid #e2e8f0">{section} — {len(items)} action{'s' if len(items)!=1 else ''}</div>
            <table style="width:100%;border-collapse:collapse">{rows}</table>
        </div>"""

    html_body = f"""
    <html>
    <body style="font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; padding: 20px; color: #1e293b; background: #f8fafc;">
        <div style="max-width: 680px; margin: 0 auto; border: 1px solid #e2e8f0; border-radius: 12px; overflow: hidden; background: white;">

            <!-- HEADER -->
            <div style="background: linear-gradient(135deg, #0f172a, #1e293b); padding: 24px 28px;">
                <h1 style="color: white; margin: 0; font-size: 22px; font-weight: 800; letter-spacing: 0.5px;">📊 INDIVIDUAL PROGRESS REPORT</h1>
                <div style="margin-top: 12px; display: flex; gap: 20px;">
                    <div>
                        <div style="font-size: 11px; text-transform: uppercase; color: #94a3b8; font-weight: 600;">Employee</div>
                        <div style="font-size: 16px; color: #f1f5f9; font-weight: 700;">{username.upper()}</div>
                    </div>
                    <div>
                        <div style="font-size: 11px; text-transform: uppercase; color: #94a3b8; font-weight: 600;">Date</div>
                        <div style="font-size: 16px; color: #f1f5f9; font-weight: 700;">{date_str}</div>
                    </div>
                    <div>
                        <div style="font-size: 11px; text-transform: uppercase; color: #94a3b8; font-weight: 600;">Session</div>
                        <div style="font-size: 16px; color: #f1f5f9; font-weight: 700;">{start_str} — {time_str}</div>
                    </div>
                </div>
            </div>

            <div style="padding: 24px 28px;">

                <!-- KPI CARDS ROW -->
                <div style="display:flex;gap:16px;margin-bottom:24px;">
                    <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;padding:16px;flex:1;text-align:center">
                        <div style="font-size:30px;font-weight:800;color:#1e293b">{len(activities)}</div>
                        <div style="font-size:10px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:0.5px">Total Actions</div>
                    </div>
                    <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;padding:16px;flex:1;text-align:center">
                        <div style="font-size:30px;font-weight:800;color:#1e293b">{len(by_section)}</div>
                        <div style="font-size:10px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:0.5px">Sections Worked</div>
                    </div>
                    <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;padding:16px;flex:1;text-align:center">
                        <div style="font-size:30px;font-weight:800;color:#1e293b">{session_hrs:.1f}</div>
                        <div style="font-size:10px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:0.5px">Hours Active</div>
                    </div>
                    <div style="background:{rating_bg};border:1px solid {rating_color}22;border-radius:10px;padding:16px;flex:1.2;text-align:center">
                        <div style="font-size:30px;font-weight:800;color:{rating_color}">{overall_pct:.0f}%</div>
                        <div style="font-size:10px;font-weight:700;color:{rating_color};text-transform:uppercase;letter-spacing:0.5px">{rating_label}</div>
                    </div>
                </div>

                <!-- AI PRODUCTIVITY ASSESSMENT -->
                <div style="background:linear-gradient(135deg,#ede9fe,#e0e7ff);border-left:4px solid #6366f1;border-radius:8px;padding:18px 20px;margin-bottom:24px;">
                    <div style="font-size:12px;font-weight:800;text-transform:uppercase;color:#4338ca;letter-spacing:1px;margin-bottom:8px;">🤖 AI Productivity Assessment</div>
                    <div style="font-size:13px;line-height:1.7;color:#1e293b;">{ai_summary}</div>
                </div>

                <!-- BENCHMARK COMPARISON TABLE -->
                <div style="margin-bottom:24px;">
                    <div style="font-size:14px;font-weight:800;color:#1e293b;text-transform:uppercase;letter-spacing:0.5px;padding-bottom:8px;border-bottom:2px solid #1e293b;margin-bottom:8px;">
                        Production vs Industry Benchmarks
                    </div>
                    <table style="width:100%;border-collapse:collapse;">
                        <thead>
                            <tr style="background:#f8fafc;">
                                <th style="padding:8px 12px;text-align:left;font-size:11px;font-weight:700;color:#64748b;text-transform:uppercase">Section</th>
                                <th style="padding:8px 12px;text-align:center;font-size:11px;font-weight:700;color:#64748b;text-transform:uppercase">Actual</th>
                                <th style="padding:8px 12px;text-align:center;font-size:11px;font-weight:700;color:#64748b;text-transform:uppercase">Target</th>
                                <th style="padding:8px 12px;text-align:center;font-size:11px;font-weight:700;color:#64748b;text-transform:uppercase">Progress</th>
                                <th style="padding:8px 12px;text-align:center;font-size:11px;font-weight:700;color:#64748b;text-transform:uppercase">Score</th>
                            </tr>
                        </thead>
                        <tbody>{bench_rows_html}</tbody>
                    </table>
                </div>

                <!-- ACTIVITY DETAIL (collapsed look) -->
                <div style="margin-bottom:16px;">
                    <div style="font-size:14px;font-weight:800;color:#1e293b;text-transform:uppercase;letter-spacing:0.5px;padding-bottom:8px;border-bottom:2px solid #1e293b;margin-bottom:8px;">
                        Activity Detail
                    </div>
                    {section_html}
                </div>

                <hr style="border: none; border-top: 1px solid #e2e8f0; margin: 20px 0;">
                <p style="font-size: 11px; color: #94a3b8; text-align: center; margin: 0;">
                    Individual Progress Report — MedPharma Hub — {date_str}
                </p>
            </div>
        </div>
    </body>
    </html>"""

    subject = f"Individual Progress: {username} — {rating_label} ({overall_pct:.0f}%) — {date_str}"

    # SMS — concise summary with rating
    section_counts = ", ".join(f"{s}: {len(items)}" for s, items in by_section.items())
    sms = (f"Progress Report: {username} | {date_str} | "
           f"{rating_label} ({overall_pct:.0f}%) | "
           f"{len(activities)} actions in {session_hrs:.1f}hrs | {section_counts}")
    if len(sms) > 155:
        sms = (f"Progress: {username} | {rating_label} ({overall_pct:.0f}%) | "
               f"{len(activities)} actions in {session_hrs:.1f}hrs")
        if len(sms) > 155:
            sms = sms[:152] + "…"

    # Fire both in background threads
    threading.Thread(target=_send_email, args=(subject, body, html_body), daemon=True).start()
    threading.Thread(target=_send_sms, args=(sms,), daemon=True).start()
    log.info(f"Individual progress report queued for {username}: {rating_label} "
             f"({overall_pct:.0f}%) — {len(activities)} actions across {len(by_section)} sections")


# ── Send helpers ──

def _send_email(subject: str, body: str, html_body: str = ""):
    """Send email notification via SendGrid v3 API.
    Uses _live_config() to read credentials fresh (avoids stale cache)."""
    cfg = _live_config()
    if cfg["IN_APP_ONLY_MODE"]:
        log.info(f"In-app notification mode active — email send simulated: {subject}")
        return

    emails = cfg["NOTIFY_EMAILS"]
    sg_key = cfg["SENDGRID_API_KEY"]
    sg_from = cfg["SENDGRID_FROM"]
    smtp_h = cfg["SMTP_HOST"]
    smtp_p = cfg["SMTP_PORT"]
    smtp_u = cfg["SMTP_USER"]
    smtp_pw = cfg["SMTP_PASS"]

    if not emails:
        log.debug("Email notification skipped — NOTIFY_EMAILS not configured")
        return

    # Primary: SendGrid
    if sg_key:
        try:
            import httpx
            content = []
            if body:
                content.append({"type": "text/plain", "value": body})
            if html_body:
                content.append({"type": "text/html", "value": html_body})
            if not content:
                content.append({"type": "text/plain", "value": "(no content)"})

            recipients = [{"email": addr} for addr in emails]
            payload = {
                "personalizations": [{"to": recipients}],
                "from": {"email": sg_from, "name": "MedPharma Hub"},
                "subject": subject,
                "content": content,
            }
            resp = httpx.post(
                "https://api.sendgrid.com/v3/mail/send",
                json=payload,
                headers={
                    "Authorization": f"Bearer {sg_key}",
                    "Content-Type": "application/json",
                },
                timeout=15,
            )
            if resp.status_code in (200, 202):
                log.info(f"Email sent via SendGrid to {', '.join(emails)}: {subject}")
                return
            log.error(f"SendGrid failed ({resp.status_code}): {resp.text}")
        except Exception as e:
            log.error(f"Failed to send email via SendGrid: {e}")

    # Fallback: SMTP
    if not smtp_h or not smtp_u or not smtp_pw:
        log.error("Email notification skipped — no working provider configured (SendGrid/SMTP)")
        return

    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"] = smtp_u or sg_from
        msg["To"] = ", ".join(emails)

        plain = body or "(no content)"
        msg.attach(MIMEText(plain, "plain"))
        if html_body:
            msg.attach(MIMEText(html_body, "html"))

        with smtplib.SMTP(smtp_h, smtp_p, timeout=20) as server:
            server.starttls()
            server.login(smtp_u, smtp_pw)
            server.sendmail(msg["From"], emails, msg.as_string())
        log.info(f"Email sent via SMTP to {', '.join(emails)}: {subject}")
    except Exception as e:
        log.error(f"Failed to send email via SMTP: {e}")


def _send_sms(message: str):
    """Send SMS notification.
    Priority:
      1. Twilio (if TWILIO_SID/TOKEN/FROM configured)
      2. Carrier email-to-SMS via SMTP (if NOTIFY_PHONE_CARRIER + SMTP configured — FREE, no sign-up)
      3. Skip silently with a log warning
    """
    cfg = _live_config()
    if cfg["IN_APP_ONLY_MODE"]:
        log.info("In-app notification mode active — SMS send simulated")
        return

    t_sid = cfg["TWILIO_SID"]
    t_tok = cfg["TWILIO_TOKEN"]
    t_from = cfg["TWILIO_FROM"]
    phone = cfg["NOTIFY_PHONE"]
    carrier = cfg["NOTIFY_PHONE_CARRIER"]

    # Path 1: Twilio
    if t_sid and t_tok and t_from and phone:
        try:
            import httpx
            url = f"https://api.twilio.com/2010-04-01/Accounts/{t_sid}/Messages.json"
            data = {"To": phone, "From": t_from, "Body": message}
            resp = httpx.post(url, data=data, auth=(t_sid, t_tok), timeout=15)
            if resp.status_code in (200, 201):
                log.info(f"SMS sent via Twilio to {phone}")
                return
            log.error(f"Twilio SMS failed ({resp.status_code}): {resp.text}")
        except Exception as e:
            log.error(f"Failed to send SMS via Twilio: {e}")
        return

    # Path 2: Carrier email-to-SMS (free — just needs SMTP + NOTIFY_PHONE_CARRIER)
    sms_email = _carrier_sms_email(phone, carrier) if carrier else None
    smtp_h = cfg["SMTP_HOST"]
    smtp_u = cfg["SMTP_USER"]
    smtp_pw = cfg["SMTP_PASS"]
    smtp_p = cfg["SMTP_PORT"]

    if sms_email and smtp_h and smtp_u and smtp_pw:
        try:
            import email.mime.text as _mt
            msg = _mt.MIMEText(message[:160])
            msg["Subject"] = ""
            msg["From"] = smtp_u
            msg["To"] = sms_email
            with smtplib.SMTP(smtp_h, smtp_p, timeout=20) as server:
                server.ehlo()
                server.starttls()
                server.login(smtp_u, smtp_pw)
                server.sendmail(smtp_u, [sms_email], msg.as_string())
            log.info(f"SMS sent via carrier email-to-SMS ({carrier}) to {sms_email}")
            return
        except Exception as e:
            log.error(f"Failed to send SMS via carrier email-to-SMS: {e}")
        return

    if not carrier:
        log.debug("SMS skipped — set NOTIFY_PHONE_CARRIER (e.g. att, verizon, tmobile) to enable free SMS via SMTP")
    else:
        log.debug(f"SMS skipped — carrier '{carrier}' configured but SMTP credentials missing (SMTP_USER/SMTP_PASS)")


def _send_email_force(subject: str, body: str, html_body: str = ""):
    """Send email — always attempts delivery (ignores IN_APP_ONLY_MODE).
    Uses stdlib urllib.request (no third-party dependency).
    Tries SendGrid first, then SMTP fallback."""
    cfg = _live_config()
    emails = cfg["NOTIFY_EMAILS"]
    sg_key = cfg["SENDGRID_API_KEY"]
    sg_from = cfg["SENDGRID_FROM"]
    smtp_h = cfg["SMTP_HOST"]
    smtp_p = cfg["SMTP_PORT"]
    smtp_u = cfg["SMTP_USER"]
    smtp_pw = cfg["SMTP_PASS"]

    if not emails:
        raise ValueError("No email recipients configured (NOTIFY_EMAIL env var)")

    log.info(f"[TEST] Email attempt — SendGrid={'YES' if sg_key else 'NO'}, SMTP={'YES' if smtp_u else 'NO'}, to={emails}")

    sg_error = None

    # Primary: SendGrid (using stdlib urllib)
    if sg_key:
        try:
            content = []
            if body:
                content.append({"type": "text/plain", "value": body})
            if html_body:
                content.append({"type": "text/html", "value": html_body})
            if not content:
                content.append({"type": "text/plain", "value": "(no content)"})

            recipients = [{"email": addr} for addr in emails]
            payload = json.dumps({
                "personalizations": [{"to": recipients}],
                "from": {"email": sg_from, "name": "MedPharma Hub"},
                "subject": subject,
                "content": content,
            }).encode("utf-8")

            req = urllib.request.Request(
                "https://api.sendgrid.com/v3/mail/send",
                data=payload,
                headers={
                    "Authorization": f"Bearer {sg_key}",
                    "Content-Type": "application/json",
                },
                method="POST",
            )
            resp = urllib.request.urlopen(req, timeout=20)
            status = resp.getcode()
            if status in (200, 202):
                log.info(f"[TEST] Email sent via SendGrid to {', '.join(emails)}: {subject}")
                return
            resp_body = resp.read().decode("utf-8", errors="replace")[:300]
            sg_error = f"SendGrid returned {status}: {resp_body}"
            log.error(f"[TEST] {sg_error}")
        except urllib.error.HTTPError as he:
            resp_body = he.read().decode("utf-8", errors="replace")[:300]
            sg_error = f"SendGrid HTTP {he.code}: {resp_body}"
            log.error(f"[TEST] {sg_error}")
        except Exception as e:
            sg_error = f"SendGrid error: {e}"
            log.error(f"[TEST] {sg_error}")

    # Fallback: SMTP
    if smtp_h and smtp_u and smtp_pw:
        try:
            msg = MIMEMultipart("alternative")
            msg["Subject"] = subject
            msg["From"] = smtp_u or sg_from
            msg["To"] = ", ".join(emails)
            msg.attach(MIMEText(body or "(no content)", "plain"))
            if html_body:
                msg.attach(MIMEText(html_body, "html"))

            with smtplib.SMTP(smtp_h, smtp_p, timeout=25) as server:
                server.starttls()
                server.login(smtp_u, smtp_pw)
                server.sendmail(msg["From"], emails, msg.as_string())
            log.info(f"[TEST] Email sent via SMTP to {', '.join(emails)}: {subject}")
            return
        except Exception as e:
            smtp_error = f"SMTP error: {e}"
            log.error(f"[TEST] {smtp_error}")
            if sg_error:
                raise RuntimeError(f"{sg_error} | {smtp_error}")
            raise RuntimeError(smtp_error)

    if sg_error:
        raise RuntimeError(sg_error)
    raise ValueError("No email provider configured — set SENDGRID_API_KEY or SMTP_USER+SMTP_PASS in Render environment")


def _send_sms_force(message: str):
    """Send SMS — always attempts delivery (ignores IN_APP_ONLY_MODE).
    Uses stdlib urllib.request (no third-party dependency).
    Reads credentials LIVE from env vars to avoid stale cached values."""
    cfg = _live_config()
    t_sid = cfg["TWILIO_SID"]
    t_tok = cfg["TWILIO_TOKEN"]
    t_from = cfg["TWILIO_FROM"]
    phone = cfg["NOTIFY_PHONE"]

    if not t_sid or not t_tok or not t_from:
        missing = [k for k, v in {"TWILIO_SID": t_sid, "TWILIO_TOKEN": t_tok, "TWILIO_FROM": t_from}.items() if not v]
        raise ValueError(f"Twilio not configured — missing: {', '.join(missing)}")
    if not phone:
        raise ValueError("No SMS recipient configured — set NOTIFY_PHONE env var")

    log.info(f"[TEST] SMS attempt — from={t_from}, to={phone}, sid={t_sid[:8]}...")

    url = f"https://api.twilio.com/2010-04-01/Accounts/{t_sid}/Messages.json"
    post_data = urllib.parse.urlencode({
        "To": phone,
        "From": t_from,
        "Body": message,
    }).encode("utf-8")
    auth_str = base64.b64encode(f"{t_sid}:{t_tok}".encode()).decode()

    req = urllib.request.Request(
        url,
        data=post_data,
        headers={
            "Authorization": f"Basic {auth_str}",
            "Content-Type": "application/x-www-form-urlencoded",
        },
        method="POST",
    )
    try:
        resp = urllib.request.urlopen(req, timeout=20)
        status = resp.getcode()
        resp_body = resp.read().decode("utf-8", errors="replace")
        if status in (200, 201):
            log.info(f"[TEST] SMS sent to {phone} (status {status})")
            # Try to extract Message SID for confirmation
            try:
                msg_data = json.loads(resp_body)
                log.info(f"[TEST] Twilio Message SID: {msg_data.get('sid', 'N/A')}")
            except Exception:
                pass
            return
        raise RuntimeError(f"Twilio SMS failed ({status}): {resp_body[:300]}")
    except urllib.error.HTTPError as he:
        resp_body = he.read().decode("utf-8", errors="replace")[:500]
        # Parse Twilio error for clearer message
        try:
            err_data = json.loads(resp_body)
            twilio_msg = err_data.get("message", resp_body[:300])
            twilio_code = err_data.get("code", "")
            raise RuntimeError(f"Twilio error {he.code} (code {twilio_code}): {twilio_msg}")
        except (json.JSONDecodeError, RuntimeError) as inner:
            if isinstance(inner, RuntimeError):
                raise
            raise RuntimeError(f"Twilio HTTP {he.code}: {resp_body[:300]}")


def send_direct_sms(message: str, to_phone: str = "") -> dict:
    """Send an immediate SMS via Twilio.
    If to_phone is omitted, uses NOTIFY_PHONE from env."""
    cfg = _live_config()
    t_sid = cfg["TWILIO_SID"]
    t_tok = cfg["TWILIO_TOKEN"]
    t_from = cfg["TWILIO_FROM"]
    phone = _normalize_phone(to_phone) if to_phone else cfg["NOTIFY_PHONE"]

    if not t_sid or not t_tok or not t_from:
        missing = [k for k, v in {"TWILIO_SID": t_sid, "TWILIO_TOKEN": t_tok, "TWILIO_FROM": t_from}.items() if not v]
        raise ValueError(f"Twilio not configured — missing: {', '.join(missing)}")
    if not phone:
        raise ValueError("No SMS recipient configured — set NOTIFY_PHONE or pass to_phone")

    text = (message or "").strip()
    if not text:
        raise ValueError("SMS message cannot be empty")

    url = f"https://api.twilio.com/2010-04-01/Accounts/{t_sid}/Messages.json"
    post_data = urllib.parse.urlencode({
        "To": phone,
        "From": t_from,
        "Body": text,
    }).encode("utf-8")
    auth_str = base64.b64encode(f"{t_sid}:{t_tok}".encode()).decode()

    req = urllib.request.Request(
        url,
        data=post_data,
        headers={
            "Authorization": f"Basic {auth_str}",
            "Content-Type": "application/x-www-form-urlencoded",
        },
        method="POST",
    )

    try:
        resp = urllib.request.urlopen(req, timeout=20)
        status = resp.getcode()
        body = resp.read().decode("utf-8", errors="replace")
        data = json.loads(body) if body else {}
        if status not in (200, 201):
            raise RuntimeError(f"Twilio SMS failed ({status}): {body[:300]}")
        return {
            "ok": True,
            "to": phone,
            "from": t_from,
            "sid": data.get("sid", ""),
            "status": data.get("status", "queued"),
        }
    except urllib.error.HTTPError as he:
        resp_body = he.read().decode("utf-8", errors="replace")[:500]
        try:
            err_data = json.loads(resp_body)
            twilio_msg = err_data.get("message", resp_body[:300])
            twilio_code = err_data.get("code", "")
            raise RuntimeError(f"Twilio error {he.code} (code {twilio_code}): {twilio_msg}")
        except (json.JSONDecodeError, RuntimeError) as inner:
            if isinstance(inner, RuntimeError):
                raise
            raise RuntimeError(f"Twilio HTTP {he.code}: {resp_body[:300]}")


def send_test_notification(triggered_by: str = "system") -> dict:
    """Send an immediate test notification to configured recipients/channels.
    Test notifications ALWAYS attempt real delivery (bypass IN_APP_ONLY_MODE).
    Runs SYNCHRONOUSLY — no threading — to ensure errors are captured clearly."""
    cfg = _live_config()
    now = datetime.now().strftime("%Y-%m-%d %I:%M %p")
    subject = f"MedPharma Hub Test Notification — {now}"
    emails = cfg["NOTIFY_EMAILS"]
    phone = cfg["NOTIFY_PHONE"]

    body = (
        f"This is a test notification from MedPharma Hub.\n"
        f"Triggered by: {triggered_by}\n"
        f"Time: {now}\n"
        f"Recipients: {', '.join(emails) if emails else 'none'}\n"
        f"Mode: {'IN_APP_ONLY (bypassed for test)' if cfg['IN_APP_ONLY_MODE'] else 'External delivery'}\n"
        f"SMS target: {phone or 'not configured'}"
    )
    html_body = f"""
    <html><body style="font-family:Arial,sans-serif;padding:20px">
      <div style="max-width:500px;margin:0 auto;border:2px solid #22c55e;border-radius:12px;overflow:hidden">
        <div style="background:linear-gradient(135deg,#0f172a,#1e293b);padding:20px;color:white">
          <h2 style="margin:0">✅ MedPharma Hub — Notifications Active</h2>
        </div>
        <div style="padding:20px">
          <p><b>Triggered by:</b> {triggered_by}</p>
          <p><b>Time:</b> {now}</p>
          <p><b>Email to:</b> {', '.join(emails)}</p>
          <p><b>SMS to:</b> {phone or 'not configured'}</p>
          <p style="color:#22c55e;font-weight:bold;font-size:16px">If you received this, your notification pipeline is working!</p>
        </div>
      </div>
    </body></html>
    """
    sms = f"MedPharma Hub test OK | by {triggered_by} | {now}"
    if len(sms) > 155:
        sms = sms[:152] + "…"

    results = {"email_sent": False, "sms_sent": False, "email_error": None, "sms_error": None}

    # ── EMAIL (synchronous) ──
    try:
        _send_email_force(subject, body, html_body)
        results["email_sent"] = True
    except Exception as e:
        results["email_error"] = str(e)
        log.error(f"[TEST] Email failed: {e}")

    # ── SMS (synchronous) ──
    try:
        _send_sms_force(sms)
        results["sms_sent"] = True
    except Exception as e:
        results["sms_error"] = str(e)
        log.error(f"[TEST] SMS failed: {e}")

    status = get_notification_status()
    status["ok"] = True
    status.update(results)
    log.info(f"[TEST] Results: email_sent={results['email_sent']}, sms_sent={results['sms_sent']}, "
             f"email_error={results['email_error']}, sms_error={results['sms_error']}")
    return status


def get_notification_status() -> dict:
    """Return current notification channel configuration status.
    Reads env vars LIVE to avoid stale module-level cache."""
    cfg = _live_config()
    sendgrid_configured = bool(cfg["SENDGRID_API_KEY"])
    smtp_configured = bool(cfg["SMTP_HOST"] and cfg["SMTP_USER"] and cfg["SMTP_PASS"])
    twilio_configured = bool(cfg["TWILIO_SID"] and cfg["TWILIO_TOKEN"] and cfg["TWILIO_FROM"] and cfg["NOTIFY_PHONE"])

    missing_twilio = []
    if not cfg["TWILIO_SID"]:
        missing_twilio.append("TWILIO_SID")
    if not cfg["TWILIO_TOKEN"]:
        missing_twilio.append("TWILIO_TOKEN")
    if not cfg["TWILIO_FROM"]:
        missing_twilio.append("TWILIO_FROM")
    if not cfg["NOTIFY_PHONE"]:
        missing_twilio.append("NOTIFY_PHONE")

    missing_email = []
    if not sendgrid_configured and not smtp_configured:
        missing_email.append("SENDGRID_API_KEY or SMTP_USER/SMTP_PASS")

    return {
        "email_recipients": cfg["NOTIFY_EMAILS"],
        "sms_target": cfg["NOTIFY_PHONE"],
        "sendgrid_configured": sendgrid_configured,
        "smtp_configured": smtp_configured,
        "twilio_configured": twilio_configured,
        "email_configured": bool(sendgrid_configured or smtp_configured),
        "missing_twilio_fields": missing_twilio,
        "missing_email_fields": missing_email,
        "notify_on_users": sorted(list(NOTIFY_ON_USERS)),
        "in_app_only_mode": cfg["IN_APP_ONLY_MODE"],
        "delivery_mode": "in_app_only" if cfg["IN_APP_ONLY_MODE"] else "external",
    }


def get_notification_debug() -> dict:
    """Return detailed diagnostic info for debugging notification delivery.
    Credential values are masked (first 4 chars shown)."""
    cfg = _live_config()

    def mask(val):
        if not val:
            return "NOT SET"
        s = str(val)
        if len(s) <= 4:
            return "****"
        return s[:4] + "*" * min(len(s) - 4, 12)

    return {
        "env_vars": {
            "SENDGRID_API_KEY": mask(cfg["SENDGRID_API_KEY"]),
            "SENDGRID_FROM": cfg["SENDGRID_FROM"],
            "NOTIFY_EMAIL": ", ".join(cfg["NOTIFY_EMAILS"]),
            "TWILIO_SID": mask(cfg["TWILIO_SID"]),
            "TWILIO_TOKEN": mask(cfg["TWILIO_TOKEN"]),
            "TWILIO_FROM": cfg["TWILIO_FROM"] or "NOT SET",
            "NOTIFY_PHONE": cfg["NOTIFY_PHONE"] or "NOT SET",
            "SMTP_HOST": cfg["SMTP_HOST"] or "NOT SET",
            "SMTP_PORT": str(cfg["SMTP_PORT"]),
            "SMTP_USER": mask(cfg["SMTP_USER"]),
            "SMTP_PASS": mask(cfg["SMTP_PASS"]),
            "NOTIFY_IN_APP_ONLY": os.getenv("NOTIFY_IN_APP_ONLY", "1"),
            "NOTIFY_ON_USERS": os.getenv("NOTIFY_ON_USERS", "jessica,rcm"),
        },
        "effective": {
            "in_app_only_mode": cfg["IN_APP_ONLY_MODE"],
            "sendgrid_ready": bool(cfg["SENDGRID_API_KEY"]),
            "smtp_ready": bool(cfg["SMTP_HOST"] and cfg["SMTP_USER"] and cfg["SMTP_PASS"]),
            "twilio_ready": bool(cfg["TWILIO_SID"] and cfg["TWILIO_TOKEN"] and cfg["TWILIO_FROM"]),
            "email_provider": "SendGrid" if cfg["SENDGRID_API_KEY"] else ("SMTP" if (cfg["SMTP_USER"] and cfg["SMTP_PASS"]) else "NONE"),
            "email_recipients": cfg["NOTIFY_EMAILS"],
            "sms_recipient": cfg["NOTIFY_PHONE"],
            "sms_from": cfg["TWILIO_FROM"] or "NOT SET",
        },
        "module_cache_vs_live": {
            "sendgrid_cached": bool(SENDGRID_API_KEY),
            "sendgrid_live": bool(cfg["SENDGRID_API_KEY"]),
            "twilio_cached": bool(TWILIO_SID),
            "twilio_live": bool(cfg["TWILIO_SID"]),
            "smtp_cached": bool(SMTP_USER),
            "smtp_live": bool(cfg["SMTP_USER"]),
            "in_app_cached": IN_APP_ONLY_MODE,
            "in_app_live": cfg["IN_APP_ONLY_MODE"],
        },
    }


# ═══════════════════════════════════════════════════════════════════════════
#  DAILY OVERALL ACCOUNT SUMMARY — Scheduled at 6 PM EST
# ═══════════════════════════════════════════════════════════════════════════

def _fmt_money(val):
    """Format a number as $X,XXX.XX"""
    return f"${val:,.2f}"


def _generate_account_ai_summary(d: dict, date_str: str) -> str:
    """
    Use OpenAI to generate a concise executive summary of the overall account
    health.  Falls back to rule-based if OpenAI unavailable.
    """
    prompt = f"""You are a healthcare RCM operations manager writing a brief end-of-day executive summary.

Date: {date_str}

Key metrics:
- Total AR: ${d['total_ar']:,.2f} across {d['total_claims']} claims ({d['active_claims']} active)
- Today: {d['submitted_today']} submitted, {d['paid_today']} paid, {d['denied_today']} denied
- MTD: {d['submitted_mtd']} submitted, {d['paid_mtd']} paid, {d['denied_mtd']} denied
- Payments today: ${d['payments_today']:,.2f} | MTD: ${d['payments_mtd']:,.2f} | YTD: ${d['payments_ytd']:,.2f}
- Clean claim rate: {d['clean_claim_rate']}% | Denial rate: {d['denial_rate']}% | Net collection: {d['net_collection_rate']}%
- Avg days to pay: {d['avg_days_to_pay']} | SLA breaches: {d['sla_breaches']}
- AR Aging: Current ${d['ar_aging']['current']:,.2f} | 31-60 ${d['ar_aging']['31_60']:,.2f} | 61-90 ${d['ar_aging']['61_90']:,.2f} | 90+ ${d['ar_aging']['90_plus']:,.2f}
- Credentialing: {d['cred_total']} total ({d['cred_approved']} approved, {d['cred_pending']} pending, {d['cred_not_started']} not started)
- EDI: {d['edi_total']} total ({d['edi_live']} live)
- User Production today: {d.get('production_submissions_today', 0)} total ({d.get('production_logs_today', 0)} log entries, {d.get('production_files_today', 0)} file uploads)
- Serving {d['total_clients']} clients | {d['today_actions']} system actions today

Industry benchmarks for context:
- Clean claim rate should be ≥95%, denial rate ≤5%, net collection ≥95%
- Avg days to pay: <30 days is excellent, 30-45 acceptable, >45 needs attention
- AR >90 days should be <15% of total AR

Write a 4-6 sentence executive summary that:
1. Highlights the overall financial health of the practice
2. Calls out any red flags (high denial rate, aging AR, SLA breaches)
3. Notes today's productivity (claims flow, payments received)
4. Provides one key recommendation for tomorrow
Keep it professional and data-driven. No greeting, no bullet points — paragraph form only.
"""
    if not OPENAI_API_KEY:
        return _rule_based_account_summary(d)

    try:
        from openai import OpenAI
        client = OpenAI(api_key=OPENAI_API_KEY)
        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "system", "content": "You are a healthcare RCM executive."},
                      {"role": "user", "content": prompt}],
            max_tokens=400,
            temperature=0.7,
        )
        return resp.choices[0].message.content.strip()
    except Exception as e:
        log.error(f"OpenAI account summary failed: {e}")
        return _rule_based_account_summary(d)


def _rule_based_account_summary(d: dict) -> str:
    """Fallback account summary without AI."""
    flags = []
    if d["denial_rate"] > 5:
        flags.append(f"denial rate of {d['denial_rate']}% exceeds the 5% industry target")
    if d["clean_claim_rate"] < 95:
        flags.append(f"clean claim rate of {d['clean_claim_rate']}% is below the 95% benchmark")
    if d["avg_days_to_pay"] > 45:
        flags.append(f"average days to pay of {d['avg_days_to_pay']} exceeds the 45-day threshold")
    if d["sla_breaches"] > 0:
        flags.append(f"{d['sla_breaches']} SLA breaches require attention")

    total_aging = sum(d["ar_aging"].values())
    if total_aging > 0 and d["ar_aging"]["90_plus"] / total_aging > 0.15:
        flags.append(f"AR >90 days represents {d['ar_aging']['90_plus']/total_aging*100:.0f}% of outstanding balances")

    summary = (
        f"The practice manages ${d['total_ar']:,.2f} in total accounts receivable across "
        f"{d['active_claims']} active claims with a net collection rate of {d['net_collection_rate']}%. "
        f"Today saw {d['submitted_today']} claims submitted, {d['paid_today']} paid, and "
        f"${d['payments_today']:,.2f} in payments posted. "
    )

    if flags:
        summary += "Areas needing attention: " + "; ".join(flags) + ". "
    else:
        summary += "All key performance indicators are within healthy ranges. "

    if d["cred_pending"] > 0:
        summary += f"Credentialing has {d['cred_pending']} applications pending completion. "

    summary += "Continue monitoring AR aging and prioritize follow-up on accounts approaching SLA deadlines."
    return summary


def send_daily_account_summary():
    """
    Compile and send the Overall Account Summary email + SMS.
    Called by the scheduler at 6 PM EST daily.
    Includes per-user production breakdown.
    """
    try:
        from app.client_db import get_daily_account_summary, get_user_production_snapshot
        d = get_daily_account_summary()
        prod_snapshot = get_user_production_snapshot()  # today's production
    except Exception as e:
        log.error(f"Failed to fetch daily account summary data: {e}")
        return

    now = datetime.now()
    date_str = now.strftime("%m/%d/%Y")
    date_str_long = now.strftime("%B %d, %Y")
    day_of_week = now.strftime("%A")

    # AI summary
    ai_summary = _generate_account_ai_summary(d, date_str_long)

    # ── Per-user production rows for email ──
    user_prod_rows_html = ""
    prod_users = prod_snapshot.get("user_stats", [])
    file_uploads = prod_snapshot.get("file_uploads", {})
    if prod_users:
        for u in prod_users:
            files = file_uploads.get(u["username"], 0)
            cats = u["categories"].replace(",", ", ") if u["categories"] else "—"
            user_prod_rows_html += f"""
            <tr>
                <td style="padding:8px 12px;border-bottom:1px solid #f1f5f9;font-size:13px;font-weight:600">{u['username'].title()}</td>
                <td style="padding:8px 12px;border-bottom:1px solid #f1f5f9;text-align:center;font-size:13px">{u['entry_count']}</td>
                <td style="padding:8px 12px;border-bottom:1px solid #f1f5f9;text-align:center;font-size:13px;font-weight:700;color:#2563eb">{u['total_hours']}h</td>
                <td style="padding:8px 12px;border-bottom:1px solid #f1f5f9;text-align:center;font-size:13px">{u['total_qty']}</td>
                <td style="padding:8px 12px;border-bottom:1px solid #f1f5f9;text-align:center;font-size:13px">{files}</td>
                <td style="padding:8px 12px;border-bottom:1px solid #f1f5f9;font-size:11px;color:#64748b">{cats}</td>
            </tr>"""
    else:
        user_prod_rows_html = """<tr><td colspan="6" style="padding:16px;text-align:center;color:#94a3b8;font-size:13px">
            ⚠️ No production data logged today. Jessica & RCM should submit daily work entries.
        </td></tr>"""

    # ── Production detail entries for email ──
    prod_detail_html = ""
    for entry in prod_snapshot.get("entries", [])[:20]:
        prod_detail_html += f"""
        <tr>
            <td style="padding:4px 8px;font-size:11px;border-bottom:1px solid #f8fafc">{entry['username'].title()}</td>
            <td style="padding:4px 8px;font-size:11px;border-bottom:1px solid #f8fafc">{entry['category']}</td>
            <td style="padding:4px 8px;font-size:11px;border-bottom:1px solid #f8fafc">{entry['task_description'][:50]}</td>
            <td style="padding:4px 8px;font-size:11px;border-bottom:1px solid #f8fafc;text-align:center">{entry['quantity']}</td>
            <td style="padding:4px 8px;font-size:11px;border-bottom:1px solid #f8fafc;text-align:center">{entry['time_spent']}h</td>
        </tr>"""

    # ── Status distribution rows ──
    status_rows_html = ""
    for status, count in sorted(d.get("status_distribution", {}).items(), key=lambda x: -x[1]):
        if status == "Paid":
            color = "#22c55e"
        elif status in ("Denied", "Appeals"):
            color = "#ef4444"
        elif status in ("Submitted", "In Progress"):
            color = "#3b82f6"
        else:
            color = "#64748b"
        status_rows_html += f"""
        <tr>
            <td style="padding:6px 12px;border-bottom:1px solid #f1f5f9;font-size:13px">{status}</td>
            <td style="padding:6px 12px;border-bottom:1px solid #f1f5f9;text-align:right;font-weight:700;font-size:13px;color:{color}">{count:,}</td>
        </tr>"""

    # ── Payor rows ──
    payor_rows_html = ""
    for p in d.get("top_payors", [])[:8]:
        payor_rows_html += f"""
        <tr>
            <td style="padding:6px 12px;border-bottom:1px solid #f1f5f9;font-size:12px">{p['payor']}</td>
            <td style="padding:6px 12px;border-bottom:1px solid #f1f5f9;text-align:right;font-size:12px">{p['count']:,}</td>
            <td style="padding:6px 12px;border-bottom:1px solid #f1f5f9;text-align:right;font-size:12px">{_fmt_money(p['charges'])}</td>
        </tr>"""

    # ── Credentialing status rows ──
    cred_rows_html = ""
    for status, count in sorted(d.get("cred_stats", {}).items(), key=lambda x: -x[1]):
        cred_rows_html += f'<span style="display:inline-block;background:#f1f5f9;border-radius:6px;padding:4px 10px;margin:2px;font-size:12px"><b>{count}</b> {status}</span>'

    # ── EDI status rows ──
    edi_rows_html = ""
    for status, count in sorted(d.get("edi_stats", {}).items(), key=lambda x: -x[1]):
        edi_rows_html += f'<span style="display:inline-block;background:#f1f5f9;border-radius:6px;padding:4px 10px;margin:2px;font-size:12px"><b>{count}</b> {status}</span>'

    # ── AR Aging bar ──
    aging = d.get("ar_aging", {})
    total_aging = sum(aging.values()) or 1

    html_body = f"""
    <html>
    <body style="font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; padding: 20px; color: #1e293b; background: #f8fafc;">
        <div style="max-width: 700px; margin: 0 auto; border: 1px solid #e2e8f0; border-radius: 12px; overflow: hidden; background: white;">

            <!-- HEADER -->
            <div style="background: linear-gradient(135deg, #1e3a5f, #2563eb); padding: 28px 32px;">
                <h1 style="color: white; margin: 0; font-size: 24px; font-weight: 800; letter-spacing: 0.5px;">📋 MEDPHARMA DAILY REPORT</h1>
                <p style="color: rgba(255,255,255,0.85); margin: 8px 0 0; font-size: 15px; font-weight: 500;">{day_of_week}, {date_str} — Daily Account Summary</p>
                <p style="color: rgba(255,255,255,0.65); margin: 4px 0 0; font-size: 12px;">MedPharma Revenue Cycle Management — {d['total_clients']} Active Client{'s' if d['total_clients']!=1 else ''}</p>
            </div>

            <div style="padding: 28px 32px;">

                <!-- AI EXECUTIVE SUMMARY -->
                <div style="background:linear-gradient(135deg,#ede9fe,#e0e7ff);border-left:4px solid #6366f1;border-radius:8px;padding:18px 20px;margin-bottom:28px;">
                    <div style="font-size:12px;font-weight:800;text-transform:uppercase;color:#4338ca;letter-spacing:1px;margin-bottom:8px;">🤖 AI Executive Summary</div>
                    <div style="font-size:13px;line-height:1.7;color:#1e293b;">{ai_summary}</div>
                </div>

                <!-- FINANCIAL KPIs -->
                <div style="font-size:14px;font-weight:800;color:#1e293b;text-transform:uppercase;letter-spacing:0.5px;padding-bottom:8px;border-bottom:2px solid #1e293b;margin-bottom:16px;">
                    💰 Financial Overview
                </div>
                <div style="display:flex;gap:12px;margin-bottom:24px;flex-wrap:wrap;">
                    <div style="background:#f0fdf4;border:1px solid #bbf7d0;border-radius:10px;padding:14px 16px;flex:1;min-width:120px;text-align:center">
                        <div style="font-size:22px;font-weight:800;color:#15803d">{_fmt_money(d['total_ar'])}</div>
                        <div style="font-size:10px;font-weight:700;color:#16a34a;text-transform:uppercase">Total AR</div>
                    </div>
                    <div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:10px;padding:14px 16px;flex:1;min-width:120px;text-align:center">
                        <div style="font-size:22px;font-weight:800;color:#2563eb">{_fmt_money(d['payments_mtd'])}</div>
                        <div style="font-size:10px;font-weight:700;color:#3b82f6;text-transform:uppercase">Payments MTD</div>
                    </div>
                    <div style="background:#fefce8;border:1px solid #fde68a;border-radius:10px;padding:14px 16px;flex:1;min-width:120px;text-align:center">
                        <div style="font-size:22px;font-weight:800;color:#ca8a04">{d['net_collection_rate']}%</div>
                        <div style="font-size:10px;font-weight:700;color:#d97706;text-transform:uppercase">Net Collection</div>
                    </div>
                    <div style="background:#fef2f2;border:1px solid #fecaca;border-radius:10px;padding:14px 16px;flex:1;min-width:120px;text-align:center">
                        <div style="font-size:22px;font-weight:800;color:#dc2626">{d['denial_rate']}%</div>
                        <div style="font-size:10px;font-weight:700;color:#ef4444;text-transform:uppercase">Denial Rate</div>
                    </div>
                </div>

                <!-- CLAIMS KPIs -->
                <div style="font-size:14px;font-weight:800;color:#1e293b;text-transform:uppercase;letter-spacing:0.5px;padding-bottom:8px;border-bottom:2px solid #1e293b;margin-bottom:16px;">
                    📄 Claims Overview
                </div>
                <div style="display:flex;gap:12px;margin-bottom:12px;flex-wrap:wrap;">
                    <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:12px 14px;flex:1;min-width:100px;text-align:center">
                        <div style="font-size:24px;font-weight:800;color:#1e293b">{d['total_claims']:,}</div>
                        <div style="font-size:10px;font-weight:600;color:#64748b;text-transform:uppercase">Total Claims</div>
                    </div>
                    <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:12px 14px;flex:1;min-width:100px;text-align:center">
                        <div style="font-size:24px;font-weight:800;color:#2563eb">{d['active_claims']:,}</div>
                        <div style="font-size:10px;font-weight:600;color:#64748b;text-transform:uppercase">Active</div>
                    </div>
                    <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:12px 14px;flex:1;min-width:100px;text-align:center">
                        <div style="font-size:24px;font-weight:800;color:#22c55e">{d['claims_paid']:,}</div>
                        <div style="font-size:10px;font-weight:600;color:#64748b;text-transform:uppercase">Paid</div>
                    </div>
                    <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:12px 14px;flex:1;min-width:100px;text-align:center">
                        <div style="font-size:24px;font-weight:800;color:#ef4444">{d['claims_denied']:,}</div>
                        <div style="font-size:10px;font-weight:600;color:#64748b;text-transform:uppercase">Denied</div>
                    </div>
                </div>

                <!-- TODAY'S ACTIVITY -->
                <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:14px;margin-bottom:12px;">
                    <div style="font-size:12px;font-weight:700;color:#64748b;text-transform:uppercase;margin-bottom:8px;">Today's Activity</div>
                    <div style="display:flex;gap:20px;flex-wrap:wrap;">
                        <div><span style="font-weight:800;color:#2563eb;font-size:18px">{d['submitted_today']}</span> <span style="font-size:12px;color:#64748b">Submitted</span></div>
                        <div><span style="font-weight:800;color:#22c55e;font-size:18px">{d['paid_today']}</span> <span style="font-size:12px;color:#64748b">Paid</span></div>
                        <div><span style="font-weight:800;color:#ef4444;font-size:18px">{d['denied_today']}</span> <span style="font-size:12px;color:#64748b">Denied</span></div>
                        <div><span style="font-weight:800;color:#16a34a;font-size:18px">{_fmt_money(d['payments_today'])}</span> <span style="font-size:12px;color:#64748b">Payments</span></div>
                        <div><span style="font-weight:800;color:#7c3aed;font-size:18px">{d.get('production_submissions_today', 0)}</span> <span style="font-size:12px;color:#64748b">User Production</span></div>
                    </div>
                </div>

                <!-- PERFORMANCE METRICS -->
                <div style="display:flex;gap:12px;margin-bottom:24px;flex-wrap:wrap;">
                    <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:12px 14px;flex:1;min-width:100px;text-align:center">
                        <div style="font-size:18px;font-weight:800;color:#1e293b">{d['clean_claim_rate']}%</div>
                        <div style="font-size:10px;font-weight:600;color:#64748b;text-transform:uppercase">Clean Claim</div>
                    </div>
                    <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:12px 14px;flex:1;min-width:100px;text-align:center">
                        <div style="font-size:18px;font-weight:800;color:#1e293b">{d['avg_days_to_pay']}</div>
                        <div style="font-size:10px;font-weight:600;color:#64748b;text-transform:uppercase">Avg Days to Pay</div>
                    </div>
                    <div style="background:{'#fee2e2' if d['sla_breaches']>0 else '#f8fafc'};border:1px solid {'#fecaca' if d['sla_breaches']>0 else '#e2e8f0'};border-radius:8px;padding:12px 14px;flex:1;min-width:100px;text-align:center">
                        <div style="font-size:18px;font-weight:800;color:{'#dc2626' if d['sla_breaches']>0 else '#1e293b'}">{d['sla_breaches']}</div>
                        <div style="font-size:10px;font-weight:600;color:#64748b;text-transform:uppercase">SLA Breaches</div>
                    </div>
                    <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:12px 14px;flex:1;min-width:100px;text-align:center">
                        <div style="font-size:18px;font-weight:800;color:#1e293b">{d['today_actions']}</div>
                        <div style="font-size:10px;font-weight:600;color:#64748b;text-transform:uppercase">Actions Today</div>
                    </div>
                </div>

                <!-- AR AGING -->
                <div style="font-size:14px;font-weight:800;color:#1e293b;text-transform:uppercase;letter-spacing:0.5px;padding-bottom:8px;border-bottom:2px solid #1e293b;margin-bottom:12px;">
                    ⏳ AR Aging Distribution
                </div>
                <div style="display:flex;gap:8px;margin-bottom:8px;">
                    <div style="flex:{aging.get('current',0)/total_aging};background:#22c55e;height:14px;border-radius:4px 0 0 4px" title="Current"></div>
                    <div style="flex:{aging.get('31_60',0)/total_aging};background:#f59e0b;height:14px" title="31-60"></div>
                    <div style="flex:{aging.get('61_90',0)/total_aging};background:#f97316;height:14px" title="61-90"></div>
                    <div style="flex:{aging.get('90_plus',0)/total_aging};background:#ef4444;height:14px;border-radius:0 4px 4px 0" title="90+"></div>
                </div>
                <div style="display:flex;gap:12px;margin-bottom:24px;flex-wrap:wrap;font-size:12px;">
                    <div><span style="display:inline-block;width:10px;height:10px;background:#22c55e;border-radius:2px;margin-right:4px"></span>Current: {_fmt_money(aging.get('current',0))}</div>
                    <div><span style="display:inline-block;width:10px;height:10px;background:#f59e0b;border-radius:2px;margin-right:4px"></span>31-60: {_fmt_money(aging.get('31_60',0))}</div>
                    <div><span style="display:inline-block;width:10px;height:10px;background:#f97316;border-radius:2px;margin-right:4px"></span>61-90: {_fmt_money(aging.get('61_90',0))}</div>
                    <div><span style="display:inline-block;width:10px;height:10px;background:#ef4444;border-radius:2px;margin-right:4px"></span>90+: {_fmt_money(aging.get('90_plus',0))}</div>
                </div>

                <!-- STATUS DISTRIBUTION + TOP PAYORS side-by-side -->
                <div style="display:flex;gap:16px;margin-bottom:24px;flex-wrap:wrap;">
                    <div style="flex:1;min-width:200px">
                        <div style="font-size:12px;font-weight:700;color:#64748b;text-transform:uppercase;margin-bottom:8px;">Claim Status Distribution</div>
                        <table style="width:100%;border-collapse:collapse">{status_rows_html}</table>
                    </div>
                    <div style="flex:1;min-width:200px">
                        <div style="font-size:12px;font-weight:700;color:#64748b;text-transform:uppercase;margin-bottom:8px;">Top Payors</div>
                        <table style="width:100%;border-collapse:collapse">
                            <thead><tr>
                                <th style="padding:4px 12px;text-align:left;font-size:10px;color:#94a3b8;text-transform:uppercase">Payor</th>
                                <th style="padding:4px 12px;text-align:right;font-size:10px;color:#94a3b8;text-transform:uppercase">Claims</th>
                                <th style="padding:4px 12px;text-align:right;font-size:10px;color:#94a3b8;text-transform:uppercase">Charges</th>
                            </tr></thead>
                            <tbody>{payor_rows_html}</tbody>
                        </table>
                    </div>
                </div>

                <!-- CREDENTIALING / EDI -->
                <div style="font-size:14px;font-weight:800;color:#1e293b;text-transform:uppercase;letter-spacing:0.5px;padding-bottom:8px;border-bottom:2px solid #1e293b;margin-bottom:16px;">
                    🏥 Credentialing & EDI
                </div>
                <div style="display:flex;gap:12px;margin-bottom:12px;flex-wrap:wrap;">
                    <div style="flex:1;min-width:140px;background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:14px;">
                        <div style="font-size:12px;font-weight:700;color:#64748b;text-transform:uppercase;margin-bottom:6px;">Credentialing ({d['cred_total']})</div>
                        <div>{cred_rows_html or '<span style="font-size:12px;color:#94a3b8">No records</span>'}</div>
                    </div>
                    <div style="flex:1;min-width:140px;background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:14px;">
                        <div style="font-size:12px;font-weight:700;color:#64748b;text-transform:uppercase;margin-bottom:6px;">EDI/ERA/EFT ({d['edi_total']})</div>
                        <div>{edi_rows_html or '<span style="font-size:12px;color:#94a3b8">No records</span>'}</div>
                    </div>
                </div>

                <!-- MTD COMPARISON -->
                <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:14px;margin-bottom:24px;">
                    <div style="font-size:12px;font-weight:700;color:#64748b;text-transform:uppercase;margin-bottom:8px;">Month-to-Date Summary</div>
                    <div style="display:flex;gap:20px;flex-wrap:wrap;font-size:13px;">
                        <div>📤 <b>{d['submitted_mtd']}</b> Submitted</div>
                        <div>✅ <b>{d['paid_mtd']}</b> Paid</div>
                        <div>❌ <b>{d['denied_mtd']}</b> Denied</div>
                        <div>💵 <b>{_fmt_money(d['payments_mtd'])}</b> Collected</div>
                        <div>📅 <b>{_fmt_money(d['payments_ytd'])}</b> YTD</div>
                    </div>
                </div>

                <!-- USER PRODUCTION SNAPSHOT -->
                <div style="font-size:14px;font-weight:800;color:#1e293b;text-transform:uppercase;letter-spacing:0.5px;padding-bottom:8px;border-bottom:2px solid #7c3aed;margin-bottom:16px;">
                    👥 User Production — {date_str}
                </div>
                <table style="width:100%;border-collapse:collapse;margin-bottom:12px;">
                    <thead>
                        <tr style="background:#f8fafc;">
                            <th style="padding:8px 12px;text-align:left;font-size:11px;font-weight:700;color:#64748b;text-transform:uppercase">Employee</th>
                            <th style="padding:8px 12px;text-align:center;font-size:11px;font-weight:700;color:#64748b;text-transform:uppercase">Entries</th>
                            <th style="padding:8px 12px;text-align:center;font-size:11px;font-weight:700;color:#64748b;text-transform:uppercase">Hours</th>
                            <th style="padding:8px 12px;text-align:center;font-size:11px;font-weight:700;color:#64748b;text-transform:uppercase">Qty</th>
                            <th style="padding:8px 12px;text-align:center;font-size:11px;font-weight:700;color:#64748b;text-transform:uppercase">Files</th>
                            <th style="padding:8px 12px;text-align:left;font-size:11px;font-weight:700;color:#64748b;text-transform:uppercase">Categories</th>
                        </tr>
                    </thead>
                    <tbody>{user_prod_rows_html}</tbody>
                </table>

                {'<div style="margin-bottom:24px;"><div style="font-size:11px;font-weight:700;color:#64748b;text-transform:uppercase;margin-bottom:6px;">Work Detail</div><table style="width:100%;border-collapse:collapse;"><thead><tr style="background:#f8fafc;"><th style="padding:4px 8px;font-size:10px;color:#94a3b8;text-transform:uppercase;text-align:left">User</th><th style="padding:4px 8px;font-size:10px;color:#94a3b8;text-transform:uppercase;text-align:left">Category</th><th style="padding:4px 8px;font-size:10px;color:#94a3b8;text-transform:uppercase;text-align:left">Description</th><th style="padding:4px 8px;font-size:10px;color:#94a3b8;text-transform:uppercase;text-align:center">Qty</th><th style="padding:4px 8px;font-size:10px;color:#94a3b8;text-transform:uppercase;text-align:center">Time</th></tr></thead><tbody>' + prod_detail_html + '</tbody></table></div>' if prod_detail_html else ''}

                <hr style="border: none; border-top: 1px solid #e2e8f0; margin: 20px 0;">
                <p style="font-size: 11px; color: #94a3b8; text-align: center; margin: 0;">
                    MedPharma Daily Report — {date_str}
                </p>
            </div>
        </div>
    </body>
    </html>"""

    # ── Plain text version ──
    # Build per-user production lines
    prod_text_lines = []
    if prod_users:
        for u in prod_users:
            files = file_uploads.get(u["username"], 0)
            prod_text_lines.append(
                f"  {u['username'].title():12s}  {u['entry_count']} entries  |  {u['total_hours']}h  |  {u['total_qty']} qty  |  {files} files  |  {u['categories']}"
            )
    else:
        prod_text_lines.append("  ⚠️ No production data logged today")

    body_lines = [
        "═══════════════════════════════════════════",
        "     MEDPHARMA DAILY REPORT",
        f"     {day_of_week}, {date_str}",
        "═══════════════════════════════════════════",
        "",
        "AI EXECUTIVE SUMMARY:",
        ai_summary,
        "",
        "───── FINANCIAL ─────",
        f"  Total AR:          {_fmt_money(d['total_ar'])}",
        f"  Payments Today:    {_fmt_money(d['payments_today'])}",
        f"  Payments MTD:      {_fmt_money(d['payments_mtd'])}",
        f"  Payments YTD:      {_fmt_money(d['payments_ytd'])}",
        f"  Net Collection:    {d['net_collection_rate']}%",
        f"  Denial Rate:       {d['denial_rate']}%",
        f"  Clean Claim Rate:  {d['clean_claim_rate']}%",
        f"  Avg Days to Pay:   {d['avg_days_to_pay']}",
        "",
        "───── CLAIMS ─────",
        f"  Total: {d['total_claims']:,}  |  Active: {d['active_claims']:,}  |  Paid: {d['claims_paid']:,}  |  Denied: {d['claims_denied']:,}",
        f"  Today: {d['submitted_today']} submitted, {d['paid_today']} paid, {d['denied_today']} denied",
        f"  MTD:   {d['submitted_mtd']} submitted, {d['paid_mtd']} paid, {d['denied_mtd']} denied",
        "",
        "───── AR AGING ─────",
        f"  Current:  {_fmt_money(aging.get('current',0))}",
        f"  31-60:    {_fmt_money(aging.get('31_60',0))}",
        f"  61-90:    {_fmt_money(aging.get('61_90',0))}",
        f"  90+:      {_fmt_money(aging.get('90_plus',0))}",
        "",
        "───── CREDENTIALING/EDI ─────",
        f"  Credentialing: {d['cred_total']} ({d['cred_approved']} approved, {d['cred_pending']} pending)",
        f"  EDI:           {d['edi_total']} ({d['edi_live']} live)",
        "",
        f"───── USER PRODUCTION — {date_str} ─────",
    ] + prod_text_lines + [
        "",
        f"  SLA Breaches: {d['sla_breaches']}  |  System Actions Today: {d['today_actions']}",
    ]
    body = "\n".join(body_lines)

    subject = f"MedPharma Daily Report — {date_str} — AR {_fmt_money(d['total_ar'])}"

    threading.Thread(target=_send_email, args=(subject, body, html_body), daemon=True).start()
    log.info(f"MedPharma Daily Report email queued: {date_str}, AR {_fmt_money(d['total_ar'])}, "
             f"{d['total_claims']} claims, {len(prod_users)} users logged production")


# ═══════════════════════════════════════════════════════════════════════════
#  SCHEDULER — 5:30 PM & 6 PM EST daily
# ═══════════════════════════════════════════════════════════════════════════

# User emails for individual reminders
USER_EMAILS = {
    "jessica": "jessica@medprosc.com",
    "rcm": "rcm@medprosc.com",
}

_scheduler_started = False


def send_production_reminders():
    """
    Send reminder emails to jessica@medprosc.com and rcm@medprosc.com
    at 5:30 PM EST if they have NOT uploaded any production data today.
    """
    from datetime import date
    today = date.today().isoformat()

    for username, email in USER_EMAILS.items():
        try:
            from app.client_db import has_production_data_today
            if has_production_data_today(username, today):
                log.info(f"Production reminder skipped for {username} — data already uploaded for {today}")
                continue

            subject = f"⏰ Reminder: Upload Your Daily Production — {datetime.now().strftime('%B %d, %Y')}"
            html_body = f"""
            <html>
            <body style="font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; padding: 20px; color: #1e293b; background: #f8fafc;">
                <div style="max-width: 560px; margin: 0 auto; border: 1px solid #e2e8f0; border-radius: 12px; overflow: hidden; background: white;">
                    <div style="background: linear-gradient(135deg, #f59e0b, #d97706); padding: 24px 28px;">
                        <h1 style="color: white; margin: 0; font-size: 20px; font-weight: 800;">⏰ Daily Production Reminder</h1>
                        <p style="color: rgba(255,255,255,0.9); margin: 6px 0 0; font-size: 14px;">
                            {datetime.now().strftime('%A, %B %d, %Y')} — 5:30 PM EST
                        </p>
                    </div>
                    <div style="padding: 24px 28px;">
                        <p style="font-size: 15px; line-height: 1.7; margin: 0 0 16px;">
                            Hi <strong>{username.title()}</strong>,
                        </p>
                        <div style="background: #fef3c7; border-left: 4px solid #f59e0b; border-radius: 8px; padding: 16px 20px; margin-bottom: 20px;">
                            <p style="font-size: 14px; line-height: 1.6; margin: 0; color: #92400e;">
                                📋 You have <strong>not uploaded</strong> any production data for today yet.
                                Please log your daily work entries or upload your production report before end of day.
                            </p>
                        </div>
                        <p style="font-size: 13px; color: #64748b; line-height: 1.6; margin: 0 0 20px;">
                            Log in to <a href="https://medpharmasc.com" style="color: #2563eb; text-decoration: none; font-weight: 600;">MedPharma Hub</a>
                            and go to <strong>User Production</strong> to submit your work for today. You can either
                            log individual tasks or upload an Excel/PDF report.
                        </p>
                        <div style="text-align: center; margin: 24px 0;">
                            <a href="https://medpharmasc.com" style="display:inline-block;background:#2563eb;color:white;padding:12px 32px;
                                border-radius:8px;text-decoration:none;font-weight:700;font-size:14px;">
                                Log In & Upload Production
                            </a>
                        </div>
                        <hr style="border: none; border-top: 1px solid #e2e8f0; margin: 20px 0;">
                        <p style="font-size: 11px; color: #94a3b8; text-align: center; margin: 0;">
                            This is an automated reminder from MedPharma Hub. If you've already submitted your data, please disregard.
                        </p>
                    </div>
                </div>
            </body>
            </html>"""
            body = (f"Hi {username.title()}, you have not uploaded production data for today ({today}). "
                    f"Please log in to MedPharma Hub and submit your daily work before end of day.")

            _send_email_to(email, subject, body, html_body)
            log.info(f"Production reminder sent to {username} ({email})")
        except Exception as e:
            log.error(f"Failed to send production reminder to {username}: {e}")


def _send_email_to(to_email: str, subject: str, body: str, html_body: str = "",
                   attachments: list = None) -> tuple[bool, str]:
    """Send email to a specific recipient via SendGrid v3 API.
    Uses _live_config() to read credentials fresh.

    ``attachments`` (optional) is a list of dicts with keys:
        - filename: str           (e.g. "report.xlsx")
        - content:  bytes         (raw file bytes; will be base64-encoded)
        - mime:     str           (e.g. "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    Returns (sent, via) where ``via`` is 'sendgrid', 'smtp', or a reason
    string when nothing went out. The previous signature returned None;
    callers that ignore the tuple still work fine.
    """
    if not to_email:
        return False, "missing recipient"
    cfg = _live_config()
    attachments = attachments or []

    # Primary: SendGrid
    sg_key = cfg["SENDGRID_API_KEY"]
    sg_from = cfg["SENDGRID_FROM"]
    if sg_key:
        try:
            import httpx, base64 as _b64
            content = []
            if body:
                content.append({"type": "text/plain", "value": body})
            if html_body:
                content.append({"type": "text/html", "value": html_body})
            if not content:
                content.append({"type": "text/plain", "value": "(no content)"})

            payload = {
                "personalizations": [{"to": [{"email": to_email}]}],
                "from": {"email": sg_from, "name": "MedPharma Hub"},
                "subject": subject,
                "content": content,
            }
            if attachments:
                payload["attachments"] = [
                    {
                        "content": _b64.b64encode(a["content"]).decode("ascii"),
                        "filename": a["filename"],
                        "type": a.get("mime", "application/octet-stream"),
                        "disposition": "attachment",
                    }
                    for a in attachments if a.get("content")
                ]
            resp = httpx.post(
                "https://api.sendgrid.com/v3/mail/send",
                json=payload,
                headers={
                    "Authorization": f"Bearer {sg_key}",
                    "Content-Type": "application/json",
                },
                timeout=30,
            )
            if resp.status_code in (200, 202):
                log.info(f"Email sent to {to_email}: {subject}")
                return True, "sendgrid"
            log.error(f"SendGrid failed ({resp.status_code}): {resp.text}")
        except Exception as e:
            log.error(f"Failed to send email to {to_email} via SendGrid: {e}")

    # Fallback: SMTP
    smtp_h = cfg["SMTP_HOST"]
    smtp_p = cfg["SMTP_PORT"]
    smtp_u = cfg["SMTP_USER"]
    smtp_pw = cfg["SMTP_PASS"]
    if not smtp_h or not smtp_u or not smtp_pw:
        log.error("Email skipped — no working provider configured (SendGrid/SMTP)")
        return False, "no provider configured (SendGrid/SMTP env vars missing)"

    try:
        # If we have attachments, use a mixed multipart wrapper so the
        # text+html alternative still renders inline while the file rides
        # alongside it.
        if attachments:
            from email.mime.base import MIMEBase
            from email import encoders as _encoders
            outer = MIMEMultipart("mixed")
            outer["Subject"] = subject
            outer["From"] = smtp_u or sg_from
            outer["To"] = to_email
            alt = MIMEMultipart("alternative")
            alt.attach(MIMEText(body or "(no content)", "plain"))
            if html_body:
                alt.attach(MIMEText(html_body, "html"))
            outer.attach(alt)
            for a in attachments:
                if not a.get("content"):
                    continue
                part = MIMEBase(*a.get("mime", "application/octet-stream").split("/", 1))
                part.set_payload(a["content"])
                _encoders.encode_base64(part)
                part.add_header(
                    "Content-Disposition",
                    f'attachment; filename="{a["filename"]}"',
                )
                outer.attach(part)
            msg = outer
        else:
            msg = MIMEMultipart("alternative")
            msg["Subject"] = subject
            msg["From"] = smtp_u or sg_from
            msg["To"] = to_email
            plain = body or "(no content)"
            msg.attach(MIMEText(plain, "plain"))
            if html_body:
                msg.attach(MIMEText(html_body, "html"))

        with smtplib.SMTP(smtp_h, smtp_p, timeout=30) as server:
            server.starttls()
            server.login(smtp_u, smtp_pw)
            server.sendmail(msg["From"], [to_email], msg.as_string())
        log.info(f"Email sent via SMTP to {to_email}: {subject}")
        return True, "smtp"
    except Exception as e:
        log.error(f"Failed to send email to {to_email} via SMTP: {e}")
        return False, f"smtp error: {e}"


def send_team_progress_reports():
    """
    Send consolidated progress reports for tracked users (e.g., Jessica/RCM)
    to configured owner recipients at 7:00 PM Eastern.
    """
    users = [u for u in NOTIFY_ON_USERS if u and u != "*"]
    if not users:
        log.info("Team progress dispatch skipped — no explicit tracked users configured")
        return

    sent = 0
    for username in sorted(users):
        try:
            flush_and_notify(username)
            sent += 1
        except Exception as e:
            log.error(f"Team progress dispatch failed for {username}: {e}")
    log.info(f"Team progress dispatch completed for {sent} user(s): {', '.join(sorted(users))}")


# ─── End-of-Day Team Report (internal — Lexi only) ───────────────────────

# Default recipients for the EOD per-user / per-client breakdown email.
# Override at runtime with the EOD_REPORT_EMAIL setting (Email Setup tab) or
# the EOD_REPORT_EMAIL env var (comma-separated).
EOD_REPORT_DEFAULT_RECIPIENTS = ["lexi@medprosc.com", "eric@medprosc.com"]

# Users to suppress from the per-user breakdown (admins, observers, etc).
# Override with EOD_REPORT_EXCLUDE_USERS env (comma-separated usernames).
_EOD_EXCLUDE_USERS_DEFAULT = {
    "lexi@medprosc.com", "admin@medprosc.com", "admin",
    "eric@medprosc.com", "eric",
}

# MedPharma brand assets — used in every email header so it always
# looks like a real product, not a debug dump.
_MEDPHARMA_LOGO_URL = "https://medpharmasc.com/wp-content/uploads/2024/11/IMG_2392.png"
_MEDPHARMA_SITE_URL = "https://medpharmasc.com"
_MEDPHARMA_HUB_URL  = os.getenv("HUB_BASE_URL", "https://medpharma-hub.onrender.com/hub")


def _eod_recipients() -> list[str]:
    # Prefer the DB setting (managed from the Email Setup tab) so recipients
    # can be changed without a redeploy; fall back to env, then defaults.
    raw = ""
    try:
        from app.client_db import get_app_setting
        raw = (get_app_setting("EOD_REPORT_EMAIL") or "").strip()
    except Exception:
        raw = ""
    if not raw:
        raw = os.getenv("EOD_REPORT_EMAIL", "").strip()
    if raw:
        return [e.strip() for e in raw.split(",") if e.strip()]
    return list(EOD_REPORT_DEFAULT_RECIPIENTS)


def _eod_excluded_users() -> set:
    raw = os.getenv("EOD_REPORT_EXCLUDE_USERS", "").strip()
    if raw:
        return {u.strip().lower() for u in raw.split(",") if u.strip()}
    return set(_EOD_EXCLUDE_USERS_DEFAULT)


def _esc_html(s) -> str:
    """Minimal HTML escape — we're building markup with f-strings."""
    return (
        str(s if s is not None else "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


# Friendly per-tab icons (matches the sidebar in client_hub.html).
_EOD_TAB_ICONS = {
    "Claims":        "💼",
    "Credentialing": "🎓",
    "Enrollment":    "📝",
    "EDI":           "🔌",
    "Production":    "⏱️",
    "Documents":     "📁",
    "Notes":         "🗒️",
    "Chat":          "💬",
    "Audit":         "🔍",
    "Pageviews":     "👁️",
}

# Map a tab name → the client `enabled_modules` slug that gates it.
# When a client doesn't have the module enabled, we suppress the column.
_TAB_TO_MODULE = {
    "Claims":        "claims",
    "Credentialing": "credentialing",
    "Enrollment":    "enrollment",
    "EDI":           "edi",
    "Production":    "production",
    "Documents":     "documents",
    "Notes":         "claims",   # notes ride with claims
    "Chat":          "chat",
    "Audit":         None,        # always shown
    "Pageviews":     None,        # always shown
}


def _filter_tabs_for_client(tabs: list, enabled_modules: list) -> list:
    """Drop tabs whose underlying module isn't enabled for this client.
    Tabs with module=None are always kept.
    """
    if not enabled_modules:
        return list(tabs)
    en = {m.lower() for m in enabled_modules}
    out = []
    for t in tabs:
        mod = _TAB_TO_MODULE.get(t)
        if mod is None or mod in en:
            out.append(t)
    return out


def _brand_email_shell(title: str, subtitle: str, accent: str, inner_html: str,
                       footer_note: str = "") -> str:
    """Wraps inner HTML in a consistent MedPharma-branded email frame.

    Uses table-based layout (not flexbox) for maximum email-client
    compatibility (Gmail, Outlook, Apple Mail).
    """
    title = _esc_html(title)
    subtitle = _esc_html(subtitle)
    accent = accent or "#2563eb"
    footer = footer_note or (
        "MedPharma Hub · Auto-generated · "
        f'<a href="{_MEDPHARMA_HUB_URL}" style="color:#94a3b8">Open Hub</a> · '
        f'<a href="{_MEDPHARMA_SITE_URL}" style="color:#94a3b8">medpharmasc.com</a>'
    )
    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{title}</title></head>
<body style="margin:0;padding:0;background:#eef2f7;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Helvetica,Arial,sans-serif;color:#0f172a">
<table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#eef2f7;padding:24px 12px">
  <tr><td align="center">
    <table role="presentation" width="720" cellpadding="0" cellspacing="0" border="0" style="max-width:720px;width:100%;background:#ffffff;border-radius:14px;overflow:hidden;box-shadow:0 8px 30px rgba(15,23,42,.08)">

      <!-- BRAND BAR -->
      <tr><td style="background:linear-gradient(135deg,#0b1733 0%,{accent} 100%);padding:22px 28px">
        <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0">
          <tr>
            <td style="vertical-align:middle">
              <img src="{_MEDPHARMA_LOGO_URL}" alt="MedPharma" height="44" style="display:block;height:44px;width:auto;border:0;outline:none;text-decoration:none">
            </td>
            <td align="right" style="vertical-align:middle">
              <div style="color:#ffffff;font-size:11px;letter-spacing:1.2px;text-transform:uppercase;font-weight:600;opacity:.78">MedPharma SC</div>
              <div style="color:#ffffff;font-size:11px;opacity:.65;margin-top:2px">medpharmasc.com</div>
            </td>
          </tr>
        </table>
        <div style="color:#ffffff;font-size:22px;font-weight:800;letter-spacing:.2px;margin:14px 0 2px">{title}</div>
        <div style="color:rgba(255,255,255,.85);font-size:13px">{subtitle}</div>
      </td></tr>

      <!-- BODY -->
      <tr><td style="padding:24px 28px 28px;background:#ffffff">
{inner_html}
      </td></tr>

      <!-- FOOTER -->
      <tr><td style="background:#0b1733;padding:14px 28px;text-align:center;color:#94a3b8;font-size:11px;line-height:1.5">
        {footer}
      </td></tr>

    </table>
  </td></tr>
</table>
</body></html>"""


def _ts_short(value) -> str:
    """Format a timestamp/datetime/iso-string as 'HH:MM' (24h)."""
    if not value:
        return ""
    if isinstance(value, str):
        # Try ISO-ish first, then common SQL "YYYY-MM-DD HH:MM:SS"
        for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M"):
            try:
                from datetime import datetime as _dt
                return _dt.strptime(value[:19], fmt).strftime("%H:%M")
            except Exception:
                continue
        # As a fallback, slice the time portion if it looks like 'YYYY-MM-DD HH:MM'
        if len(value) >= 16 and value[10] in (" ", "T"):
            return value[11:16]
        return ""
    try:
        return value.strftime("%H:%M")
    except Exception:
        return ""


def _render_eod_report_html(report: dict) -> tuple[str, str]:
    """Render the EOD report dict into (text_body, html_body).

    The report dict comes from app.client_db.get_eod_team_report().

    Design rules (after Lexi feedback 2026-06-10):
      * Branded shell with MedPharma logo header
      * Admins (Lexi/Eric) suppressed by default — operators only
      * Per-client section uses a vertical card (no giant 11-col table)
      * Module-filtered badges so EDI/Credentialing don't show for clients
        who haven't enabled those services
      * Every activity row is timestamped (HH:MM 24h)
    """
    date_iso = report.get("report_date", "")
    try:
        from datetime import datetime as _dt
        d = _dt.strptime(date_iso, "%Y-%m-%d")
        date_long = d.strftime("%A, %B %d, %Y")
    except Exception:
        date_long = date_iso

    headlines = report.get("headlines", {}) or {}
    tabs_all = report.get("tab_keys", []) or list(_EOD_TAB_ICONS.keys())
    users_all = report.get("users", []) or []
    excluded = _eod_excluded_users()

    # Filter out admins/observers (default: Lexi + Eric + admin@).
    users = [
        u for u in users_all
        if (u.get("username") or "").strip().lower() not in excluded
        and (u.get("email") or "").strip().lower() not in excluded
    ]

    # ── plain-text fallback ──
    text_lines = [
        f"MedPharma End-of-Day Team Report — {date_long}",
        "",
        f"Active operators today: {len(users)}",
        f"Claims created: {headlines.get('claims_new', 0)} · updated: {headlines.get('claims_touched', 0)}",
        f"Credentialing new: {headlines.get('cred_new', 0)} · Enrollment new: {headlines.get('enroll_new', 0)} · EDI new: {headlines.get('edi_new', 0)}",
        f"Production entries: {headlines.get('production_rows', 0)} ({headlines.get('production_hours', 0)} hrs)",
        f"Notes: {headlines.get('notes_new', 0)} · Files uploaded: {headlines.get('files_uploaded', 0)} · Chat messages: {headlines.get('chat_messages', 0)} · Audit events: {headlines.get('audit_events', 0)}",
        "",
        "Per-operator breakdown:",
    ]
    for u in users:
        text_lines.append(
            f"\n* {u.get('contact_name') or u.get('username')} "
            f"<{u.get('email','')}> — {u.get('active_hours',0)}h active / "
            f"{u.get('idle_hours',0)}h idle / {u.get('total_actions',0)} actions"
        )
        for cname, cb in (u.get("clients") or {}).items():
            chunks = [f"{k}={v}" for k, v in cb.get("totals", {}).items() if v]
            if not chunks:
                continue
            text_lines.append(f"    - {cname}: {', '.join(chunks)}")
            for it in (cb.get("items") or [])[:6]:
                ts = _ts_short(it.get("ts") or it.get("at") or "")
                ts_prefix = f"[{ts}] " if ts else ""
                text_lines.append(
                    f"        · {ts_prefix}[{it.get('tab')}] {it.get('action','')} — {it.get('title','')}"
                )
    text_body = "\n".join(text_lines)

    # ── HTML headline ribbon (single row, table-based so Outlook is happy) ──
    def _ribbon_cell(label, value, color="#1d4ed8"):
        return (
            f'<td align="center" valign="top" width="20%" style="padding:14px 6px">'
            f'<div style="font-size:10px;color:#64748b;text-transform:uppercase;letter-spacing:.8px;font-weight:700">{label}</div>'
            f'<div style="font-size:24px;color:{color};font-weight:800;margin-top:4px;line-height:1.1">{value}</div>'
            f'</td>'
        )

    ribbon_row1 = (
        _ribbon_cell("Operators", len(users)) +
        _ribbon_cell("Claims new", headlines.get("claims_new", 0)) +
        _ribbon_cell("Claims touched", headlines.get("claims_touched", 0), "#2563eb") +
        _ribbon_cell("Credentialing", headlines.get("cred_new", 0), "#7c3aed") +
        _ribbon_cell("Enrollment", headlines.get("enroll_new", 0), "#7c3aed")
    )
    ribbon_row2 = (
        _ribbon_cell("EDI", headlines.get("edi_new", 0), "#0891b2") +
        _ribbon_cell("Prod hrs", f"{headlines.get('production_hours', 0)}", "#16a34a") +
        _ribbon_cell("Notes", headlines.get("notes_new", 0), "#0f766e") +
        _ribbon_cell("Files", headlines.get("files_uploaded", 0), "#0f766e") +
        _ribbon_cell("Chat", headlines.get("chat_messages", 0), "#db2777")
    )

    ribbon_html = (
        '<table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" '
        'style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;margin:0 0 18px">'
        f'<tr>{ribbon_row1}</tr>'
        '<tr><td colspan="5" style="border-top:1px solid #e2e8f0;padding:0;line-height:0">&nbsp;</td></tr>'
        f'<tr>{ribbon_row2}</tr>'
        '</table>'
    )

    # ── Per-operator cards ──
    def _badges_for(totals_dict, only_tabs=None):
        out = ""
        chips = only_tabs if only_tabs is not None else tabs_all
        for t in chips:
            v = (totals_dict or {}).get(t, 0)
            if not v:
                continue
            out += (
                f'<span style="display:inline-block;background:#eef2ff;'
                f'color:#1d4ed8;border-radius:6px;padding:4px 9px;'
                f'margin:2px 4px 2px 0;font-size:12px;font-weight:600">'
                f'{_EOD_TAB_ICONS.get(t,"")} {t} <b>{v}</b></span>'
            )
        return out

    def _client_block(cname, cb):
        """Render ONE client section inside an operator card.
        Drops modules the client doesn't have enabled.
        """
        enabled = cb.get("enabled_modules") or []
        visible_tabs = _filter_tabs_for_client(tabs_all, enabled)
        # Suppress sections that are entirely zero for this client.
        if not any((cb.get("totals", {}) or {}).get(t, 0) for t in visible_tabs):
            return ""
        badges = _badges_for(cb.get("totals", {}), only_tabs=visible_tabs) or (
            '<span style="color:#94a3b8;font-size:12px;font-style:italic">No tracked work captured</span>'
        )
        # Items (chronologically) — already trimmed/sorted by aggregator.
        items_html = ""
        items_iter = cb.get("items") or []
        # Sort by ts if available
        try:
            items_iter = sorted(items_iter, key=lambda it: (it.get("ts") or "") , reverse=False)
        except Exception:
            pass
        for it in items_iter[:10]:
            tab = it.get("tab", "")
            if visible_tabs and tab not in visible_tabs:
                continue
            ts = _ts_short(it.get("ts") or it.get("at") or "")
            ts_html = (
                f'<span style="display:inline-block;background:#0f172a;color:#fff;'
                f'border-radius:4px;padding:1px 6px;font-size:10px;font-weight:700;'
                f'font-family:Menlo,Consolas,monospace;margin-right:8px;letter-spacing:.5px">{ts}</span>'
                if ts else ''
            )
            items_html += (
                f'<tr><td style="padding:6px 0;border-bottom:1px solid #f1f5f9;font-size:12px;color:#334155">'
                f'{ts_html}'
                f'<span style="color:#1d4ed8;font-weight:600;margin-right:6px">{_EOD_TAB_ICONS.get(tab,"")} {_esc_html(tab)}</span>'
                f'<span style="color:#64748b">{_esc_html(it.get("action","")) }</span> '
                f'— {_esc_html(it.get("title",""))}'
                f'</td></tr>'
            )
        if not items_html:
            items_html = (
                '<tr><td style="padding:8px 0;color:#94a3b8;font-size:12px;font-style:italic">'
                'No itemised changes captured for this client today.</td></tr>'
            )
        return f"""
        <div style="margin:14px 0 0;padding:14px 16px;background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px">
          <div style="font-size:14px;font-weight:800;color:#0f172a;margin-bottom:6px">{_esc_html(cname)}</div>
          <div style="margin:2px 0 10px">{badges}</div>
          <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0">{items_html}</table>
        </div>"""

    cards_html = ""
    if not users:
        cards_html = (
            '<div style="padding:22px;color:#94a3b8;font-size:14px;text-align:center;'
            'background:#f8fafc;border:1px dashed #cbd5e1;border-radius:10px">'
            'No operator activity recorded today.</div>'
        )
    else:
        for u in users:
            uname = _esc_html(u.get("contact_name") or u.get("username", ""))
            email = _esc_html(u.get("email", ""))
            role  = _esc_html((u.get("role") or "staff").title())
            hrs_a = u.get("active_hours", 0)
            hrs_i = u.get("idle_hours", 0)
            acts  = u.get("total_actions", 0)
            first = _ts_short(u.get("first_seen") or "")
            last  = _ts_short(u.get("last_seen") or "")
            session_line = (
                f"{first}–{last}" if first and last else (first or last or "—")
            )
            tab_badges = _badges_for(u.get("totals", {})) or (
                '<span style="color:#94a3b8;font-size:12px;font-style:italic">Logged in but no tracked work</span>'
            )
            highlights = u.get("highlights") or []
            highlights_html = ""
            if highlights:
                hi_list = "".join(
                    f'<li style="font-size:12px;color:#475569;padding:2px 0">{_esc_html(h)}</li>'
                    for h in highlights[:6]
                )
                highlights_html = (
                    f'<div style="margin-top:10px;padding:10px 12px;background:#fff7ed;border-left:3px solid #f59e0b;border-radius:6px">'
                    f'<div style="font-size:11px;color:#92400e;font-weight:800;text-transform:uppercase;letter-spacing:.6px;margin-bottom:4px">Audit highlights</div>'
                    f'<ul style="margin:4px 0 0;padding-left:20px">{hi_list}</ul></div>'
                )
            client_blocks = ""
            for cname, cb in (u.get("clients") or {}).items():
                client_blocks += _client_block(cname, cb)
            if not client_blocks:
                client_blocks = (
                    '<div style="padding:14px;color:#94a3b8;font-size:12px;font-style:italic;'
                    'background:#f8fafc;border:1px dashed #cbd5e1;border-radius:8px;margin-top:12px">'
                    'No per-client work captured — only page visits.</div>'
                )
            cards_html += f"""
            <div style="background:#ffffff;border:1px solid #e2e8f0;border-radius:12px;padding:18px 20px;margin:14px 0;box-shadow:0 1px 3px rgba(15,23,42,.04)">
              <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0">
                <tr>
                  <td style="vertical-align:top">
                    <div style="font-size:17px;font-weight:800;color:#0f172a">{uname} <span style="font-size:10px;color:#94a3b8;font-weight:700;text-transform:uppercase;letter-spacing:.6px;margin-left:6px">{role}</span></div>
                    <div style="font-size:12px;color:#64748b;margin-top:2px">{email}</div>
                  </td>
                  <td align="right" style="vertical-align:top;white-space:nowrap">
                    <div style="font-size:10px;color:#64748b;text-transform:uppercase;letter-spacing:.6px;font-weight:700">Time on hub</div>
                    <div style="font-size:18px;font-weight:800;color:#16a34a;line-height:1.1">{hrs_a}h <span style="color:#94a3b8;font-size:11px;font-weight:600">active</span></div>
                    <div style="font-size:11px;color:#64748b;margin-top:2px">{hrs_i}h idle · {acts} actions</div>
                    <div style="font-size:11px;color:#64748b;margin-top:2px">{session_line}</div>
                  </td>
                </tr>
              </table>
              <div style="margin-top:12px">{tab_badges}</div>
              {highlights_html}
              {client_blocks}
            </div>
            """

    inner = (
        ribbon_html +
        '<h2 style="margin:6px 0 4px;font-size:15px;color:#0f172a;font-weight:800;letter-spacing:.2px">Operator Breakdown</h2>'
        '<div style="font-size:12px;color:#64748b;margin-bottom:6px">Workers only — admins suppressed (configure with <code>EOD_REPORT_EXCLUDE_USERS</code>).</div>' +
        cards_html
    )

    html_body = _brand_email_shell(
        title="End-of-Day Team Report",
        subtitle=f"{date_long} — per operator, per client, per module",
        accent="#2563eb",
        inner_html=inner,
    )
    return text_body, html_body


def send_eod_team_report(report_date: str = None, force: bool = False) -> dict:
    """Compose and email the end-of-day team report.

    Default recipients: lexi@medprosc.com + eric@medprosc.com
    (override with EOD_REPORT_EMAIL, comma-separated).

    Each recipient is sent individually via _send_email_to so SendGrid
    delivery is per-recipient (we can see who bounced).

    Returns a delivery report dict {date, recipients, sent, failed,
    headlines, user_count}.
    """
    from datetime import datetime as _dt
    try:
        from app.client_db import get_eod_team_report
    except Exception as e:
        log.error(f"EOD report aggregator import failed: {e}")
        return {"ok": False, "error": str(e)}

    if not report_date:
        report_date = _dt.now().strftime("%Y-%m-%d")

    try:
        report = get_eod_team_report(report_date)
    except Exception as e:
        log.error(f"EOD report aggregation failed: {e}")
        return {"ok": False, "error": str(e), "date": report_date}

    headlines = report.get("headlines", {}) or {}
    if not force and not report.get("users") and sum(headlines.values()) == 0:
        log.info(f"EOD report skipped — zero activity for {report_date}")
        return {"ok": True, "skipped": "no activity", "date": report_date}

    text_body, html_body = _render_eod_report_html(report)
    try:
        from datetime import datetime as _dt2
        d_long = _dt2.strptime(report_date, "%Y-%m-%d").strftime("%a %b %d, %Y")
    except Exception:
        d_long = report_date
    subject = (
        f"📋 MedPharma EOD — {d_long} — "
        f"{headlines.get('active_users', 0)} active, "
        f"{headlines.get('claims_new', 0)} new claims, "
        f"{headlines.get('production_hours', 0)}h logged"
    )

    recipients = _eod_recipients()

    # Persist the rendered report BEFORE attempting email so it's available
    # in the in-app archive even if SendGrid is unconfigured / fails.
    archive_id = 0
    try:
        from app.client_db import save_eod_report
        archive_id = save_eod_report(
            report_date=report_date,
            headlines=headlines,
            summary=report,
            html_body=html_body,
            text_body=text_body,
            generated_by="scheduled" if not force else "manual",
            email_status="pending",
            email_recipients=recipients,
        )
    except Exception:
        log.exception("save_eod_report failed for %s", report_date)

    # Fan out an in-app notification to every active admin/staff user so
    # they see "Your EOD report is ready" the moment they sign in, no matter
    # what email is doing.
    try:
        from app.client_db import get_db, fanout_notification
        conn = get_db()
        try:
            cur = conn.cursor()
            cur.execute(
                "SELECT id FROM clients "
                "WHERE COALESCE(is_active,1)=1 "
                "AND lower(COALESCE(role,'client')) IN ('admin','staff')"
            )
            staff_ids = [int(r[0]) for r in cur.fetchall()]
        finally:
            conn.close()
        if staff_ids:
            fanout_notification(
                user_ids=staff_ids,
                kind="eod_report",
                title=f"EOD report ready · {d_long}",
                body=(
                    f"{headlines.get('active_users', 0)} active users · "
                    f"{headlines.get('claims_new', 0)} new claims · "
                    f"{headlines.get('production_hours', 0)}h logged"
                ),
                link=(f"/hub?eod={archive_id}" if archive_id else "/hub"),
                related_type="eod_report",
                related_id=archive_id or None,
            )
    except Exception:
        log.exception("EOD in-app fanout failed for %s", report_date)

    sent, failed = [], []
    if not recipients:
        log.error("EOD report has no recipients configured")
        try:
            from app.client_db import update_eod_report_email_status
            if archive_id:
                update_eod_report_email_status(archive_id, "no_recipients", [])
        except Exception:
            pass
        return {"ok": False, "error": "no recipients", "date": report_date,
                "archive_id": archive_id}

    for to_email in recipients:
        try:
            ok_sent, via = _send_email_to(to_email, subject, text_body, html_body)
        except Exception as e:
            log.error(f"EOD report send to {to_email} crashed: {e}")
            failed.append({"email": to_email, "error": str(e), "via": "exception"})
            continue
        if ok_sent:
            sent.append({"email": to_email, "via": via})
        else:
            failed.append({"email": to_email, "via": via})

    # Update archive with final delivery status so the in-app history view
    # tells the operator exactly what happened.
    try:
        from app.client_db import update_eod_report_email_status
        if archive_id:
            if sent and not failed:
                final_status = "delivered"
            elif sent and failed:
                final_status = "partial"
            elif failed:
                # Distinguish "no provider configured" from "provider rejected".
                via0 = (failed[0].get("via") or "").lower()
                if "not configured" in via0 or "no provider" in via0 or "missing" in via0:
                    final_status = "no_provider"
                else:
                    final_status = "failed"
            else:
                final_status = "unknown"
            update_eod_report_email_status(archive_id, final_status,
                                           [s.get("email") for s in sent])
    except Exception:
        log.exception("update_eod_report_email_status failed for %s", archive_id)

    log.info(
        f"EOD report dispatched for {report_date}: sent={len(sent)} failed={len(failed)} "
        f"users={len(report.get('users', []))} archive_id={archive_id}"
    )
    return {
        "ok": True,
        "date": report_date,
        "archive_id": archive_id,
        "recipients": recipients,
        "sent": sent,
        "failed": failed,
        "user_count": len(report.get("users", [])),
        "headlines": headlines,
    }


def _build_demo_eod_report() -> dict:
    """Fabricate a realistic EOD report so the email preview shows what
    the final product looks like even when the live DB is quiet. Same
    shape as get_eod_team_report() — fed directly to the renderer.
    """
    from datetime import datetime as _dt
    today = _dt.now().strftime("%Y-%m-%d")
    tab_keys = ["Claims", "Credentialing", "Enrollment", "EDI", "Production",
                "Documents", "Notes", "Chat", "Audit", "Pageviews"]

    def _tabs(**vals):
        out = {k: 0 for k in tab_keys}
        out.update(vals)
        return out

    users = [
        {
            "username": "jessica@medprosc.com",
            "contact_name": "Jessica",
            "email": "jessica@medprosc.com",
            "role": "staff",
            "active_hours": 6.4,
            "idle_hours": 1.1,
            "actions": 142,
            "first_seen": f"{today}T08:53:12",
            "last_seen": f"{today}T17:18:44",
            "totals": _tabs(Claims=18, Notes=7, Production=3, Pageviews=92, Audit=4),
            "total_actions": 124,
            "highlights": [
                "claim_status_update claims_master — moved CLM-44218 from A/R Follow-Up → Paid",
                "note_added notes_log — Aetna confirmed reprocess; eta 14 days",
            ],
            "clients": {
                "Apex Pain Management": {
                    "totals": _tabs(Claims=12, Notes=4, Pageviews=58, Audit=2),
                    "items": [
                        {"tab": "Claims",     "action": "created", "title": "CLM-44218 (Billed/Submitted)"},
                        {"tab": "Claims",     "action": "updated", "title": "CLM-44109 (Paid)"},
                        {"tab": "Claims",     "action": "updated", "title": "CLM-44091 (Appeals)"},
                        {"tab": "Notes",      "action": "noted",   "title": "Claim CLM-44091 — sent reconsideration packet to Cigna"},
                        {"tab": "Notes",      "action": "noted",   "title": "Claim CLM-44218 — Aetna confirmed receipt"},
                    ],
                },
                "Coastal Orthopedics": {
                    "totals": _tabs(Claims=6, Notes=3, Production=3, Pageviews=34, Audit=2),
                    "items": [
                        {"tab": "Claims",     "action": "updated", "title": "CLM-22107 (Denied → Appeals)"},
                        {"tab": "Production", "action": "logged",  "title": "A/R Follow-Up: Worked Cigna denials batch (12 · 2.5h)"},
                        {"tab": "Notes",      "action": "noted",   "title": "Claim CLM-22107 — Cigna requires CPT modifier docs"},
                    ],
                },
            },
        },
        {
            "username": "susan@medprosc.com",
            "contact_name": "Susan",
            "email": "susan@medprosc.com",
            "role": "staff",
            "active_hours": 5.8,
            "idle_hours": 0.7,
            "actions": 96,
            "first_seen": f"{today}T09:02:50",
            "last_seen": f"{today}T16:44:18",
            "totals": _tabs(Credentialing=9, Enrollment=4, Documents=2, Notes=5, Pageviews=61, Audit=3),
            "total_actions": 84,
            "highlights": [
                "credentialing_submit credentialing — Dr Chen / BCBS-SC initial app submitted",
                "file_upload client_files — uploaded CAQH attestation PDF",
            ],
            "clients": {
                "Apex Pain Management": {
                    "totals": _tabs(Credentialing=6, Notes=3, Documents=1, Pageviews=38, Audit=2),
                    "items": [
                        {"tab": "Credentialing", "action": "created", "title": "Dr Chen · BCBS-SC (Submitted)"},
                        {"tab": "Credentialing", "action": "updated", "title": "Dr Patel · Aetna (Approved)"},
                        {"tab": "Documents",     "action": "uploaded","title": "Patel-CAQH-attestation.pdf · Credentialing"},
                        {"tab": "Notes",         "action": "noted",   "title": "Credentialing Dr Chen — recruiter packet forwarded"},
                    ],
                },
                "Magnolia Imaging": {
                    "totals": _tabs(Credentialing=3, Enrollment=4, Documents=1, Notes=2, Pageviews=23, Audit=1),
                    "items": [
                        {"tab": "Enrollment",   "action": "created", "title": "Dr Lee · Humana (Submitted)"},
                        {"tab": "Enrollment",   "action": "updated", "title": "Dr Lee · Cigna (Approved)"},
                        {"tab": "Credentialing","action": "updated", "title": "Dr Lee · BCBS-SC (Follow-up scheduled)"},
                    ],
                },
            },
        },
        {
            "username": "rcm@medprosc.com",
            "contact_name": "RCM",
            "email": "rcm@medprosc.com",
            "role": "admin",
            "active_hours": 4.2,
            "idle_hours": 0.6,
            "actions": 71,
            "first_seen": f"{today}T10:11:09",
            "last_seen": f"{today}T15:33:51",
            "totals": _tabs(Production=5, EDI=3, Notes=2, Chat=4, Pageviews=40, Audit=2),
            "total_actions": 56,
            "highlights": [
                "production_log team_production — 4.5h posted (ERA reconciliation)",
                "edi_status edi_setup — Apex GoLive confirmed for 06/12",
            ],
            "clients": {
                "Apex Pain Management": {
                    "totals": _tabs(Production=3, EDI=2, Notes=1, Chat=2, Pageviews=24, Audit=1),
                    "items": [
                        {"tab": "Production", "action": "logged", "title": "ERA Reconciliation: Tied out Aetna 06/03 ERA (28 · 2.0h)"},
                        {"tab": "EDI",        "action": "updated","title": "Apex · UHC (ERA Active, EFT Active)"},
                        {"tab": "Chat",       "action": "messaged","title": "room 'Apex — RCM Daily Standup'"},
                    ],
                },
                "Coastal Orthopedics": {
                    "totals": _tabs(Production=2, EDI=1, Notes=1, Chat=2, Pageviews=16, Audit=1),
                    "items": [
                        {"tab": "Production", "action": "logged",  "title": "Posting: BCBS-SC EOB batch (14 · 1.0h)"},
                        {"tab": "EDI",        "action": "updated", "title": "Coastal · Aetna (ERA Pending → Active)"},
                        {"tab": "Chat",       "action": "messaged","title": "room 'Coastal — Posting Queue'"},
                    ],
                },
            },
        },
        {
            "username": "eric@medprosc.com",
            "contact_name": "Eric",
            "email": "eric@medprosc.com",
            "role": "admin",
            "active_hours": 2.3,
            "idle_hours": 0.4,
            "actions": 38,
            "first_seen": f"{today}T11:24:01",
            "last_seen": f"{today}T15:08:22",
            "totals": _tabs(Chat=11, Pageviews=27, Audit=2),
            "total_actions": 40,
            "highlights": [
                "client_review clients — opened Apex profile; reviewed open AR queue",
                "chat_room_create chat_rooms — created 'Apex — RCM Daily Standup'",
            ],
            "clients": {
                "Apex Pain Management": {
                    "totals": _tabs(Chat=7, Pageviews=18, Audit=1),
                    "items": [
                        {"tab": "Chat", "action": "messaged", "title": "room 'Apex — RCM Daily Standup'"},
                        {"tab": "Chat", "action": "messaged", "title": "room 'Apex — Credentialing Sync'"},
                    ],
                },
                "Coastal Orthopedics": {
                    "totals": _tabs(Chat=4, Pageviews=9, Audit=1),
                    "items": [
                        {"tab": "Chat", "action": "messaged", "title": "room 'Coastal — Posting Queue'"},
                    ],
                },
            },
        },
        {
            "username": "admin@medprosc.com",
            "contact_name": "Lexi",
            "email": "lexi@medprosc.com",
            "role": "admin",
            "active_hours": 1.7,
            "idle_hours": 0.2,
            "actions": 24,
            "first_seen": f"{today}T13:02:11",
            "last_seen": f"{today}T15:42:00",
            "totals": _tabs(Audit=6, Pageviews=18),
            "total_actions": 24,
            "highlights": [
                "client_create clients — Magnolia Imaging onboarded",
                "user_invite clients — invited admin@magnoliaimg.com (staff)",
                "module_toggle clients — enabled Documents + Chat for Coastal",
            ],
            "clients": {
                "Magnolia Imaging": {
                    "totals": _tabs(Audit=4, Pageviews=11),
                    "items": [
                        {"tab": "Audit", "action": "created", "title": "client_create — Magnolia Imaging onboarded"},
                        {"tab": "Audit", "action": "updated", "title": "user_invite — admin@magnoliaimg.com"},
                    ],
                },
                "Coastal Orthopedics": {
                    "totals": _tabs(Audit=2, Pageviews=7),
                    "items": [
                        {"tab": "Audit", "action": "updated", "title": "module_toggle — enabled Documents + Chat"},
                    ],
                },
            },
        },
    ]

    team_totals = {k: 0 for k in tab_keys}
    for u in users:
        for k, v in u["totals"].items():
            team_totals[k] = team_totals.get(k, 0) + v

    headlines = {
        "claims_new": 18,
        "claims_touched": 23,
        "cred_new": 9,
        "enroll_new": 4,
        "edi_new": 3,
        "production_rows": 5,
        "production_hours": 8.5,
        "notes_new": 14,
        "files_uploaded": 3,
        "chat_messages": 15,
        "audit_events": 17,
        "active_users": len(users),
    }

    return {
        "report_date": today,
        "generated_at": _dt.now().isoformat(timespec="seconds"),
        "tab_keys": tab_keys,
        "users": users,
        "team_totals": team_totals,
        "headlines": headlines,
        "client_count": 3,
    }


def send_eod_team_report_demo() -> dict:
    """Render & email a fully-populated DEMO end-of-day report.

    Same recipients as the real one (lexi@medprosc.com + eric@medprosc.com
    by default, overridable via EOD_REPORT_EMAIL). Use this when the live
    DB is quiet but you want to verify the email actually looks right.
    """
    report = _build_demo_eod_report()
    text_body, html_body = _render_eod_report_html(report)
    headlines = report["headlines"]
    try:
        from datetime import datetime as _dt
        d_long = _dt.strptime(report["report_date"], "%Y-%m-%d").strftime("%a %b %d, %Y")
    except Exception:
        d_long = report["report_date"]
    subject = (
        f"📋 MedPharma EOD [DEMO] — {d_long} — "
        f"{headlines['active_users']} active, "
        f"{headlines['claims_new']} new claims, "
        f"{headlines['production_hours']}h logged"
    )

    recipients = _eod_recipients()
    sent, failed = [], []
    for to_email in recipients:
        try:
            ok_sent, via = _send_email_to(to_email, subject, text_body, html_body)
        except Exception as e:
            log.error(f"EOD demo send to {to_email} crashed: {e}")
            failed.append({"email": to_email, "via": f"exception: {e}"})
            continue
        if ok_sent:
            sent.append({"email": to_email, "via": via})
        else:
            failed.append({"email": to_email, "via": via})

    log.info(f"EOD DEMO dispatched: sent={len(sent)} failed={len(failed)}")
    return {
        "ok": True,
        "demo": True,
        "date": report["report_date"],
        "recipients": recipients,
        "sent": sent,
        "failed": failed,
        "user_count": len(report["users"]),
        "headlines": headlines,
    }


# ─── Client-facing daily production report ──────────────────────────────

# Friendly section headers + ordering for the client email + Excel sheet.
_CLIENT_SECTION_META = [
    ("claims",        "Claims Activity",      "claims",        "💼", "#2563eb"),
    ("credentialing", "Credentialing",        "credentialing", "🎓", "#7c3aed"),
    ("enrollment",    "Payor Enrollment",     "enrollment",    "📝", "#7c3aed"),
    ("edi",           "EDI / ERA / EFT",      "edi",           "🔌", "#0891b2"),
    ("production",    "Production Hours",     "production",    "⏱️", "#16a34a"),
    ("notes",         "Notes",                "claims",        "🗒️", "#0f766e"),
    ("documents",     "Documents Uploaded",   "documents",     "📁", "#0f766e"),
]


def _client_section_visible(section_key: str, enabled_modules: list) -> bool:
    """Skip a section when the client hasn't enabled its module."""
    for k, _label, mod, _ico, _c in _CLIENT_SECTION_META:
        if k == section_key:
            if not enabled_modules:
                return True
            return mod in {m.lower() for m in enabled_modules}
    return True


def _ts_long(value) -> str:
    """Format 'HH:MM' (24h) — same logic as _ts_short but exposed for symmetry."""
    return _ts_short(value)


def _render_client_daily_report_html(report: dict) -> tuple[str, str]:
    """Render the per-client daily production report into (text_body, html_body).

    Layout: stacked vertical sections, organised by production area
    (claims movement → credentialing → enrollment → EDI → production
    hours → notes → documents). Skips sections the client hasn't
    enabled. Every row is timestamped.
    """
    company  = report.get("company") or ""
    date_iso = report.get("report_date") or ""
    contact  = report.get("contact_name") or ""
    try:
        from datetime import datetime as _dt
        d = _dt.strptime(date_iso, "%Y-%m-%d")
        date_long = d.strftime("%A, %B %d, %Y")
    except Exception:
        date_long = date_iso

    headlines = report.get("headlines", {}) or {}
    sections  = report.get("sections", {}) or {}
    operators = report.get("operators", []) or []
    enabled   = report.get("enabled_modules") or []

    # ── plain-text fallback ──
    txt = [
        f"MedPharma Daily Production Report — {company}",
        f"{date_long}",
        "",
        f"Claims: {headlines.get('claims_new',0)} new · {headlines.get('claims_touched',0)} touched · "
        f"{headlines.get('claims_paid',0)} paid · {headlines.get('claims_denied',0)} denied/appealed",
        f"Credentialing new: {headlines.get('cred_new',0)} · Enrollment new: {headlines.get('enroll_new',0)} · "
        f"EDI new: {headlines.get('edi_new',0)}",
        f"Production hours: {headlines.get('production_hours',0)} · Notes: {headlines.get('notes_new',0)} · "
        f"Files: {headlines.get('files_uploaded',0)}",
        f"Operators on your account: {headlines.get('operators',0)}",
        "",
    ]
    for key, label, _mod, ico, _color in _CLIENT_SECTION_META:
        if not _client_section_visible(key, enabled):
            continue
        rows = sections.get(key) or []
        if not rows:
            continue
        txt.append(f"== {label} ({len(rows)}) ==")
        for r in rows[:50]:
            ts = _ts_short(r.get("ts") or "")
            ts_prefix = f"[{ts}] " if ts else ""
            if key == "claims":
                txt.append(f"  {ts_prefix}{r.get('ClaimKey','')} — {r.get('ClaimStatus','')} ({r.get('action','')}) by {r.get('Owner','')}")
            elif key in ("credentialing", "enrollment", "edi"):
                txt.append(f"  {ts_prefix}{r.get('ProviderName','')} · {r.get('Payor','')} — {r.get('Status','')} ({r.get('action','')}) by {r.get('Owner','')}")
            elif key == "production":
                txt.append(f"  {ts_prefix}{r.get('Owner','')}: {r.get('Category','')} — {r.get('Task','')} ({r.get('Qty',0)} · {r.get('Hours',0)}h)")
            elif key == "notes":
                txt.append(f"  {ts_prefix}{r.get('Author','')}: {r.get('Subject','')} — {r.get('Note','')}")
            elif key == "documents":
                txt.append(f"  {ts_prefix}{r.get('Filename','')} ({r.get('Category','')}) by {r.get('UploadedBy','')}")
        txt.append("")

    text_body = "\n".join(txt)

    # ── headline ribbon ──
    def _ribbon(label, value, color="#1d4ed8"):
        return (
            f'<td align="center" valign="top" width="16.66%" style="padding:14px 6px">'
            f'<div style="font-size:10px;color:#64748b;text-transform:uppercase;letter-spacing:.8px;font-weight:700">{label}</div>'
            f'<div style="font-size:22px;color:{color};font-weight:800;margin-top:4px;line-height:1.1">{value}</div>'
            f'</td>'
        )
    ribbon_html = (
        '<table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" '
        'style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;margin:0 0 18px">'
        '<tr>'
        + _ribbon("Claims new",     headlines.get("claims_new", 0))
        + _ribbon("Claims touched", headlines.get("claims_touched", 0), "#2563eb")
        + _ribbon("Paid",           headlines.get("claims_paid", 0), "#16a34a")
        + _ribbon("Denied / Appeals", headlines.get("claims_denied", 0), "#dc2626")
        + _ribbon("Cred / Enroll",  f"{headlines.get('cred_new',0)} / {headlines.get('enroll_new',0)}", "#7c3aed")
        + _ribbon("Prod hrs",       f"{headlines.get('production_hours', 0)}", "#16a34a")
        + '</tr></table>'
    )

    # ── per-section tables ──
    def _row(cells, td_style="padding:6px 10px;border-bottom:1px solid #f1f5f9;font-size:12px;color:#334155"):
        return "<tr>" + "".join(f'<td style="{td_style}">{c}</td>' for c in cells) + "</tr>"

    def _section_table(key, label, ico, color):
        rows = sections.get(key) or []
        if not _client_section_visible(key, enabled) or not rows:
            return ""
        # Header row varies per section
        if key == "claims":
            header = ["Time", "Claim", "Status", "Action", "Owner"]
            body_rows = "".join(_row([
                _ts_short(r.get("ts","")),
                _esc_html(r.get("ClaimKey","")),
                _esc_html(r.get("ClaimStatus","")),
                _esc_html(r.get("action","")),
                _esc_html(r.get("Owner","")),
            ]) for r in rows[:30])
        elif key in ("credentialing", "enrollment", "edi"):
            header = ["Time", "Provider", "Payor", "Status", "Action", "Owner"]
            body_rows = "".join(_row([
                _ts_short(r.get("ts","")),
                _esc_html(r.get("ProviderName","")),
                _esc_html(r.get("Payor","")),
                _esc_html(r.get("Status","")),
                _esc_html(r.get("action","")),
                _esc_html(r.get("Owner","")),
            ]) for r in rows[:30])
        elif key == "production":
            header = ["Time", "Owner", "Category", "Task", "Qty", "Hours"]
            body_rows = "".join(_row([
                _ts_short(r.get("ts","")),
                _esc_html(r.get("Owner","")),
                _esc_html(r.get("Category","")),
                _esc_html(r.get("Task","")),
                _esc_html(r.get("Qty","")),
                _esc_html(r.get("Hours","")),
            ]) for r in rows[:30])
        elif key == "notes":
            header = ["Time", "Author", "Subject", "Note"]
            body_rows = "".join(_row([
                _ts_short(r.get("ts","")),
                _esc_html(r.get("Author","")),
                _esc_html(r.get("Subject","")),
                _esc_html(r.get("Note","")),
            ]) for r in rows[:30])
        elif key == "documents":
            header = ["Time", "File", "Category", "Uploaded by"]
            body_rows = "".join(_row([
                _ts_short(r.get("ts","")),
                _esc_html(r.get("Filename","")),
                _esc_html(r.get("Category","")),
                _esc_html(r.get("UploadedBy","")),
            ]) for r in rows[:30])
        else:
            return ""
        header_html = "<tr>" + "".join(
            f'<th align="left" style="padding:8px 10px;background:#f8fafc;border-bottom:2px solid {color};'
            f'font-size:10px;text-transform:uppercase;letter-spacing:.6px;color:#475569;font-weight:700">{h}</th>'
            for h in header
        ) + "</tr>"
        more_note = ""
        if len(rows) > 30:
            more_note = (
                f'<div style="font-size:11px;color:#64748b;padding:6px 12px;font-style:italic">'
                f'…and {len(rows) - 30} more rows — see the attached Excel for the full list.</div>'
            )
        return f"""
        <div style="margin:18px 0;background:#fff;border:1px solid #e2e8f0;border-radius:12px;overflow:hidden">
          <div style="padding:10px 14px;background:linear-gradient(90deg,{color}10,#ffffff);border-bottom:1px solid #e2e8f0">
            <span style="font-size:14px;font-weight:800;color:{color}">{ico} {label}</span>
            <span style="font-size:11px;color:#64748b;font-weight:600;margin-left:6px">({len(rows)} {'row' if len(rows)==1 else 'rows'})</span>
          </div>
          <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0">
            <thead>{header_html}</thead>
            <tbody>{body_rows}</tbody>
          </table>
          {more_note}
        </div>"""

    sections_html = ""
    for key, label, _mod, ico, color in _CLIENT_SECTION_META:
        sections_html += _section_table(key, label, ico, color)

    if not sections_html:
        sections_html = (
            '<div style="padding:30px;background:#f8fafc;border:1px dashed #cbd5e1;border-radius:10px;'
            'text-align:center;color:#64748b;font-size:13px">'
            'No production activity recorded today for your account. '
            'Tomorrow\'s report will resume automatically.</div>'
        )

    # Operator footer
    ops_html = ""
    if operators:
        rows_html = ""
        for op in operators:
            rows_html += (
                f'<tr><td style="padding:6px 10px;font-size:12px;color:#334155">'
                f'<b>{_esc_html(op.get("contact_name") or op.get("username"))}</b> '
                f'<span style="color:#94a3b8">({_esc_html(op.get("role") or "staff")})</span>'
                f'</td><td align="right" style="padding:6px 10px;font-size:12px;color:#475569">'
                f'{op.get("actions", 0)} actions · {op.get("hours", 0)}h</td></tr>'
            )
        ops_html = f"""
        <div style="margin:18px 0 0;padding:14px 16px;background:#0b1733;border-radius:12px">
          <div style="font-size:10px;color:#94a3b8;text-transform:uppercase;letter-spacing:.8px;font-weight:700;margin-bottom:8px">
            MedPharma operators on your account today
          </div>
          <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#ffffff;border-radius:8px">
            {rows_html}
          </table>
        </div>"""

    intro = (
        f'<div style="font-size:13px;color:#334155;margin-bottom:16px">'
        f'Hi {_esc_html(contact.split()[0]) if contact else "team"}, here\'s a summary of what '
        f'<b>MedPharma</b> handled on your account today. A spreadsheet with the full row-level '
        f'detail is attached. Reply to this email to reach your account team.'
        f'</div>'
    )

    inner = intro + ribbon_html + sections_html + ops_html

    html_body = _brand_email_shell(
        title=f"Daily Production Report — {company}",
        subtitle=date_long,
        accent="#16a34a",
        inner_html=inner,
        footer_note=(
            f"Sent to {_esc_html(report.get('email',''))} · "
            f'<a href="{_MEDPHARMA_HUB_URL}" style="color:#94a3b8">Open Hub</a> · '
            f'<a href="{_MEDPHARMA_SITE_URL}" style="color:#94a3b8">medpharmasc.com</a><br>'
            f'To stop receiving this email, contact your MedPharma account manager.'
        ),
    )
    return text_body, html_body


def _build_client_report_xlsx(report: dict) -> bytes:
    """Build an .xlsx workbook with one sheet per non-empty production
    section. Returns the raw bytes so the email layer can attach it.

    Falls back gracefully (returns empty bytes) when openpyxl isn't
    installed — the email still ships, just without the attachment.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
    except Exception as e:
        log.warning(f"openpyxl unavailable, skipping XLSX attachment: {e}")
        return b""

    wb = Workbook()
    # Replace the default sheet with our Summary
    summary = wb.active
    summary.title = "Summary"
    summary.append(["MedPharma Daily Production Report"])
    summary.append([f"Client: {report.get('company','')}"])
    summary.append([f"Date:   {report.get('report_date','')}"])
    summary.append([f"Generated: {report.get('generated_at','')}"])
    summary.append([])
    summary.append(["Headline", "Value"])
    h = report.get("headlines", {}) or {}
    summary.append(["Claims new",     h.get("claims_new", 0)])
    summary.append(["Claims touched", h.get("claims_touched", 0)])
    summary.append(["Claims paid",    h.get("claims_paid", 0)])
    summary.append(["Claims denied/appealed", h.get("claims_denied", 0)])
    summary.append(["Credentialing new", h.get("cred_new", 0)])
    summary.append(["Enrollment new",    h.get("enroll_new", 0)])
    summary.append(["EDI new",           h.get("edi_new", 0)])
    summary.append(["Production hours",  h.get("production_hours", 0)])
    summary.append(["Notes",             h.get("notes_new", 0)])
    summary.append(["Files uploaded",    h.get("files_uploaded", 0)])
    summary.append(["Operators",         h.get("operators", 0)])
    # Bold the first row
    for cell in summary[1]:
        cell.font = Font(bold=True, size=14, color="2563EB")
    for cell in summary[6]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F1F5F9")
    summary.column_dimensions["A"].width = 32
    summary.column_dimensions["B"].width = 20

    # Section sheets
    section_headers = {
        "claims":        ["Time", "ClaimKey", "ClaimStatus", "Action", "Owner"],
        "credentialing": ["Time", "Provider", "Payor", "Status", "Action", "Owner"],
        "enrollment":    ["Time", "Provider", "Payor", "Status", "Action", "Owner"],
        "edi":           ["Time", "Provider", "Payor", "Status", "Action", "Owner"],
        "production":    ["Time", "Owner", "Category", "Task", "Qty", "Hours"],
        "notes":         ["Time", "Author", "Subject", "Note"],
        "documents":     ["Time", "Filename", "Category", "Uploaded By"],
    }
    section_row_fn = {
        "claims":        lambda r: [_ts_short(r.get("ts","")), r.get("ClaimKey",""),
                                    r.get("ClaimStatus",""), r.get("action",""), r.get("Owner","")],
        "credentialing": lambda r: [_ts_short(r.get("ts","")), r.get("ProviderName",""),
                                    r.get("Payor",""), r.get("Status",""),
                                    r.get("action",""), r.get("Owner","")],
        "enrollment":    lambda r: [_ts_short(r.get("ts","")), r.get("ProviderName",""),
                                    r.get("Payor",""), r.get("Status",""),
                                    r.get("action",""), r.get("Owner","")],
        "edi":           lambda r: [_ts_short(r.get("ts","")), r.get("ProviderName",""),
                                    r.get("Payor",""), r.get("Status",""),
                                    r.get("action",""), r.get("Owner","")],
        "production":    lambda r: [_ts_short(r.get("ts","")), r.get("Owner",""),
                                    r.get("Category",""), r.get("Task",""),
                                    r.get("Qty", 0), r.get("Hours", 0)],
        "notes":         lambda r: [_ts_short(r.get("ts","")), r.get("Author",""),
                                    r.get("Subject",""), r.get("Note","")],
        "documents":     lambda r: [_ts_short(r.get("ts","")), r.get("Filename",""),
                                    r.get("Category",""), r.get("UploadedBy","")],
    }
    sections = report.get("sections", {}) or {}
    enabled  = report.get("enabled_modules") or []
    for key, label, _mod, _ico, _color in _CLIENT_SECTION_META:
        if not _client_section_visible(key, enabled):
            continue
        rows = sections.get(key) or []
        if not rows:
            continue
        sheet_title = label[:31]   # Excel sheet name limit
        ws = wb.create_sheet(sheet_title)
        ws.append(section_headers[key])
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2563EB")
            cell.alignment = Alignment(horizontal="left", vertical="center")
        for r in rows:
            ws.append(section_row_fn[key](r))
        # Auto-fit-ish column widths
        for col_idx in range(1, len(section_headers[key]) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 22

    # Operators sheet
    operators = report.get("operators", []) or []
    if operators:
        ws = wb.create_sheet("Operators")
        ws.append(["Username", "Contact name", "Role", "Actions", "Hours"])
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0B1733")
        for op in operators:
            ws.append([
                op.get("username",""),
                op.get("contact_name",""),
                op.get("role",""),
                op.get("actions", 0),
                op.get("hours", 0),
            ])
        for col in "ABCDE":
            ws.column_dimensions[col].width = 20

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _client_report_recipients(report: dict) -> list[str]:
    """Compose the recipient list for a client report: primary email +
    any extras the admin set on the profile. Lowercased + deduped.
    """
    out = []
    seen = set()
    for raw in [report.get("email")] + list(report.get("report_recipients") or []):
        if not raw:
            continue
        addr = str(raw).strip().lower()
        if addr and addr not in seen and "@" in addr:
            seen.add(addr)
            out.append(addr)
    return out


def send_client_daily_report(client_id: int, report_date: str = None,
                             force: bool = False, demo: bool = False) -> dict:
    """Compose + email the per-client daily production report (with Excel).

    Args:
        client_id:   PK of the client.
        report_date: YYYY-MM-DD (defaults to today).
        force:       send even when there's zero activity.
        demo:        ignore the DB and use a populated showcase payload.

    Returns delivery report dict (sent / failed / recipients / headlines).
    """
    from datetime import datetime as _dt
    if demo:
        report = _build_demo_client_daily_report()
    else:
        try:
            from app.client_db import get_client_daily_report
        except Exception as e:
            log.error(f"client_db.get_client_daily_report import failed: {e}")
            return {"ok": False, "error": str(e), "client_id": client_id}
        if not report_date:
            report_date = _dt.now().strftime("%Y-%m-%d")
        report = get_client_daily_report(client_id, report_date)
        if not report or not report.get("ok"):
            return {"ok": False, "error": (report or {}).get("error", "no report"),
                    "client_id": client_id}

    headlines = report.get("headlines", {}) or {}
    has_activity = any(v for v in headlines.values() if isinstance(v, (int, float)) and v)
    if not force and not demo and not has_activity:
        log.info(f"Client {client_id} report skipped — no activity for {report.get('report_date')}")
        return {"ok": True, "skipped": "no activity", "client_id": client_id,
                "date": report.get("report_date")}

    text_body, html_body = _render_client_daily_report_html(report)
    xlsx_bytes = _build_client_report_xlsx(report)

    try:
        d_long = _dt.strptime(report["report_date"], "%Y-%m-%d").strftime("%a %b %d, %Y")
    except Exception:
        d_long = report.get("report_date", "")

    company = report.get("company", "")
    demo_tag = " [DEMO]" if demo else ""
    subject = (
        f"📊 MedPharma Daily Report{demo_tag} — {company} — {d_long} — "
        f"{headlines.get('claims_new', 0)} new claims · "
        f"{headlines.get('production_hours', 0)}h logged"
    )

    attachments = []
    if xlsx_bytes:
        safe_co = "".join(c if c.isalnum() else "_" for c in (company or "client"))[:40]
        attachments.append({
            "filename": f"MedPharma_{safe_co}_{report.get('report_date','')}.xlsx",
            "content":  xlsx_bytes,
            "mime":     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        })

    recipients = _client_report_recipients(report)
    if not recipients:
        log.warning(f"Client {client_id} report has no deliverable recipients")
        return {"ok": False, "error": "no recipients", "client_id": client_id,
                "date": report.get("report_date")}

    sent, failed = [], []
    for to_email in recipients:
        try:
            ok_sent, via = _send_email_to(to_email, subject, text_body, html_body,
                                          attachments=attachments)
        except Exception as e:
            log.error(f"Client report send to {to_email} crashed: {e}")
            failed.append({"email": to_email, "via": f"exception: {e}"})
            continue
        if ok_sent:
            sent.append({"email": to_email, "via": via})
        else:
            failed.append({"email": to_email, "via": via})

    log.info(
        f"Client {client_id} ({company}) report dispatched: "
        f"sent={len(sent)} failed={len(failed)} attachment={'yes' if xlsx_bytes else 'no'}"
    )
    return {
        "ok": True,
        "demo": demo,
        "client_id": client_id,
        "company": company,
        "date": report.get("report_date"),
        "recipients": recipients,
        "sent": sent,
        "failed": failed,
        "attachment_bytes": len(xlsx_bytes) if xlsx_bytes else 0,
        "headlines": headlines,
    }


def send_all_client_daily_reports(report_date: str = None, force: bool = False) -> dict:
    """Iterate every opted-in client (has email + daily_report_optin=1)
    and dispatch their per-client production report. Used by the 6:30 PM
    EST scheduler so the admin doesn't have to push a button per client.
    """
    try:
        from app.client_db import list_clients_optin_for_daily_report
    except Exception as e:
        log.error(f"list_clients_optin_for_daily_report import failed: {e}")
        return {"ok": False, "error": str(e)}

    clients = list_clients_optin_for_daily_report()
    results = []
    sent_total = failed_total = skipped = 0
    for c in clients:
        try:
            result = send_client_daily_report(c["client_id"],
                                              report_date=report_date,
                                              force=force, demo=False)
        except Exception as e:
            log.error(f"send_client_daily_report({c['client_id']}) crashed: {e}")
            result = {"ok": False, "error": str(e), "client_id": c["client_id"]}
        if result.get("skipped"):
            skipped += 1
        else:
            sent_total += len(result.get("sent") or [])
            failed_total += len(result.get("failed") or [])
        results.append({
            "client_id": c["client_id"],
            "company":   c["company"],
            "ok":        result.get("ok"),
            "skipped":   result.get("skipped"),
            "sent":      len(result.get("sent") or []),
            "failed":    len(result.get("failed") or []),
        })
    log.info(
        f"Client report fan-out complete: "
        f"clients={len(clients)} sent={sent_total} failed={failed_total} skipped={skipped}"
    )
    return {
        "ok": True,
        "date": report_date,
        "client_count":  len(clients),
        "sent_total":    sent_total,
        "failed_total":  failed_total,
        "skipped":       skipped,
        "results":       results,
    }


def _build_demo_client_daily_report() -> dict:
    """Fabricated single-client production report for previewing the email
    layout when the live DB has no activity for the target client.
    """
    from datetime import datetime as _dt
    today = _dt.now().strftime("%Y-%m-%d")
    ts = lambda hh, mm: f"{today} {hh:02d}:{mm:02d}:00"
    return {
        "ok": True,
        "client_id": 0,
        "company":   "Apex Pain Management (DEMO)",
        "contact_name": "Dr. Elena Vargas",
        "email":     "lexi@medprosc.com",
        "report_date": today,
        "generated_at": _dt.now().isoformat(timespec="seconds"),
        "enabled_modules": ["claims", "credentialing", "enrollment",
                            "production", "documents", "chat"],
        "report_recipients": [],
        "daily_report_optin": 1,
        "headlines": {
            "claims_new":       8,
            "claims_touched":   12,
            "claims_paid":      4,
            "claims_denied":    2,
            "cred_new":         3,
            "enroll_new":       1,
            "edi_new":          0,
            "production_hours": 5.5,
            "notes_new":        7,
            "files_uploaded":   2,
            "operators":        3,
        },
        "sections": {
            "claims": [
                {"ts": ts(9, 4),  "ClaimKey": "CLM-44218", "ClaimStatus": "Billed/Submitted", "Owner": "jessica", "action": "created"},
                {"ts": ts(9, 38), "ClaimKey": "CLM-44219", "ClaimStatus": "Billed/Submitted", "Owner": "jessica", "action": "created"},
                {"ts": ts(10, 22),"ClaimKey": "CLM-44091", "ClaimStatus": "Appeals",          "Owner": "jessica", "action": "updated"},
                {"ts": ts(11, 14),"ClaimKey": "CLM-44109", "ClaimStatus": "Paid",             "Owner": "rcm",     "action": "updated"},
                {"ts": ts(13, 47),"ClaimKey": "CLM-44177", "ClaimStatus": "Paid",             "Owner": "rcm",     "action": "updated"},
                {"ts": ts(14, 9), "ClaimKey": "CLM-44188", "ClaimStatus": "Denied",           "Owner": "jessica", "action": "updated"},
            ],
            "credentialing": [
                {"ts": ts(10, 1), "ProviderName": "Dr Chen",  "Payor": "BCBS-SC", "Status": "Submitted", "Owner": "susan", "action": "created"},
                {"ts": ts(11, 56),"ProviderName": "Dr Patel", "Payor": "Aetna",   "Status": "Approved",  "Owner": "susan", "action": "updated"},
                {"ts": ts(15, 12),"ProviderName": "Dr Chen",  "Payor": "Cigna",   "Status": "Submitted", "Owner": "susan", "action": "created"},
            ],
            "enrollment": [
                {"ts": ts(13, 5), "ProviderName": "Dr Patel", "Payor": "Humana", "Status": "Submitted", "Owner": "susan", "action": "created"},
            ],
            "edi": [],
            "production": [
                {"ts": ts(11, 0), "Owner": "rcm",    "Category": "ERA Reconciliation", "Task": "Tied out Aetna 06/03 ERA", "Qty": 28, "Hours": 2.0},
                {"ts": ts(13, 30),"Owner": "rcm",    "Category": "Posting",            "Task": "BCBS-SC EOB batch",        "Qty": 14, "Hours": 1.0},
                {"ts": ts(15, 0), "Owner": "jessica","Category": "A/R Follow-Up",      "Task": "Worked Cigna denials batch","Qty": 12, "Hours": 2.5},
            ],
            "notes": [
                {"ts": ts(10, 30),"Author": "jessica","Subject": "Claim CLM-44091","Note": "Sent reconsideration packet to Cigna with corrected modifier"},
                {"ts": ts(11, 20),"Author": "rcm",    "Subject": "Claim CLM-44109","Note": "Aetna posted 06/03 ERA — $1,247.18 to patient acct"},
                {"ts": ts(13, 45),"Author": "susan",  "Subject": "Cred Dr Chen",   "Note": "BCBS-SC initial app submitted; expect 60-day window"},
                {"ts": ts(14, 50),"Author": "jessica","Subject": "Claim CLM-44188","Note": "Denied 197 — auth not on file; pulling pre-auth from EMR"},
                {"ts": ts(15, 20),"Author": "susan",  "Subject": "Cred Dr Chen",   "Note": "Cigna initial app submitted; CAQH attestation confirmed"},
            ],
            "documents": [
                {"ts": ts(11, 5), "Filename": "Patel-CAQH-attestation.pdf",   "Category": "Credentialing", "UploadedBy": "susan"},
                {"ts": ts(15, 18),"Filename": "Chen-Cigna-W9-2026.pdf",        "Category": "Credentialing", "UploadedBy": "susan"},
            ],
        },
        "operators": [
            {"username": "jessica","contact_name": "Jessica","role": "staff","actions": 9, "hours": 2.5},
            {"username": "rcm",    "contact_name": "RCM",    "role": "admin","actions": 4, "hours": 3.0},
            {"username": "susan",  "contact_name": "Susan",  "role": "staff","actions": 6, "hours": 0.0},
        ],
    }


def send_client_daily_report_demo(to_email: str = None) -> dict:
    """Send the showcase client report (with Excel attachment).

    If ``to_email`` is given it overrides the default lexi recipient
    (handy when an admin wants to preview the layout in their own
    inbox without actually emailing a customer).
    """
    report = _build_demo_client_daily_report()
    if to_email:
        report["email"] = to_email
        report["report_recipients"] = []
    return _send_demo_client_report(report)


def _send_demo_client_report(report: dict) -> dict:
    """Internal: same dispatch path as send_client_daily_report but
    bypasses the DB lookup since the report dict is fabricated."""
    from datetime import datetime as _dt
    text_body, html_body = _render_client_daily_report_html(report)
    xlsx_bytes = _build_client_report_xlsx(report)
    try:
        d_long = _dt.strptime(report["report_date"], "%Y-%m-%d").strftime("%a %b %d, %Y")
    except Exception:
        d_long = report.get("report_date", "")
    headlines = report.get("headlines", {}) or {}
    subject = (
        f"📊 MedPharma Daily Report [DEMO] — {report.get('company','')} — {d_long} — "
        f"{headlines.get('claims_new', 0)} new claims · "
        f"{headlines.get('production_hours', 0)}h logged"
    )
    attachments = []
    if xlsx_bytes:
        safe_co = "".join(c if c.isalnum() else "_" for c in (report.get("company") or "client"))[:40]
        attachments.append({
            "filename": f"MedPharma_{safe_co}_{report.get('report_date','')}.xlsx",
            "content":  xlsx_bytes,
            "mime":     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        })
    recipients = _client_report_recipients(report)
    sent, failed = [], []
    for to_email in recipients:
        try:
            ok_sent, via = _send_email_to(to_email, subject, text_body, html_body,
                                          attachments=attachments)
        except Exception as e:
            log.error(f"Demo client report send to {to_email} crashed: {e}")
            failed.append({"email": to_email, "via": f"exception: {e}"})
            continue
        if ok_sent:
            sent.append({"email": to_email, "via": via})
        else:
            failed.append({"email": to_email, "via": via})
    log.info(f"Client report DEMO dispatched: sent={len(sent)} failed={len(failed)}")
    return {
        "ok": True,
        "demo": True,
        "client_id": 0,
        "company": report.get("company"),
        "date": report.get("report_date"),
        "recipients": recipients,
        "sent": sent,
        "failed": failed,
        "attachment_bytes": len(xlsx_bytes) if xlsx_bytes else 0,
        "headlines": headlines,
    }


def send_bizdev_followup_reminders() -> dict:
    """Remind BizDev (and admins) about leads that haven't been followed up
    in 2+ days. Fires in-app notifications always, and emails when an email
    provider is configured. Runs on a 2-day cadence per lead so a single
    overdue lead is not re-sent every day."""
    from app.client_db import claim_leads_for_reminder, list_chat_eligible_users, fanout_notification

    due = claim_leads_for_reminder()
    if not due:
        return {"ok": True, "due": 0, "emailed": 0, "notified": 0}

    recipients = [u for u in list_chat_eligible_users()
                  if (u.get("role") or "") in ("bizdev", "admin")]
    if not recipients:
        return {"ok": True, "due": len(due), "emailed": 0, "notified": 0,
                "note": "no bizdev/admin recipients"}

    # In-app notification for every recipient — works with no email provider.
    notified = 0
    try:
        ids = [int(u["id"]) for u in recipients if u.get("id")]
        title = f"⏰ {len(due)} lead{'s' if len(due) != 1 else ''} need follow-up"
        body = "; ".join(
            f"{d.get('practice_name') or d.get('contact_name') or ('Lead #' + str(d.get('id')))}"
            f" ({d.get('days_since_contact')}d)"
            for d in due[:10]
        )
        notified = fanout_notification(ids, "lead_followup", title, body,
                                       link="/hub?panel=leads",
                                       related_type="lead")
    except Exception as e:
        log.error(f"bizdev follow-up in-app notify failed: {e}")

    # Build the email body.
    def _lead_line(d):
        name = d.get("practice_name") or d.get("contact_name") or f"Lead #{d.get('id')}"
        who = d.get("contact_name") or ""
        phone = d.get("contact_phone") or ""
        email = d.get("contact_email") or ""
        days = d.get("days_since_contact")
        bits = [f"{name} — {days} days since last contact"]
        meta = ", ".join(x for x in (who, phone, email) if x)
        if meta:
            bits.append(f"   {meta}")
        if d.get("status"):
            bits.append(f"   Status: {d.get('status')}")
        return "\n".join(bits)

    text_lines = "\n\n".join(_lead_line(d) for d in due)
    subject = f"⏰ BizDev follow-ups due: {len(due)} lead{'s' if len(due) != 1 else ''}"
    text_body = (
        "These leads haven't been contacted in 2+ days and are due for a "
        "follow-up:\n\n"
        f"{text_lines}\n\n"
        "Open the Leads pipeline in the MedPharma Hub, follow up, then click "
        "\"Mark followed up\" on each lead to reset its 2-day clock.\n"
    )
    rows_html = "".join(
        f"<tr><td style='padding:6px 10px;border-bottom:1px solid #e2e8f0'>"
        f"<b>{(d.get('practice_name') or d.get('contact_name') or ('Lead #' + str(d.get('id'))))}</b>"
        f"<br><span style='color:#64748b;font-size:12px'>"
        f"{', '.join(x for x in (d.get('contact_name') or '', d.get('contact_phone') or '', d.get('contact_email') or '') if x)}</span></td>"
        f"<td style='padding:6px 10px;border-bottom:1px solid #e2e8f0;text-align:right;color:#b91c1c;font-weight:600'>"
        f"{d.get('days_since_contact')} days</td></tr>"
        for d in due
    )
    html_body = (
        "<div style='font-family:system-ui,Segoe UI,Arial,sans-serif;max-width:600px;margin:0 auto;color:#0f172a'>"
        "<h2 style='color:#1d4ed8;margin:0 0 6px'>⏰ Follow-ups due</h2>"
        f"<p>These {len(due)} lead(s) haven't been contacted in 2+ days:</p>"
        "<table style='width:100%;border-collapse:collapse;font-size:14px'>"
        f"{rows_html}</table>"
        "<p style='font-size:13px;color:#475569;margin-top:16px'>Open the Leads "
        "pipeline, follow up, then click <b>Mark followed up</b> to reset the "
        "2-day clock.</p></div>"
    )

    emailed = 0
    for u in recipients:
        addr = (u.get("email") or "").strip()
        if not addr or "@" not in addr:
            continue
        try:
            sent, _via = _send_email_to(addr, subject, text_body, html_body)
            if sent:
                emailed += 1
        except Exception as e:
            log.error(f"bizdev follow-up email to {addr} failed: {e}")

    log.info(f"BizDev follow-up reminders: due={len(due)} notified={notified} emailed={emailed}")
    return {"ok": True, "due": len(due), "emailed": emailed, "notified": notified}


def send_bizdev_weekly_report(week_start: str = None) -> dict:
    """Email Victor's weekly Business-Development report to the team
    (Lexi + Eric, or whoever EOD_REPORT_EMAIL is set to). Runs Monday 8 AM EST
    and is also exposed for on-demand sending.

    Business Development is about the *type* of clients in the pipeline
    (RCM / Payor / Workflow / Compliance / Combination) — NOT dollar value.
    No monetary figures appear anywhere in this report.
    """
    from app.client_db import get_leads_weekly_report

    rep = get_leads_weekly_report(week_start)
    cats = rep.get("categories", {}) or {}
    rng = f"{rep.get('week_start','')} – {rep.get('week_end','')}"

    # Public MedPharma logo (same asset shown on the hub login screen).
    LOGO_URL = "https://medpharmasc.com/wp-content/uploads/2024/11/IMG_2392.png"

    cat_lines = "\n".join([
        f"  • RCM: {cats.get('rcm', 0)}",
        f"  • Payor: {cats.get('payor', 0)}",
        f"  • Workflow: {cats.get('workflow', 0)}",
        f"  • Compliance: {cats.get('compliance', 0)}",
        f"  • Combination (2+ services): {cats.get('combination', 0)}",
        f"  • Open total: {cats.get('open_total', 0)}",
        f"  • Closed: {cats.get('closed', 0)}",
    ])
    text_body = (
        f"MedPharma — Business Development Weekly Report\n"
        f"Week: {rng}\n\n"
        f"New leads this week: {rep.get('new_this_week', 0)}\n"
        f"Closed this week: {rep.get('closed_this_week', 0)}\n"
        f"Open pipeline (leads): {cats.get('open_total', 0)}\n\n"
        f"Pipeline by client type:\n{cat_lines}\n"
    )

    def _esc(s):
        return _esc_html(s) if s is not None else ""

    def _types(r):
        lines = r.get("service_lines") or []
        if len(lines) >= 2:
            return "Combination (" + ", ".join(lines) + ")"
        return ", ".join(lines) or "—"

    rows_html = "".join(
        f"<tr><td style='padding:6px 10px;border-bottom:1px solid #e2e8f0'>"
        f"<b>{_esc(r.get('practice_name') or r.get('contact_name') or ('Lead #' + str(r.get('id'))))}</b></td>"
        f"<td style='padding:6px 10px;border-bottom:1px solid #e2e8f0'>{_esc(_types(r))}</td>"
        f"<td style='padding:6px 10px;border-bottom:1px solid #e2e8f0'>{_esc(r.get('status') or '')}</td>"
        f"<td style='padding:6px 10px;border-bottom:1px solid #e2e8f0'>{_esc(r.get('owner') or '')}</td></tr>"
        for r in (rep.get("rows") or [])[:40]
    ) or "<tr><td colspan='4' style='padding:10px;color:#64748b'>No lead activity this week.</td></tr>"

    def _chip(label, val, color):
        return (
            f"<td style='padding:0 6px'><div style='background:{color}15;border:1px solid {color}40;"
            f"border-radius:10px;padding:10px 6px;text-align:center'>"
            f"<div style='font-size:22px;font-weight:800;color:{color}'>{val}</div>"
            f"<div style='font-size:11px;color:#475569;text-transform:uppercase;letter-spacing:.4px'>{label}</div>"
            f"</div></td>"
        )

    html_body = (
        "<div style='font-family:system-ui,Segoe UI,Arial,sans-serif;max-width:660px;margin:0 auto;color:#0f172a'>"
        # ── Branded header with the actual MedPharma logo ──
        "<div style='background:#0b2233;border-radius:12px 12px 0 0;padding:22px 24px;text-align:center'>"
        f"<img src='{LOGO_URL}' alt='MedPharma' style='max-width:280px;width:80%;height:auto;display:block;margin:0 auto'>"
        "</div>"
        "<div style='border:1px solid #e2e8f0;border-top:0;border-radius:0 0 12px 12px;padding:24px'>"
        "<h2 style='color:#1d4ed8;margin:0 0 4px'>Business Development — Weekly Report</h2>"
        f"<p style='color:#64748b;margin:0 0 18px'>{rng} &nbsp;·&nbsp; Pipeline by client type</p>"
        # ── Top-line counts (no dollars) ──
        "<table style='width:100%;border-collapse:separate;border-spacing:0;margin-bottom:8px'><tr>"
        f"{_chip('New This Week', rep.get('new_this_week', 0), '#2563eb')}"
        f"{_chip('Open Pipeline', cats.get('open_total', 0), '#7c3aed')}"
        f"{_chip('Closed This Week', rep.get('closed_this_week', 0), '#0891b2')}"
        "</tr></table>"
        # ── Breakdown by client type ──
        "<h3 style='font-size:15px;margin:18px 0 8px'>Open pipeline by client type</h3>"
        "<table style='width:100%;border-collapse:separate;border-spacing:0;margin-bottom:8px'><tr>"
        f"{_chip('RCM', cats.get('rcm', 0), '#2563eb')}"
        f"{_chip('Payor', cats.get('payor', 0), '#059669')}"
        f"{_chip('Workflow', cats.get('workflow', 0), '#d97706')}"
        f"{_chip('Compliance', cats.get('compliance', 0), '#dc2626')}"
        f"{_chip('Combination', cats.get('combination', 0), '#7c3aed')}"
        "</tr></table>"
        # ── Lead activity table (type-focused, no value) ──
        "<h3 style='font-size:15px;margin:18px 0 8px'>Lead activity this week</h3>"
        "<table style='width:100%;border-collapse:collapse;font-size:13px'>"
        "<tr style='text-align:left;color:#64748b'><th style='padding:6px 10px'>Practice</th><th style='padding:6px 10px'>Client Type</th><th style='padding:6px 10px'>Status</th><th style='padding:6px 10px'>Owner</th></tr>"
        f"{rows_html}</table>"
        "<p style='color:#94a3b8;font-size:12px;margin-top:18px'>MedPharma © 2026 · medpharmasc.com</p>"
        "</div></div>"
    )

    recipients = _eod_recipients()
    subject = f"MedPharma · BizDev Weekly Report — {rng}"
    sent = []
    failed = []
    for addr in recipients:
        addr = (addr or "").strip()
        if not addr or "@" not in addr:
            continue
        try:
            ok, via = _send_email_to(addr, subject, text_body, html_body)
            (sent if ok else failed).append({"email": addr, "via": via})
        except Exception as e:
            failed.append({"email": addr, "via": f"error: {e}"})

    log.info(f"BizDev weekly report: sent={len(sent)} failed={len(failed)} recipients={recipients}")
    return {"ok": True, "recipients": recipients, "sent": sent, "failed": failed,
            "week": rng}


def send_chat_unread_reminders(min_age_minutes: int = 120):
    """Email a one-time nudge to anyone who was @mentioned in a chat message
    they STILL haven't read after ``min_age_minutes`` (default 2 hours).

    This replaces the old "email on every message" behaviour. Read messages
    never trigger a reminder, and each mention is reminded at most once — so
    the inbox stays quiet unless someone is genuinely being waited on.
    PHI-safe: the message body is never included, only "open the room to view".
    """
    try:
        from app.client_db import (list_unread_mention_reminders,
                                    mark_chat_reminder_sent)
    except Exception:
        log.exception("chat reminder: db helpers unavailable")
        return {"ok": False, "sent": 0}

    try:
        pending = list_unread_mention_reminders(min_age_minutes=min_age_minutes)
    except Exception:
        log.exception("chat reminder: failed to list pending reminders")
        return {"ok": False, "sent": 0}

    if not pending:
        return {"ok": True, "sent": 0}

    try:
        from app.config import HUB_BASE_URL as _hub_base  # type: ignore
        hub_base = (_hub_base or "").strip().rstrip("/")
    except Exception:
        hub_base = ""

    sent = 0
    for item in pending:
        addr = (item.get("email") or "").strip()
        if not addr or "@" not in addr:
            continue
        room_name = item.get("room_name") or "a chat room"
        room_id = item.get("room_id")
        sender = item.get("sender_name") or "a teammate"
        who = (item.get("contact_name") or item.get("username") or "there").split()[0]
        deep_link = (f"{hub_base}/hub?chat={room_id}" if hub_base
                     else f"/hub?chat={room_id}")
        subject = f"💬 You were mentioned in '{room_name}' — still unread"
        text_body = (
            f"Hi {who},\n\n"
            f"{sender} mentioned you in the chat room '{room_name}' over two "
            f"hours ago and it's still unread.\n\n"
            f"Open the room to read and reply (the message itself isn't "
            f"included here for HIPAA compliance):\n{deep_link}\n"
        )
        html_body = (
            f"<div style='font-family:Segoe UI,Arial,sans-serif;color:#0f172a'>"
            f"<p>Hi {who},</p>"
            f"<p><b>{sender}</b> mentioned you in the chat room "
            f"<b>{room_name}</b> over two hours ago and it's still unread.</p>"
            f"<p style='color:#475569;font-size:13px'>The message body isn't "
            f"included in this email (HIPAA-protected content stays inside the "
            f"hub). Open the room to read and reply.</p>"
            f"<p style='margin:18px 0'>"
            f"<a href='{deep_link}' style='display:inline-block;padding:10px 22px;"
            f"background:#1d4ed8;color:#fff;text-decoration:none;border-radius:8px;"
            f"font-weight:600'>Open the chat room →</a></p>"
            f"</div>"
        )
        try:
            ok, _via = _send_email_to(addr, subject, text_body, html_body)
            if ok:
                sent += 1
            # Mark sent regardless so a hard-failing address doesn't loop every
            # cycle; in-app unread badge still nudges them inside the hub.
            mark_chat_reminder_sent(item["message_id"], item["user_id"])
        except Exception:
            log.exception("chat reminder: send failed for %s", addr)

    log.info(f"Chat unread-mention reminders: sent={sent} of {len(pending)} pending")
    return {"ok": True, "sent": sent, "pending": len(pending)}


def start_daily_scheduler():
    """
        Start APScheduler to fire:
            - send_production_reminders at 5:30 PM EST (for jessica & rcm)
            - send_daily_account_summary at 6:00 PM EST
    Safe to call multiple times — only starts once.
    """
    global _scheduler_started
    if _scheduler_started:
        return
    _scheduler_started = True

    try:
        from apscheduler.schedulers.background import BackgroundScheduler
        from apscheduler.triggers.cron import CronTrigger
        import pytz

        est = pytz.timezone("US/Eastern")
        scheduler = BackgroundScheduler(daemon=True)

        # 5:30 PM EST — Production reminders
        scheduler.add_job(
            send_production_reminders,
            CronTrigger(hour=17, minute=30, timezone=est),
            id="daily_production_reminders",
            name="5:30 PM EST Production Reminders",
            replace_existing=True,
        )

        # 6:00 PM EST — Account summary report
        scheduler.add_job(
            send_daily_account_summary,
            CronTrigger(hour=18, minute=0, timezone=est),
            id="daily_account_summary",
            name="6 PM EST Overall Account Summary",
            replace_existing=True,
        )

        # 6:30 PM EST — End-of-Day per-user / per-client / per-tab team report
        scheduler.add_job(
            send_eod_team_report,
            CronTrigger(hour=18, minute=30, timezone=est),
            id="daily_eod_team_report",
            name="6:30 PM EST EOD Team Report (lexi)",
            replace_existing=True,
        )

        # 6:35 PM EST — Per-client production reports (with Excel attachments)
        scheduler.add_job(
            send_all_client_daily_reports,
            CronTrigger(hour=18, minute=35, timezone=est),
            id="daily_client_reports",
            name="6:35 PM EST Per-Client Production Reports",
            replace_existing=True,
        )

        # 9:00 AM EST — BizDev follow-up reminders (every 2 days per lead)
        scheduler.add_job(
            send_bizdev_followup_reminders,
            CronTrigger(hour=9, minute=0, timezone=est),
            id="bizdev_followup_reminders",
            name="9 AM EST BizDev Follow-up Reminders",
            replace_existing=True,
        )

        # Monday 8:00 AM EST — BizDev weekly pipeline report to the team
        scheduler.add_job(
            send_bizdev_weekly_report,
            CronTrigger(day_of_week="mon", hour=8, minute=0, timezone=est),
            id="bizdev_weekly_report",
            name="Mon 8 AM EST BizDev Weekly Report",
            replace_existing=True,
        )

        # Every 30 min — nudge people @mentioned in chat who still haven't
        # read the message after 2 hours (replaces email-on-every-message).
        from apscheduler.triggers.interval import IntervalTrigger
        scheduler.add_job(
            send_chat_unread_reminders,
            IntervalTrigger(minutes=30, timezone=est),
            id="chat_unread_reminders",
            name="Every 30 min Chat Unread-Mention Reminders (2h)",
            replace_existing=True,
        )

        scheduler.start()
        log.info("Daily scheduler started — 5:00 national pull, 5:30 reminders, 6:00 summary, 6:30 EOD team, 6:35 client reports")
    except ImportError:
        # Fallback: use a simple threading timer that checks every 60 seconds
        log.warning("apscheduler not installed — falling back to threading-based scheduler")
        _start_thread_scheduler()
    except Exception as e:
        log.error(f"Failed to start scheduler: {e}")
        _start_thread_scheduler()


def _start_thread_scheduler():
    """Fallback scheduler using threading — checks every 60s for 5:30 and 6:00 PM EST."""
    import time as _time

    def _check_loop():
        last_reminder_date = None
        last_sent_date = None
        last_eod_date = None
        last_clients_date = None
        while True:
            try:
                # Get current time in US/Eastern
                try:
                    import pytz
                    est = pytz.timezone("US/Eastern")
                    now_est = datetime.now(est)
                except ImportError:
                    # No pytz — approximate EST as UTC-5
                    from datetime import timedelta, timezone
                    est_tz = timezone(timedelta(hours=-5))
                    now_est = datetime.now(est_tz)

                today = now_est.date()

                # 5:30 PM — Production reminders
                if now_est.hour == 17 and 30 <= now_est.minute < 35 and last_reminder_date != today:
                    last_reminder_date = today
                    log.info("Thread scheduler firing production reminders")
                    send_production_reminders()

                # 6:00 PM — Daily account summary
                if now_est.hour == 18 and now_est.minute < 5 and last_sent_date != today:
                    last_sent_date = today
                    log.info("Thread scheduler firing daily account summary")
                    send_daily_account_summary()

                # 6:30 PM — End-of-Day per-user / per-client / per-tab report
                if now_est.hour == 18 and 30 <= now_est.minute < 35 and last_eod_date != today:
                    last_eod_date = today
                    log.info("Thread scheduler firing EOD team report")
                    send_eod_team_report()

                # 6:35 PM — Per-client production reports with Excel
                if now_est.hour == 18 and 35 <= now_est.minute < 40 and last_clients_date != today:
                    last_clients_date = today
                    log.info("Thread scheduler firing per-client production reports")
                    send_all_client_daily_reports()

            except Exception as e:
                log.error(f"Thread scheduler error: {e}")
            _time.sleep(60)

    t = threading.Thread(target=_check_loop, daemon=True)
    t.start()
    log.info("Fallback thread scheduler started — 5:30 reminders + 6:00 summary + 6:30 EOD team + 6:35 client reports")
